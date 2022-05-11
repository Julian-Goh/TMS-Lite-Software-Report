from openpyxl import load_workbook
# from openpyxl_image_loader import SheetImageLoader #Editted the default SheetImageLoader class to fix an unintended behavior.

from collections import OrderedDict

from sheet_image_loader import SheetImageLoader

from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.colors import COLOR_INDEX

from fontTools.ttLib import TTFont
from fontTools.ttLib.tables._c_m_a_p import CmapSubtable
# print(COLOR_INDEX)

import openpyxl
import os
from os import path
import re
from datetime import datetime
import copy

from PIL import Image, ImageFont
import numpy as np
import copy

from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU

from misc_module.os_create_folder import create_save_folder

p2e = pixels_to_EMU
c2e = cm_to_EMU

#Default XL; column width:8 units = 1.65cm,  row height:14.4units = 0.5cm

custom_col_char_dict = {}
curr_char = 'A'
for i in range(0, 26):
    custom_col_char_dict[curr_char] = i + 1
    curr_char = chr(ord(curr_char) + 1)

del curr_char


def type_int(arg):
    if (type(arg)) == int or (isinstance(arg, np.integer) == True):
        return True
    else:
        return False

def getTextWidth(ttfont_obj, text, pointSize, glyph_set = None, font_cmap = None, units_per_em = None):
    if glyph_set is None:
        glyph_set = ttfont_obj.getGlyphSet()

    if font_cmap is None:
        cmap = ttfont_obj['cmap']
        font_cmap = cmap.getcmap(3,1).cmap

    if units_per_em is None:
        units_per_em = ttfont_obj['head'].unitsPerEm

    s = glyph_set
    t = font_cmap

    total = 0
    for c in text:
        if ord(c) in t and t[ord(c)] in s:
            total += s[t[ord(c)]].width
        else:
            total += s['.notdef'].width
    total = np.divide( np.multiply(total,float(pointSize)), units_per_em);
    return total

def xl_cm_to_emu(xl_cm, rel_pos = float(0)):
    result = None

    if type(rel_pos) == float:
        if float(0) <= rel_pos <= float(1):
            result = c2e(np.multiply(rel_pos, xl_cm))
        else:
            raise ValueError("rel_pos must be a float value ranging from 0 to 1")

    else:
        raise ValueError("rel_pos must be a float value ranging from 0 to 1")
    
    return result

def validate_unsign_int(value):
    if (type(value) == int or isinstance(value, np.integer) == True):
        if value >= 0:
            return True
        else:
            return False
    else:
        return False

def validate_unsign_float(value):
    if (type(value) == float or isinstance(value, np.floating) == True):
        if value >= 0:
            return True
        else:
            return False
    else:
        return False

def round_float(value, place = 2):
    # print(value)
    return float(f'{value:.{place}f}')

def xl_col_label_num(char):
    ### Excel 2007-2010 column range is: 'A' to 'XFD'#A = 1, #XFD = 16,384
    col_num = None
    try:
        if len(char) == 1:
            col_num = custom_col_char_dict[char]

        elif 1 < len(char) <= 3:
            col_num = 0
            for order, c in enumerate(char[::-1]):
                # print(c)
                col_num = col_num + np.multiply(custom_col_char_dict[c], np.power(26, order))

        else:
            raise KeyError

        return int(col_num)

    except KeyError:
        # print('KeyError')
        col_num = None
        return col_num


def xl_col_label_char(num):
    ### Excel 2007-2010 column range is: 'A' to 'XFD'#A = 1, #XFD = 16,384
    col_char = None
    if (type(num) == int or isinstance(num, np.integer) == True):
        if num > 0:
            char_list = []
            while num > 0:
                num, remainder = divmod(num, 26)
                # check for exact division and borrow if needed
                if remainder == 0:
                    remainder = 26
                    num -= 1
                char_list.append(chr(remainder+64))

            col_char = ''.join(reversed(char_list))

        else:
            raise ValueError("'num' argument must be a non-zero positive int-type data")

    else:
        raise ValueError("'num' argument must be a non-zero positive int-type data")

    return col_char


def compute_col_separator(col_separator): #compute border separator
    result_list = []
    if type(col_separator) == list:
        for element in col_separator:
            if type(element) == str:
                result = xl_col_label_num(element)
                if result is not None and type(result) == int:
                    result_list.append(result)

            elif type(element) == int and (1 <= element <= 16834):
                result_list.append(element)
    else:
        raise AttributeError("'col_separator' must be a list containing column characters A - XFD or integers from 1 - 16834")

    return result_list

def compute_row_separator(row_separator): #compute border separator
    result_list = []
    if type(row_separator) == list:
        for element in row_separator:
            if type(element) == int and (1 <= element <= 1048576):
                result_list.append(element)
    else:
        raise AttributeError("'row_separator' must be a list containing row numbers from 1 - 1048576")

    return result_list

def get_maximum_rows(ws):
    rows = 0
    iter_rows_obj = ws.iter_rows(min_row=1, max_row=ws.max_row
        , min_col=1, max_col=ws.max_column, values_only=False)

    # for i, row in enumerate(iter_rows_obj):
    for row in iter_rows_obj:
        # print(row)
        # print(i, all(col.value is None for col in row))
        if not all(col.value is None for col in row):
            rows += 1

    return rows

def xl_unmerge_cell(ws, ref_col_start, ref_row_start, ref_col_end = None, ref_row_end = None
    , del_value = False
    , border_type = None, border_style = 'medium'):
    
    unmerge_init = True
    unmerge_bool = True
    err_flag_arr = np.zeros((4,), dtype = bool)
    ref_col_start_num = None
    ref_col_end_num = None

    if False == type_int(ref_col_start):
        if type(ref_col_start) == str:
            ref_col_start_num = xl_col_label_num(ref_col_start)

            if type_int(ref_col_start_num) == False:
                err_flag_arr[0] = True
                raise ValueError("ref_col_start must be a string corresponding to Excel column characters from 'A' to 'XFD'")

            elif type_int(ref_col_start_num) == True:
                if 1 <= ref_col_start_num <= 16384:
                    pass
                else:
                    err_flag_arr[0] = True
                    raise ValueError("ref_col_start must be an integer from 1 to 16384")

        else:
            err_flag_arr[0] = True
            raise ValueError("ref_col_start must be an int type or str type input")

    elif True == type_int(ref_col_start):
        if 1 <= ref_col_start <= 16384:
            ref_col_start_num = ref_col_start

        else:
            err_flag_arr[0] = True
            raise ValueError("ref_col_start must be an integer from 1 to 16384")


    if False == type_int(ref_col_end):
        if type(ref_col_end) == str:
            ref_col_end_num = xl_col_label_num(ref_col_end)
            if ref_col_end_num is None:
                err_flag_arr[2] = True
                raise ValueError("ref_col_end must be a string corresponding to Excel column characters from 'A' to 'XFD'")

        elif ref_col_end is None and err_flag_arr[0] == False:
            ref_col_end_num = ref_col_start_num

        else:
            err_flag_arr[2] = True
            raise ValueError("ref_col_end must be an int type or str type input")

    elif True == type_int(ref_col_end):
        if 1 <= ref_col_end <= 16384:
            ref_col_end_num = ref_col_end

        else:
            err_flag_arr[2] = True
            raise ValueError("ref_col_end must be an integer from 1 to 16384")


    ############################################################################################
    if False == type_int(ref_row_start):
        err_flag_arr[1] = True
        raise ValueError("ref_row_start must be an int type input from 1 to 1048576")

    elif True == type_int(ref_row_start):
        if 1 <= ref_row_start <= 1048576:
            pass
        else:
            err_flag_arr[1] = True
            raise ValueError("ref_row_start must be an int type input from 1 to 1048576")

    if False == type_int(ref_row_end):
        if ref_row_end is None and err_flag_arr[1] == False:
            ref_row_end = ref_row_start
        else:
            err_flag_arr[3] = True
            raise ValueError("ref_row_end must be an int type input from 1 to 1048576")

    elif True == type_int(ref_row_end):
        if 1 <= ref_row_end <= 1048576:
            pass
        else:
            err_flag_arr[3] = True
            raise ValueError("ref_row_end must be an int type input from 1 to 1048576")


    # print(ref_col_start_num)
    # print(ref_row_start)
    # print(ref_col_end_num)
    # print(ref_row_end)
    '''To remove merge cells the supplied column and row inputs must match'''
    if np.any(err_flag_arr) == False:
        if ref_col_start_num > ref_col_end_num:
            unmerge_init = False
            raise ValueError("'ref_col_start' must be smaller/equal in magnitude compare to 'ref_col_end'.")

        if ref_row_start > ref_row_end:
            unmerge_init = False
            raise ValueError("'ref_row_start' must be smaller/equal in magnitude compare to 'ref_row_end'.")

        else:
            unmerge_init = True

    if unmerge_init == True:
        # print(ws.merged_cells.ranges)
        ws_merge_obj = copy.copy(ws.merged_cells.ranges)
        # print(ws_merge_obj)
        for cell_group in ws_merge_obj:
            start_cell = str(cell_group).split(':')[0]
            end_cell = str(cell_group).split(':')[-1]
            # print(start_cell, end_cell)
            start_col_num = xl_col_label_num(re.findall('(\\D+)', start_cell)[0])
            start_row_num = int(re.findall('(\\d+)', start_cell)[0])

            end_col_num = xl_col_label_num(re.findall('(\\D+)', end_cell)[0])
            end_row_num = int(re.findall('(\\d+)', end_cell)[0])

            # print(str(cell_group))
            if start_col_num >= ref_col_start_num and end_col_num <= ref_col_end_num:
                if start_row_num >= ref_row_start and end_row_num <= ref_row_end:
                    ws.unmerge_cells(str(cell_group))
                    # print("Unmerging: ", str(cell_group))
                    set_outer_border(ws, border_type = border_type, border_style = border_style
                        , min_col = start_col_num, max_col = end_col_num, min_row = start_row_num, max_row = end_row_num)

                    if del_value == True:
                        ws[start_cell].value = None

        del ws_merge_obj

def set_outer_border(ws, border_type = 'default', border_style = 'medium'
    , col_separator = [], row_separator = [], print_bool = False, **iter_rows_kwargs):
    """border_type = 'default', 'top', 'bottom', 'left-right', None """
    '''
    BORDER_NONE = None
    BORDER_DASHDOT = 'dashDot'
    BORDER_DASHDOTDOT = 'dashDotDot'
    BORDER_DASHED = 'dashed'
    BORDER_DOTTED = 'dotted'
    BORDER_DOUBLE = 'double'
    BORDER_HAIR = 'hair'
    BORDER_MEDIUM = 'medium'
    BORDER_MEDIUMDASHDOT = 'mediumDashDot'
    BORDER_MEDIUMDASHDOTDOT = 'mediumDashDotDot'
    BORDER_MEDIUMDASHED = 'mediumDashed'
    BORDER_SLANTDASHDOT = 'slantDashDot'
    BORDER_THICK = 'thick'
    BORDER_THIN = 'thin'
    '''
    #load parameters from iter_rows_kwargs : START
    try:
        min_row = iter_rows_kwargs['min_row']
        if type(min_row) == int or (isinstance(min_row, np.integer) == True):
            if 1 <= min_row <= 1048576:
                pass
            else:
                raise ValueError("'min_row' must be an integer-type and 1 <= min_row <= 1048576")
        else:
            raise ValueError("'min_row' must be an integer-type and 1 <= min_row <= 1048576")

    except KeyError:
        min_row = None
        raise Exception("Requires 'min_row' argument!")

    try:
        min_col = iter_rows_kwargs['min_col']
        if type(min_col) == int or (isinstance(min_col, np.integer) == True):
            if 1 <= min_col <= 16384:
                pass
            else:
                raise ValueError("If 'min_col' is an integer-type, then 1 <= min_col <= 16834."
                    + "\nElse if 'min_col' is a string-type, then 'A' <= min_col <= 'XFD'.")
        elif type(min_col) == str:
            min_col = xl_col_label_num(min_col)
            if min_col is None:
                raise ValueError("If 'min_col' is an integer-type, then 1 <= min_col <= 16834."
                + "\nElse if 'min_col' is a string-type, then 'A' <= min_col <= 'XFD'.")
        else:
            raise ValueError("If 'min_col' is an integer-type, then 1 <= min_col <= 16834."
                 + "\nElse if 'min_col' is a string-type, then 'A' <= min_col <= 'XFD'.")

    except KeyError:
        min_col = None
        raise Exception("Requires 'min_col' argument!")

    try:
        max_row = iter_rows_kwargs['max_row']
        if type(max_row) == int or (isinstance(max_row, np.integer) == True):
            if 1 <= max_row <= 1048576 and (min_row < max_row):
                pass
            else:
                if min_row is not None and (type(min_row) == int or (isinstance(min_row, np.integer) == True)):
                    max_row = min_row
                else:    
                    max_row = None
        else:
            if min_row is not None and (type(min_row) == int or (isinstance(min_row, np.integer) == True)):
                max_row = min_row
            else:    
                max_row = None

    except KeyError:
        if min_row is not None and (type(min_row) == int or (isinstance(min_row, np.integer) == True)):
            max_row = min_row
        else:    
            max_row = None

    try:
        max_col = iter_rows_kwargs['max_col']
        if type(max_col) == int or (isinstance(max_col, np.integer) == True):
            if 1 <= max_col <= 16384 and (min_col < max_col):
                pass
            else:
                if min_col is not None and (type(min_col) == int or (isinstance(min_col, np.integer) == True)):
                    max_col = min_col
                else:    
                    max_col = None

        elif type(max_col) == str:
            max_col = xl_col_label_num(max_col)
            if max_col is None:
                if min_col is not None and (type(min_col) == int or (isinstance(min_col, np.integer) == True)):
                    max_col = min_col
                else:    
                    max_col = None

            elif type(max_col) == int:
                if 1 <= max_col <= 16384 and (min_col < max_col):
                    pass
                else:
                    if min_col is not None and (type(min_col) == int or (isinstance(min_col, np.integer) == True)):
                        max_col = min_col
                    else:    
                        max_col = None
        else:
            if min_col is not None and (type(min_col) == int or (isinstance(min_col, np.integer) == True)):
                max_col = min_col
            else:    
                max_col = None

    except KeyError:
        if min_col is not None and (type(min_col) == int or (isinstance(min_col, np.integer) == True)):
            max_col = min_col
        else:    
            max_col = None

    try:
        values_only = iter_rows_kwargs['values_only']
        if type(values_only) == bool:
            pass

        else:
            values_only = False
    except KeyError:
        values_only = False
    
    #load parameters from iter_rows_kwargs : END
    i = 0
    iter_rows_obj = ws.iter_rows(min_row=min_row, max_row=max_row
        , min_col=min_col, max_col=max_col, values_only=values_only)

    # print(col_separator)
    col_separator = compute_col_separator(col_separator)
    # print(col_separator)
    # row_separator = compute_row_separator(row_separator)
    # print('col_separator: ', col_separator)
    # print('min_row, max_row: ', min_row, max_row)
    row_range = max_row - min_row
    col_range = max_col - min_col
    # print(row_range, col_range)

    for i, row in enumerate(iter_rows_obj):
        if border_type is not None and isinstance(border_type, str) == True:
            if i == 0: ##The 1st Row from the Top
                j = 0
                for j, c in enumerate(row):
                    if row_range > 0:
                        if border_type == 'default' or border_type == 'top':
                            if j == 0:
                                if col_range > 0:
                                    c.border = Border(top = Side(border_style=border_style, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000'))
                                elif col_range == 0:
                                    c.border = Border(top = Side(border_style=border_style, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))

                            elif j == (len(row)-1):
                                c.border = Border(top = Side(border_style=border_style, color='FF000000')
                                    , right = Side(border_style=border_style, color='FF000000'))
                            else:
                                if (j + 1) in col_separator:
                                    c.border = Border(top = Side(border_style=border_style, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))
                                else:
                                    c.border = Border(top = Side(border_style=border_style, color='FF000000'))
                            
                        elif border_type == 'bottom' or border_type == 'left-right':
                            if j == 0:
                                if col_range > 0:
                                    c.border = Border(top = Side(border_style=None, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000'))
                                elif col_range == 0:
                                    c.border = Border(top = Side(border_style=None, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))

                            elif j == (len(row)-1):
                                c.border = Border(top = Side(border_style=None, color='FF000000')
                                    , right = Side(border_style=border_style, color='FF000000'))
                            else:
                                if (j + 1) in col_separator:
                                    c.border = Border(top = Side(border_style=None, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))
                                else:
                                    c.border = Border(top = Side(border_style=None, color='FF000000'))


                    elif row_range == 0:
                        if border_type == 'default':
                            if j == 0:
                                if col_range > 0:
                                    c.border = Border(top = Side(border_style=border_style, color='FF000000')
                                        , bottom = Side(border_style=border_style, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000'))

                                elif col_range == 0:
                                    c.border = Border(top = Side(border_style=border_style, color='FF000000')
                                        , bottom = Side(border_style=border_style, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))

                            elif j == (len(row)-1):
                                c.border = Border(top = Side(border_style=border_style, color='FF000000')
                                    , bottom = Side(border_style=border_style, color='FF000000')
                                    , right = Side(border_style=border_style, color='FF000000'))
                            else:
                                if (j + 1) in col_separator:
                                    c.border = Border(top = Side(border_style=border_style, color='FF000000')
                                        , bottom = Side(border_style=border_style, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))
                                else:
                                    c.border = Border(top = Side(border_style=border_style, color='FF000000')
                                        , bottom = Side(border_style=border_style, color='FF000000'))

                        elif border_type == 'top':
                            if j == 0:
                                if col_range > 0:
                                    c.border = Border(top = Side(border_style=border_style, color='FF000000')
                                        , bottom = Side(border_style=None, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000'))
                                elif col_range == 0:
                                    c.border = Border(top = Side(border_style=border_style, color='FF000000')
                                        , bottom = Side(border_style=None, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))

                            elif j == (len(row)-1):
                                c.border = Border(top = Side(border_style=border_style, color='FF000000')
                                    , bottom = Side(border_style=None, color='FF000000')
                                    , right = Side(border_style=border_style, color='FF000000'))
                            else:
                                if (j + 1) in col_separator:
                                    c.border = Border(top = Side(border_style=border_style, color='FF000000')
                                        , bottom = Side(border_style=None, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))
                                else:
                                    c.border = Border(top = Side(border_style=border_style, color='FF000000')
                                        , bottom = Side(border_style=None, color='FF000000'))

                        elif border_type == 'bottom':
                            if j == 0:
                                if col_range > 0:
                                    c.border = Border(top = Side(border_style=None, color='FF000000')
                                        , bottom = Side(border_style=border_style, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000'))
                                elif col_range == 0:
                                    c.border = Border(top = Side(border_style=None, color='FF000000')
                                        , bottom = Side(border_style=border_style, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))

                            elif j == (len(row)-1):
                                c.border = Border(top = Side(border_style=None, color='FF000000')
                                    , bottom = Side(border_style=border_style, color='FF000000')
                                    , right = Side(border_style=border_style, color='FF000000'))
                            else:
                                if (j + 1) in col_separator:
                                    c.border = Border(top = Side(border_style=None, color='FF000000')
                                        , bottom = Side(border_style=border_style, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))
                                else:
                                    c.border = Border(top = Side(border_style=None, color='FF000000')
                                        , bottom = Side(border_style=border_style, color='FF000000'))

                        elif border_type == 'left-right':
                            if j == 0:
                                if col_range > 0:
                                    c.border = Border(top = Side(border_style=None, color='FF000000')
                                        , bottom = Side(border_style=None, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000'))
                                elif col_range == 0:
                                    c.border = Border(top = Side(border_style=None, color='FF000000')
                                        , bottom = Side(border_style=None, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))

                            elif j == (len(row)-1):
                                c.border = Border(top = Side(border_style=None, color='FF000000')
                                    , bottom = Side(border_style=None, color='FF000000')
                                    , right = Side(border_style=border_style, color='FF000000'))
                            else:
                                if (j + 1) in col_separator:
                                    c.border = Border(top = Side(border_style=None, color='FF000000')
                                        , bottom = Side(border_style=None, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))
                                else:
                                    c.border = Border(top = Side(border_style=None, color='FF000000')
                                        , bottom = Side(border_style=None, color='FF000000'))

                    del c
                del j

            elif i > 0: ##The Remaining Rows
                if i < (max_row - min_row):
                    if col_range > 0:
                        row[0].border = Border(left = Side(border_style=border_style, color='FF000000'))
                        row[-1].border = Border(right = Side(border_style=border_style, color='FF000000'))
                        for element in col_separator:
                            row[element - 1].border = Border(right = Side(border_style=border_style, color='FF000000'))
                            del element

                    elif col_range == 0:
                        row[0].border = Border(left = Side(border_style=border_style, color='FF000000')
                            , right = Side(border_style=border_style, color='FF000000'))

                elif i == (max_row - min_row):
                    for j, c in enumerate(row):
                        if border_type == 'default' or border_type == 'bottom':
                            if j == 0:
                                if col_range > 0:
                                    c.border = Border(bottom = Side(border_style=border_style, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000'))
                                elif col_range == 0:
                                    c.border = Border(bottom = Side(border_style=border_style, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))

                            elif j == (len(row)-1):
                                c.border = Border(bottom = Side(border_style=border_style, color='FF000000')
                                    , right = Side(border_style=border_style, color='FF000000'))
                            else:
                                if (j + 1) in col_separator:
                                    c.border = Border(bottom = Side(border_style=border_style, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))
                                else:
                                    c.border = Border(bottom = Side(border_style=border_style, color='FF000000'))

                        elif border_type == 'top' or border_type == 'left-right':
                            if j == 0:
                                if col_range > 0:
                                    c.border = Border(bottom = Side(border_style=None, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000'))
                                elif col_range == 0:
                                    c.border = Border(bottom = Side(border_style=None, color='FF000000')
                                        , left = Side(border_style=border_style, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))

                            elif j == (len(row)-1):
                                c.border = Border(bottom = Side(border_style=None, color='FF000000')
                                    , right = Side(border_style=border_style, color='FF000000'))
                            else:
                                if (j + 1) in col_separator:
                                    c.border = Border(bottom = Side(border_style=None, color='FF000000')
                                        , right = Side(border_style=border_style, color='FF000000'))
                                else:
                                    c.border = Border(bottom = Side(border_style=None, color='FF000000'))

                        del c

                    del j

                    
        elif border_type is None: # Clear all borders from cell(s)
            for c in row:
                c.border = Border(top = Side(border_style=None)
                            , bottom = Side(border_style=None)
                            , left = Side(border_style=None)
                            , right = Side(border_style=None))
            del c
        del row

    del i

    del col_separator
    del row_separator

def xl_read_worksheet(ws, **iter_rows_kwargs):
    # print("Reading Data...")
    try:
        min_row = iter_rows_kwargs['min_row']
        if type(min_row) == int:
            if 1 <= min_row <= 1048576:
                pass
            else:
                raise ValueError("'min_row' must be an integer-type and 1 <= min_row <= 1048576")
        else:
            raise ValueError("'min_row' must be an integer-type and 1 <= min_row <= 1048576")

    except KeyError:
        min_row = None
        raise Exception("Requires 'min_row' argument!")

    try:
        min_col = iter_rows_kwargs['min_col']
        if type(min_col) == int:
            if 1 <= min_col <= 16384:
                pass
            else:
                raise ValueError("If 'min_col' is an integer-type, then 1 <= min_col <= 16834."
                    + "\nElse if 'min_col' is a string-type, then 'A' <= min_col <= 'XFD'.")
        elif type(min_col) == str:
            min_col = xl_col_label_num(min_col)
            if min_col is None:
                raise ValueError("If 'min_col' is an integer-type, then 1 <= min_col <= 16834."
                + "\nElse if 'min_col' is a string-type, then 'A' <= min_col <= 'XFD'.")
        else:
            raise ValueError("If 'min_col' is an integer-type, then 1 <= min_col <= 16834."
                 + "\nElse if 'min_col' is a string-type, then 'A' <= min_col <= 'XFD'.")

    except KeyError:
        min_col = None
        raise Exception("Requires 'min_col' argument!")

    try:
        max_row = iter_rows_kwargs['max_row']
        if type(max_row) == int:
            if 1 <= max_row <= 1048576 and (min_row < max_row):
                pass
            else:
                if min_row is not None and type(min_row) == int:
                    max_row = min_row
                else:    
                    max_row = None
        else:
            if min_row is not None and type(min_row) == int:
                max_row = min_row
            else:    
                max_row = None

    except KeyError:
        if min_row is not None and type(min_row) == int:
            max_row = min_row
        else:    
            max_row = None

    try:
        max_col = iter_rows_kwargs['max_col']
        if type(max_col) == int:
            if 1 <= max_col <= 16384 and (min_col < max_col):
                pass
            else:
                if min_col is not None and type(min_col) == int:
                    max_col = min_col
                else:    
                    max_col = None

        elif type(max_col) == str:
            max_col = xl_col_label_num(max_col)
            if max_col is None:
                if min_col is not None and type(min_col) == int:
                    max_col = min_col
                else:    
                    max_col = None

            elif type(max_col) == int:
                if 1 <= max_col <= 16384 and (min_col < max_col):
                    pass
                else:
                    if min_col is not None and type(min_col) == int:
                        max_col = min_col
                    else:    
                        max_col = None
        else:
            if min_col is not None and type(min_col) == int:
                max_col = min_col
            else:    
                max_col = None

    except KeyError:
        if min_col is not None and type(min_col) == int:
            max_col = min_col
        else:    
            max_col = None

    try:
        values_only = iter_rows_kwargs['values_only']
        if type(values_only) == bool:
            pass

        else:
            values_only = False
    except KeyError:
        values_only = False

    iter_rows_obj = ws.iter_rows(min_row=min_row, max_row=max_row
        , min_col=min_col, max_col=max_col, values_only=values_only)
    # print(iter_rows_obj)
    for rows in iter_rows_obj:
        for cell in rows:
            yield cell

class XL_WorkBook():
    def __init__(self):
        self.ref_font = TTFont(os.getcwd() + "\\Font\\" + "arial.ttf")
        cmap = self.ref_font['cmap']
        self.font_cmap = cmap.getcmap(3,1).cmap
        self.glyph_set = self.ref_font.getGlyphSet()
        self.units_per_em = self.ref_font['head'].unitsPerEm

        self.font_size = 16
        self.unit_char_width = getTextWidth(self.ref_font, '0', self.font_size, glyph_set = self.glyph_set, font_cmap = self.font_cmap, units_per_em = self.units_per_em)
        # print(self.unit_char_width)

        self.width_font_factor = np.array([1.8, 1.67], dtype = np.float16) #This is only applicable for Arial font 16!! #1st value is used for 1st and last char, else use 2nd-value
        # print(self.width_font_factor)
        self.height_font_factor = np.divide(21, 15) #This is only applicable for Arial font 16!!

        self.__save_dir = os.path.join(os.environ['USERPROFILE'], "TMS_Saved_Reports")

        self.__height_cm_xlunit_ratio = np.divide(0.51, 14.4)
        self.__width_cm_xlunit_ratio = np.divide(1.65, 8)
        self.__xl_img_cm_tolerance = 0.15 #0.3 ### IMPORTANT! Tolerance value: 0.3cm, If Image covers a cell by 0.3cm, user can still view text(s) inside the cell.

        self.workbook = None
        self.worksheet = None

        self._init_col_char = 'A'
        self._init_row_num = 2

    def xl_get_column_width(self, worksheet, start_char = 'A', end_char = None):
        #### This function returns (a) the width, and (b) number of columns

        if start_char is not None and end_char is not None:
            start_col_num = xl_col_label_num(start_char)
            end_col_num = xl_col_label_num(end_char)

            if start_col_num is not None and end_col_num is not None:
                if start_col_num > end_col_num:
                    raise ValueError("end_char is at a lower alphabetical order compared to start_char!")

                elif start_col_num == end_col_num:
                    return worksheet.column_dimensions[start_char].width - 0.78, 1

                elif start_col_num < end_col_num:
                    column_width = 0
                    for i in range(start_col_num, end_col_num + 1):
                        curr_char = get_column_letter(i)
                        column_width = column_width + worksheet.column_dimensions[curr_char].width - 0.78
                    # print(column_width)
                    return column_width, (end_col_num - start_col_num + 1) 
            else:
                raise ValueError("start_char and end_char must be an alphabetical string 'from A to Z', respectively.")

        elif start_char is not None and end_char is None:
            start_col_num = xl_col_label_num(start_char)
            if start_col_num is not None:
                return worksheet.column_dimensions[start_char].width - 0.78, 1

            else:
                raise ValueError("start_char must be an alphabetical string 'from A to Z'")

        return None, None

    def xl_get_row_height(self, worksheet, start_num = 1, end_num = 1):
        error_bool = False
        if not (type(start_num) == int and type(end_num) == int):
            error_bool = True
            raise ValueError("end_num and start_num must be positive integer(s) and non-zero")

        elif (type(start_num) == int and type(end_num) == int):
            if start_num > 0 and end_num > 0:
                error_bool = False
                pass
            else:
                error_bool = True
                raise ValueError("end_num and start_num must be positive integer(s) and non-zero")

        if error_bool == False:
            if start_num > end_num:
                row_height = worksheet.row_dimensions[start_num].height
                if row_height is None:
                    return 14.4, 1
                else:
                    return row_height, 1

            elif start_num == end_num:
                row_height = worksheet.row_dimensions[start_num].height
                if row_height is None:
                    return 14.4, 1
                else:
                    return row_height, 1

            elif start_num < end_num:
                row_height = 0
                for row_num in range(start_num, end_num + 1):
                    cell_height = worksheet.row_dimensions[row_num].height
                    if cell_height is None:
                        cell_height = 14.4
                    # print(cell_height)
                    row_height = row_height + cell_height

                return row_height, (end_num - start_num + 1)

        return None, None

    def xl_set_column_width(self, worksheet, start_char = 'A', end_char = 'A', col_width = 8.78):

        start_col_num = xl_col_label_num(start_char)
        end_col_num = xl_col_label_num(end_char)

        if start_col_num > end_col_num:
            raise ValueError("end_char is at a lower alphabetical order compared to start_char!")

        elif start_col_num == end_col_num:
            worksheet.column_dimensions[start_char].width = col_width

        elif start_col_num < end_col_num:
            for i in range(start_col_num, end_col_num + 1):
                curr_char = get_column_letter(i)
                # print(curr_char)
                worksheet.column_dimensions[curr_char].width = col_width

    def xl_file_save(self, folder, workbook, new_file_name = None, file_format = '.xlsx', file_path = None):
        # print('new_file_name: ', new_file_name)
        # print('file_path: ', file_path)
        error_event = False

        if new_file_name is not None and file_path is None:
            index = 0
            loop = True
            dt_string = datetime.now().strftime("%d-%m-%Y_")
            while loop == True:
                save_path = folder + '\\'+ dt_string + new_file_name + '_' + str(index) + file_format
                if (path.exists(save_path)) == True:
                    index = index + 1
                elif (path.exists(save_path)) == False:
                    loop = False
            
            try:
                workbook.save(save_path)
                workbook.close()
            except Exception:
                workbook.close()
                error_event = True

            return save_path, error_event

        elif new_file_name is None and file_path is not None:
            save_path = file_path
            try:
                workbook.save(save_path)
                workbook.close()
            except Exception:
                workbook.close()
                error_event = True

            return save_path, error_event

        else:
            return None, error_event

    def xl_file_save_2(self, folder, workbook, new_file_name = None, file_format = '.xlsx', file_path = None):
        # print('new_file_name: ', new_file_name)
        # print('file_path: ', file_path)

        if new_file_name is not None and file_path is None:
            index = 0
            loop = True
            dt_string = datetime.now().strftime("%d-%m-%Y_")
            while loop == True:
                save_path = folder + '\\'+ dt_string + new_file_name + '_' + str(index) + file_format
                if (path.exists(save_path)) == True:
                    index = index + 1
                elif (path.exists(save_path)) == False:
                    loop = False

            workbook.save(save_path)
            return save_path

        elif new_file_name is None and file_path is not None:
            save_path = file_path

            workbook.save(save_path)
            return save_path

        else:
            return None

    def xl_new_workbook(self, sheet_name = None):
        self.workbook = Workbook()
        self.worksheet = self.workbook.worksheets[0]
        self.xl_select_worksheet(self.workbook.worksheets[0].title) 
        if type(sheet_name) == str:
            if sheet_name == self.worksheet.title:
                pass
            else:
                self.xl_rename_worksheet(sheet_name, override = True)
                self.xl_select_worksheet(sheet_name)

        return self.worksheet.title

        
    def xl_check_worksheet(self, sheet_name):
        if sheet_name in self.workbook.sheetnames:
            # curr_id_list = re.findall("\\(\\d\\)", sheet_name)
            # curr_id_list = re.findall("\\([0-9]\\)|\\([0-9][0-9]\\)", sheet_name)
            curr_id_list = re.findall("\\([0-9]\\)$|\\([0-9][0-9]\\)$", sheet_name)
            # print(curr_id_list)
            if len(curr_id_list) > 0:
                new_id = int(curr_id_list[-1].strip('()'))
                # print(new_id)
                # new_name = re.sub("\\(" + str(new_id) + "\\)", "", sheet_name).rstrip()
                # base_name = copy.copy(new_name)
                # print(new_name, sheet_name, curr_id_list[-1])
                new_name = copy.copy(sheet_name)
                # base_name = re.sub("\\(" + str(new_id) + "\\)", "", sheet_name).rstrip()
                base_name = re.sub("\\(" + str(new_id) + "\\)$", "", sheet_name).rstrip()
            else:
                new_id = 1
                new_name = copy.copy(sheet_name)
                base_name = copy.copy(sheet_name)
        else:
            new_id = 1
            new_name = copy.copy(sheet_name)
            base_name = copy.copy(sheet_name)

        while True:
            if new_name in self.workbook.sheetnames:
                new_name_id = '(' + str(new_id) + ')'
                new_name = base_name + new_name_id
                # print('Checking...: ', new_name, len(new_name))
                if len(new_name) > 31:
                    xtra_str = len(new_name) - 31
                    crop_name = base_name[:-xtra_str]
                    new_name = crop_name + new_name_id
                    print(crop_name, len(crop_name))
                new_id += 1

            else:
                break

        del base_name

        # print('Finished...: ', new_name)
        return new_name

    def xl_new_worksheet(self, sheet_name):
        sheet_name = self.xl_check_worksheet(sheet_name)
        self.workbook.create_sheet(sheet_name)
        self.worksheet = self.workbook.worksheets[-1]

        return sheet_name

    def xl_remove_worksheet(self, sheet_name):
        # self.workbook.remove_sheet(sheet_name)
        del self.workbook[sheet_name]

    def xl_rename_worksheet(self, sheet_name, override = False):
        # print(self.workbook.sheetnames)
        ret_err = False
        if self.worksheet.title == sheet_name:
            return ret_err, sheet_name
        else:
            if sheet_name in self.workbook.sheetnames:
                if override == False:
                    ret_err = True
                    return ret_err, sheet_name

                elif override == True:
                    sheet_name = self.xl_check_worksheet(sheet_name)
                    self.worksheet.title = sheet_name
                    ret_err = False
                    return ret_err, sheet_name

            else:
                self.worksheet.title = sheet_name
                ret_err = False
                return ret_err, sheet_name

    def xl_select_worksheet(self, sheet_name):
        if sheet_name in self.workbook.sheetnames:
            self.worksheet = self.workbook[sheet_name]

        else:
            raise Exception("The current Excel File does not have '{}' worksheet".format(sheet_name))

    def xl_save_workbook(self, file_path = None):
        
        report_folder = create_save_folder(folder_dir = self.__save_dir)
        # print('file_path: ', file_path)
        if file_path is None:
            save_file, error_event = self.xl_file_save(report_folder, self.workbook, 'Report', '.xlsx')
            # save_file = self.xl_file_save_2(report_folder, self.workbook, 'Report', '.xlsx')

        elif file_path is not None:
            save_file, error_event = self.xl_file_save(report_folder, self.workbook, file_format = '.xlsx', file_path = file_path)
            # save_file = self.xl_file_save_2(report_folder, self.workbook, file_format = '.xlsx', file_path = file_path)

        return save_file, error_event
        # return save_file

    def xl_open_workbook(self, file_path, sheet_name = None):
        self.workbook = load_workbook(file_path)
        if type(sheet_name) == str:
            self.xl_select_worksheet(sheet_name)
        else:
            self.worksheet = self.workbook.worksheets[-1]

    def xl_close_workbook(self, file_path = None):
        event_error = False

        if file_path is not None:
            try:
                self.workbook.save(file_path)
                self.workbook.close()
            except Exception:
                event_error = True
        else:
            try:
                self.workbook.close()
            except Exception:
                event_error = True

        return event_error

    def xl_update_workbook(self, file_path, sheet_name = None, new_ws_bool = False):
        self.workbook = load_workbook(file_path)

        if new_ws_bool == False:
            if type(sheet_name) == str:
                try:
                    self.worksheet = self.workbook[sheet_name]
                except Exception:
                    self.worksheet = self.workbook.worksheets[-1]
            else:
                self.worksheet = self.workbook.worksheets[-1]

        elif new_ws_bool == True:
            if type(sheet_name) == str:
                sheet_name = self.xl_new_worksheet(sheet_name)
            else:
                raise AttributeError("Please insert a 'sheet_name' into this function if you want to create a new worksheet."
                    + "\n'sheet_name' is a str-type.")

        # Have to Set this Default Width for Some of these Columns....
        self.xl_set_column_width(self.worksheet, 'B', 'F')
        self.xl_set_column_width(self.worksheet, 'G', 'L')
        self.xl_set_column_width(self.worksheet, 'Q', 'V')
        self.xl_set_column_width(self.worksheet, 'W', 'AB')
        self.xl_set_column_width(self.worksheet, 'AC', 'AH')

        self.xl_set_column_width(self.worksheet, 'AI', 'AN')

        self.xl_init_workbook(init_row_num = 2, init_col_char = 'A')

        return sheet_name

    def xl_image_clear_all(self):
        self.worksheet._images *= 0

    def xl_read_image(self, ws, sort_col_bool = False, sort_row_bool = False):
        return SheetImageLoader(ws, sort_col_bool, sort_row_bool)

    def xl_init_workbook(self, init_row_num = 1, init_col_char = 'A'):
        '''
            Only use xl_init_workbook() before inserting new data to Excel worksheet!!
        '''

        __ws_max_row = 39 ## Maximum row for TMS-Lite Lab Specification Report

        self.xl_set_column_width(self.worksheet, 'A', 'AN') #Initialize all the column width size! Otherwise xl_get_column_width will return weird column width when cell is empty

        self.worksheet.merge_cells('A2:F2')
        set_outer_border(ws = self.worksheet, min_row = 2, min_col = 'A', max_col='F')

        self.cell_insert_data(self.worksheet, 'Client Photo Sample', 'string', 'A2', auto_fit_col = False, auto_fit_row = False)

        self.worksheet.merge_cells('B3:F3')
        set_outer_border(ws = self.worksheet, min_row = 3, min_col = 'A')
        set_outer_border(ws = self.worksheet, min_row = 3, min_col = 'B', max_col='F')

        self.cell_insert_data(self.worksheet, 'Dimension(mm): ', 'string', 'A3'
            , auto_fit_override = (True, True)
            , alignment_dict = dict(wrapText=True, horizontal = 'left', vertical = 'center'))

        self.worksheet.merge_cells('B4:F4')
        set_outer_border(ws = self.worksheet, min_row = 4, min_col = 'A')
        set_outer_border(ws = self.worksheet, min_row = 4, min_col = 'B', max_col='F')

        self.cell_insert_data(self.worksheet, 'Background: ', 'string', 'A4'
            , auto_fit_override = (False, True)
            , alignment_dict = dict(wrapText=True, horizontal = 'left', vertical = 'center'))

        # self.worksheet.merge_cells('A4:F39')
        set_outer_border(ws = self.worksheet, min_row = 5, max_row = 39, min_col = 'A', max_col='F')
        #####################################################################################################

        self.worksheet.merge_cells('G2:L2')
        set_outer_border(ws = self.worksheet, min_row = 2, min_col = 'G', max_col='L')
        self.cell_insert_data(self.worksheet, 'Lighting Drawing', 'string', 'G2', auto_fit_col = False, auto_fit_row = False)

        self.worksheet.merge_cells('G3:L3')
        set_outer_border(ws = self.worksheet, min_row = 3, max_row = 39, min_col = 'G', max_col='L')
        self.cell_insert_data(self.worksheet, 'LIGHTING SERIES', 'string', 'G3', auto_fit_col = False, auto_fit_row = False)
        #####################################################################################################
        set_outer_border(ws = self.worksheet, min_row = 2, max_row = 39, min_col = 'M', max_col='P')
        self.worksheet.column_dimensions['M'].width = 25
        self.worksheet.column_dimensions['N'].width = 25
        self.worksheet.column_dimensions['O'].width = 25
        self.worksheet.column_dimensions['P'].width = 2
        self.worksheet.merge_cells('M2:P2')

        set_outer_border(ws = self.worksheet, min_row = 2, min_col = 'M', max_col = 'P')
        self.cell_insert_data(self.worksheet, 'Specification', 'string', 'M2', auto_fit_col = False, auto_fit_row = False)

        self.worksheet.merge_cells('M3:P3')
        set_outer_border(ws = self.worksheet, min_row = 3, min_col = 'M', max_col='P')

        self.cell_insert_data(self.worksheet, "Note rows with '*' must be filled.", 'string', 'M3', auto_fit_col = False, auto_fit_row = False
            , font_dict = dict(name='Arial',
            size=16,
            bold=False,
            italic=False, 
            vertAlign=None,
            underline='none',
            strike=False,
            color='FFFF0000')
            , alignment_dict = dict(wrapText=True, horizontal = 'left', vertical = 'center'))

        ###### Testing Criteria Table
        test_criteria_table = ['A = Lighting Dimension (mm):', 'B = Lighting Thickness (mm):'
        , 'C = Lighting Inner Diameter (mm):', 'D = Lighting Working Distance (mm):', 'E = Lens Working Distance (mm):'
        , 'F = Field of View (mm x mm):']

        self.worksheet.merge_cells('M4:N4')
        set_outer_border(ws = self.worksheet, min_row = 4, min_col = 'M', max_col='N')
        self.cell_insert_data(self.worksheet, 'Testing Criteria', 'string', 'M4', auto_fit_col = False, auto_fit_row = False)

        set_outer_border(ws = self.worksheet, min_row = 4, min_col = 'O')
        self.cell_insert_data(self.worksheet, 'Length', 'string', 'O4')

        for list_index, i in enumerate(range(5, 5+len(test_criteria_table))):
            self.worksheet.merge_cells('M' +  str(i) + ':' + 'N' + str(i))
            set_outer_border(ws = self.worksheet, min_row = i, min_col = 'M', max_col='N', border_style = 'thin')

            self.cell_insert_data(self.worksheet, test_criteria_table[list_index]
                , 'string', 'M' + str(i), auto_fit_col = False, auto_fit_row = False
                , alignment_dict = dict(wrapText=True, horizontal = 'left', vertical = 'center'))

            set_outer_border(ws = self.worksheet, min_row = i, min_col = 'O', border_style  = 'thin')
            self.cell_insert_data(self.worksheet, '*', 'string', 'P' + str(i)
                , auto_fit_override = (False, True)
                , font_dict = dict(name='Arial',
                size=16,
                bold=False,
                italic=False, 
                vertAlign=None,
                underline='none',
                strike=False,
                color='FFFF0000')
                , alignment_dict = dict(wrapText=True, horizontal = 'left', vertical = 'center'))

        del test_criteria_table
        ###### Lighting Model Table
        light_model_table = ['Lighting Model:', 'Lighting Color:'
        , 'Accessories\n' + '(Polarizer/Diffuser/Extension\nCable/Mounting Bracket):'
        , 'Lighting Voltage / Current (mA):']

        self.worksheet['M12'].fill = PatternFill(start_color = COLOR_INDEX[51], fill_type = 'solid')

        for list_index, i in enumerate(range(12, 12+len(light_model_table))):
            set_outer_border(ws = self.worksheet, min_row = i, min_col = 'M', border_style = 'thin')
            self.cell_insert_data(self.worksheet, light_model_table[list_index]
                , 'string', 'M' + str(i)
                , auto_fit_override = (False, True)
                , alignment_dict = dict(wrapText=True, horizontal = 'left', vertical = 'center'))

            set_outer_border(ws = self.worksheet, min_row = i, min_col = 'N', border_style = 'thin')

        del light_model_table

        ###### Controller Model Table
        ctrl_model_table = ['Controller Model:', 'Controller Mode (Constant /\n'+ 'Strobe / Trigger):'
        , 'Current Multiplier:', 'Intensity:']
        self.worksheet['M17'].fill = PatternFill(start_color = COLOR_INDEX[11], fill_type = 'solid')

        for list_index, i in enumerate(range(17, 17+len(ctrl_model_table))):
            set_outer_border(ws = self.worksheet, min_row = i, min_col = 'M', border_style = 'thin')
            self.cell_insert_data(self.worksheet, ctrl_model_table[list_index]
                , 'string', 'M' + str(i)
                , auto_fit_override = (False, True)
                , alignment_dict = dict(wrapText=True, horizontal = 'left', vertical = 'center'))

            set_outer_border(ws = self.worksheet, min_row = i, min_col = 'N', border_style = 'thin')

        del ctrl_model_table

        ###### Camera Model Table
        camera_model_table = ['Camera Model:', 'Sensor Size:', 'Resolution (pixel):'
        , 'Megapixel:', 'Sensor Type:', 'Pixel size (' + chr(ord('\u03BC')) + 'M):', 'Shutter Type:'
        , 'Mono/Color:', 'Y axis (mm):', 'X axis(mm):']
        self.worksheet['M22'].fill = PatternFill(start_color = COLOR_INDEX[5], fill_type = 'solid')

        for list_index, i in enumerate(range(22, 22+len(camera_model_table))):
            set_outer_border(ws = self.worksheet, min_row = i, min_col = 'M', border_style = 'thin')
            self.cell_insert_data(self.worksheet, camera_model_table[list_index]
                , 'string', 'M' + str(i)
                , auto_fit_override = (False, True)
                , alignment_dict = dict(wrapText=True, horizontal = 'left', vertical = 'center'))

            set_outer_border(ws = self.worksheet, min_row = i, min_col = 'N', border_style = 'thin')
            self.cell_insert_data(self.worksheet, '*', 'string', 'O' + str(i)
                , font_dict = dict(name='Arial',
                size=16,
                bold=False,
                italic=False, 
                vertAlign=None,
                underline='none',
                strike=False,
                color='FFFF0000')
                , alignment_dict = dict(wrapText=True, horizontal = 'left', vertical = 'center'))

        del camera_model_table

        ###### Lens Model Table
        lens_model_table = ['Lens Model:', 'Focal Length (mm):', 'Mount:'
        , 'Sensor Size (max.):', 'TV distortion (%):', 'F.O.V (mm):']
        self.worksheet['M33'].fill = PatternFill(start_color = COLOR_INDEX[40], fill_type = 'solid')

        for list_index, i in enumerate(range(33, 33+len(lens_model_table))):
            set_outer_border(ws = self.worksheet, min_row = i, min_col = 'M', border_style = 'thin')
            self.cell_insert_data(self.worksheet, lens_model_table[list_index]
                , 'string', 'M' + str(i)
                , auto_fit_override = (False, True)
                , alignment_dict = dict(wrapText=True, horizontal = 'left', vertical = 'center'))

            set_outer_border(ws = self.worksheet, min_row = i, min_col = 'N', border_style = 'thin')
            self.cell_insert_data(self.worksheet, '*', 'string', 'O' + str(i)
                , font_dict = dict(name='Arial',
                size=16,
                bold=False,
                italic=False, 
                vertAlign=None,
                underline='none',
                strike=False,
                color='FFFF0000')
                , alignment_dict = dict(wrapText=True, horizontal = 'left', vertical = 'center'))

        del lens_model_table
        #####################################################################################################
        set_outer_border(ws = self.worksheet, min_row = 3, max_row = 39, min_col = 'Q', max_col = 'V')
        self.worksheet.merge_cells('Q2:V2')
        set_outer_border(ws = self.worksheet, min_row = 2, min_col = 'Q', max_col = 'V')
        self.cell_insert_data(self.worksheet, 'Targeted Image', 'string', 'Q2', auto_fit_col = False, auto_fit_row = False)

        #####################################################################################################
        set_outer_border(ws = self.worksheet, min_row = 3, max_row = 39, min_col = 'W', max_col = 'AB')
        self.worksheet.merge_cells('W2:AB2')
        set_outer_border(ws = self.worksheet, min_row = 2, min_col = 'W', max_col = 'AB')
        self.cell_insert_data(self.worksheet, 'Grayscale', 'string', 'W2', auto_fit_col = False, auto_fit_row = False)
        
        #####################################################################################################
        set_outer_border(ws = self.worksheet, min_row = 3, max_row = 39, min_col = 'AC', max_col = 'AH')
        self.worksheet.merge_cells('AC2:AH2')
        set_outer_border(ws = self.worksheet, min_row = 2, min_col = 'AC', max_col = 'AH')
        self.cell_insert_data(self.worksheet, 'Binary', 'string', 'AC2', auto_fit_col = False, auto_fit_row = False)

        #####################################################################################################
        set_outer_border(ws = self.worksheet, min_row = 3, max_row = 39, min_col = 'AI', max_col = 'AN')
        self.worksheet.merge_cells('AI2:AN2')
        set_outer_border(ws = self.worksheet, min_row = 2, min_col = 'AI', max_col = 'AN')
        self.cell_insert_data(self.worksheet, 'Setup', 'string', 'AI2', auto_fit_col = False, auto_fit_row = False)
        
        # print('self.worksheet.max_row: ', self.worksheet.max_row)

    def cell_insert_img(self, ws, img, row_num, start_col = 'A', end_col = None
        , check_row_num = 1, check_col_char = 'A', get_text_cell_id = False):

        if end_col is None:
            end_col = start_col

        if get_text_cell_id == True:
            text_cell_id_list = []

        new_max_row = row_num #We initialize it to row_num, assuming that is the starting point of the 1st image anchor point.
        padding_space = 1

        if (isinstance(img, np.ndarray)) == True:
            start_col_num = xl_col_label_num(start_col)
            # col_width, no_of_col = self.xl_get_column_width(ws, start_col, end_col)
            # xl_2_cm = np.multiply(no_of_col, 1.65) - np.multiply(1.65, 0.3) #column width:8 units = 1.65cm,  row height:14.4units = 0.5cm
            imwidth = self.cell_compute_imwidth(ws, start_col, end_col)

            cell_id = start_col + str(row_num)

            CHECK_cell_id = check_col_char + str(int(check_row_num))

            img_exist_dict = self.cell_check_img(ws, CHECK_cell_id, imwidth)
            # print(img_exist_dict.items())
            if len(img_exist_dict) > 0:
                ## Calculating the new starting cell id for img anchor...
                new_cell_id, _, _ = self.__cell_next_anchor_img(ws, cell_id = cell_id, img_exist_dict = img_exist_dict, img_spacing = padding_space)

            else:
                new_cell_id = cell_id

            # print('imwidth: ', imwidth, 'col_width: ', col_width)
            img_obj = openpyxl.drawing.image.Image(img) #data can be a path to the img in string, or np.array img matrix

            new_width, new_height = self.image_resize_compute(img_obj.width, img_obj.height, new_W = imwidth)
            width_cm = round_float(np.divide(new_width, 37.795), 4)
            height_cm = round_float(np.divide(new_height, 37.795), 4)

            img_cm_size = (width_cm, height_cm)

            self.cell_anchor_img(ws, img_obj, new_cell_id, imwidth)

            ws.add_image(img_obj) # adding in the image

            _, imtext_cell_id, new_max_row = \
            self.__cell_next_anchor_img(ws, cell_id = new_cell_id, img_cm_size = img_cm_size, img_spacing = padding_space)

            del width_cm, height_cm, img_cm_size

            if get_text_cell_id == True:
                text_cell_id_list.append(imtext_cell_id)

        elif type(img) == str:
            if os.path.isfile(img) == False:
                raise AttributeError("Image File does not exists!")
            elif os.path.isfile(img) == True:
                start_col_num = xl_col_label_num(start_col)

                # col_width, no_of_col = self.xl_get_column_width(ws, start_col, end_col)
                # xl_2_cm = np.multiply(no_of_col, 1.65) - np.multiply(1.65, 0.3) #column width:8 units = 1.65cm,  row height:14.4units = 0.5cm

                imwidth = self.cell_compute_imwidth(ws, start_col, end_col)

                cell_id = start_col + str(row_num)

                CHECK_cell_id = check_col_char + str(int(check_row_num))

                img_exist_dict = self.cell_check_img(ws, CHECK_cell_id, imwidth)
                # print(img_exist_dict.items())
                if len(img_exist_dict) > 0:
                    ## Calculating the new starting cell id for img anchor...
                    new_cell_id, _, _ = self.__cell_next_anchor_img(ws, cell_id = cell_id, img_exist_dict = img_exist_dict, img_spacing = padding_space)

                else:
                    new_cell_id = cell_id

                # print('imwidth: ', imwidth, 'col_width: ', col_width)
                img_obj = openpyxl.drawing.image.Image(img) #data can be a path to the img in string, or np.array img matrix

                new_width, new_height = self.image_resize_compute(img_obj.width, img_obj.height, new_W = imwidth)
                width_cm = round_float(np.divide(new_width, 37.795), 4)
                height_cm = round_float(np.divide(new_height, 37.795), 4)

                img_cm_size = (width_cm, height_cm)

                self.cell_anchor_img(ws, img_obj, new_cell_id, imwidth)

                ws.add_image(img_obj) # adding in the image

                _, imtext_cell_id, new_max_row = \
                self.__cell_next_anchor_img(ws, cell_id = new_cell_id, img_cm_size = img_cm_size, img_spacing = padding_space)

                # print("IMG PATH: ", new_max_row)

                del width_cm, height_cm, img_cm_size

                if get_text_cell_id == True:
                    text_cell_id_list.append(imtext_cell_id)

        elif type(img) == dict or type(img) == list or isinstance(img, OrderedDict):
            start_col_num = xl_col_label_num(start_col)

            # col_width, no_of_col = self.xl_get_column_width(ws, start_col, end_col)
            # xl_2_cm = np.multiply(no_of_col, 1.65) - np.multiply(1.65, 0.3) #column width:8 units = 1.65cm,  row height:14.4units = 0.5cm
            imwidth = self.cell_compute_imwidth(ws, start_col, end_col)

            cell_id = start_col + str(row_num)
            
            CHECK_cell_id = check_col_char + str(int(check_row_num))

            img_exist_dict = self.cell_check_img(ws, CHECK_cell_id, imwidth)
            # print(img_exist_dict.items())
            if len(img_exist_dict) > 0:
                ## Calculating the new starting cell id for img anchor...
                new_cell_id, _, _ = self.__cell_next_anchor_img(ws, cell_id = cell_id, img_exist_dict = img_exist_dict, img_spacing = padding_space)

            else:
                new_cell_id = cell_id

            if type(img) == list:
                if len(img) > 0:
                    # img_no = 0 #image number. if a list is provided, this is used to track the img no.
                    # for img_no, imdata in enumerate(img):
                    for imdata in img:
                        if type(imdata) == str or (isinstance(imdata, np.ndarray) == True) or (isinstance(imdata, openpyxl.drawing.image.Image) == True):
                            if type(imdata) == str or (isinstance(imdata, np.ndarray) == True):
                                img_obj = openpyxl.drawing.image.Image(imdata)

                            else:
                                img_obj = imdata

                            new_width, new_height = self.image_resize_compute(img_obj.width, img_obj.height, new_W = imwidth)
                            width_cm = round_float(np.divide(new_width, 37.795), 4)
                            height_cm = round_float(np.divide(new_height, 37.795), 4)

                            img_cm_size = (width_cm, height_cm)

                            self.cell_anchor_img(ws, img_obj, new_cell_id, imwidth)

                            ws.add_image(img_obj)

                            new_cell_id, imtext_cell_id, new_max_row = \
                            self.__cell_next_anchor_img(ws, cell_id = new_cell_id, img_cm_size = img_cm_size, img_spacing = padding_space)

                            del width_cm, height_cm, img_cm_size

                            if get_text_cell_id == True:
                                text_cell_id_list.append(imtext_cell_id)

            elif type(img) == dict or isinstance(img, OrderedDict):
                if len(img) > 0:
                    for imdata in img.values():
                        if type(imdata) == str or (isinstance(imdata, np.ndarray) == True) or (isinstance(imdata, openpyxl.drawing.image.Image) == True):
                            if type(imdata) == str or (isinstance(imdata, np.ndarray) == True):
                                img_obj = openpyxl.drawing.image.Image(imdata)

                            else:
                                img_obj = imdata
                                
                            new_width, new_height = self.image_resize_compute(img_obj.width, img_obj.height, new_W = imwidth)
                            width_cm = round_float(np.divide(new_width, 37.795), 2)
                            height_cm = round_float(np.divide(new_height, 37.795), 2)

                            img_cm_size = (width_cm, height_cm)

                            self.cell_anchor_img(ws, img_obj, new_cell_id, imwidth)

                            ws.add_image(img_obj)

                            new_cell_id, imtext_cell_id, new_max_row = \
                            self.__cell_next_anchor_img(ws, cell_id = new_cell_id, img_cm_size = img_cm_size, img_spacing = padding_space)

                            # print(new_max_row)
                            # print('img new_cell_id: ', new_cell_id)

                            del width_cm, height_cm, img_cm_size

                            if get_text_cell_id == True:
                                text_cell_id_list.append(imtext_cell_id)

        
        if get_text_cell_id == True:
            return new_max_row, text_cell_id_list

        else:
            return new_max_row

    def cell_compute_imwidth(self, ws, start_col, end_col):
        col_width, _ = self.xl_get_column_width(ws, start_col, end_col)
        start_col_width, _ = self.xl_get_column_width(ws, start_col)

        # print('col_width: ', col_width)

        col_cm = np.multiply(col_width, self.__width_cm_xlunit_ratio)
        start_col_cm = np.multiply(start_col_width, self.__width_cm_xlunit_ratio)
        xl_2_cm = col_cm - np.multiply(start_col_cm, 0.4)

        imwidth = np.multiply(xl_2_cm, 37.795) #cm to pixel conversion

        del col_width, start_col_width, col_cm, start_col_cm, xl_2_cm

        return imwidth


    def cell_check_img(self, worksheet, cell_id, imwidth):
        img_exist_dict = {} #Image dictionary which contains anchor location (dictionary key) & imheight (dictionary value)
        # width_cm = round_float(np.divide(pixel_width, 37.795), 4) ####IMPORTANT FOMULA PIXEL --> CM.
        # height_cm = round_float(np.divide(pixel_height, 37.795), 4) ####IMPORTANT FOMULA PIXEL --> CM.
        
        xl_imloader = SheetImageLoader(worksheet, sort_col_bool = True)
        # print(xl_imloader._images)

        col_char = re.findall('(\\D+)',cell_id)[0]
        row_num = int(re.findall('(\\d+)',cell_id)[0])

        deletion_list = [] ### We will prepare the deletion list to re-anchor the existing image(s) with resizing process.
        ### Important! Currently, no plans on re-adjusting the image anchor.

        for ref_cell_id, img_obj in xl_imloader._images.items():
            ### We Calculate the arbitary dimensions of existing image(s) with respect to the column(s) width (e.g. A to F)
            ### Important! We only do this if ref_col_char is the same as the col_char. We will only iterate on 1 column/merge column this way.

            ref_col_char = re.findall('(\\D+)', ref_cell_id)[0]
            if ref_col_char == col_char:
                pil_image = xl_imloader.get(ref_cell_id)
                # print(pil_image)
                deletion_list.append(img_obj) #We Delete Existing Images, and Place it Again with Standard Resize Properties.

                new_img_obj = openpyxl.drawing.image.Image(np.array(pil_image))
                del pil_image

                self.cell_anchor_img(worksheet, new_img_obj, ref_cell_id, imwidth)
                worksheet.add_image(new_img_obj)

                new_width, new_height = self.image_resize_compute(img_obj.width, img_obj.height, new_W = imwidth)
                width_cm = round_float(np.divide(new_width, 37.795), 4)
                height_cm = round_float(np.divide(new_height, 37.795), 4)

                img_exist_dict[ref_cell_id] = (width_cm, height_cm)

        # print('img_exist_dict: ', img_exist_dict)
        for deletion_id in deletion_list:
            del worksheet._images[worksheet._images.index(deletion_id)]

        del deletion_list
        del xl_imloader

        return img_exist_dict

    def __cell_next_anchor_img(self, ws, cell_id, img_exist_dict = {}, img_cm_size = None, img_spacing = 0): #calculate the location of the anchor for next image
        img_col_char = re.findall('(\\D+)', cell_id)[0]
        ### We 1st initialize the new_cell_id equal to cell_id indicating that there is NO new anchor location.
        new_cell_id = "{}{}".format(img_col_char, min(int(re.findall('(\\d+)', cell_id)[0]), 1048576))

        ### We 1st initialize the final_row_id
        final_row_id = min(int(re.findall('(\\d+)', cell_id)[0]), 1048576)

        ### We 1st initialize the imtext_cell_id, imtext_cell_id is the cell_id for image text(s)/description(s). Normally, located right after the end of image anchor.
        imtext_cell_id = "{}{}".format(img_col_char, min(int(re.findall('(\\d+)', cell_id)[0]) + 1, 1048576))


        if len(img_exist_dict) > 0:
            end_imrow_arr = np.zeros((len(img_exist_dict)), dtype=np.uint16) ## Array containing the ending row number for each scanned/existing image(s) in the file
            ### We then find the max(end_imrow_arr) to determine the next ost relevant row number for new set of image(s)
            i = 0

            for img_cell_id, imsize in img_exist_dict.items():
                img_row_num = int(re.findall('(\\d+)', img_cell_id)[0])
                height_cm = imsize[1]
                if img_row_num < 1048576: ## Maximum row number in Excel is 1048576
                    cusum_height = 0
                    iter_index = img_row_num
                    loop_break = False
                    while loop_break == False:
                        row_height, _ = self.xl_get_row_height(ws, iter_index)
                        row_height = round_float(np.multiply(row_height, self.__height_cm_xlunit_ratio), 2)

                        cusum_height = cusum_height + row_height
                        if (cusum_height + self.__xl_img_cm_tolerance) >= height_cm:
                            loop_break = True
                            break

                        if iter_index > 1048576:
                            loop_break = True
                            break

                        iter_index += 1
                    
                    del loop_break

                    end_row_num = iter_index
                else:
                    end_row_num = 1048576
                
                end_imrow_arr[i] = int(end_row_num)

                i += 1

            del new_cell_id
            new_cell_id = "{}{}".format(img_col_char, min(max(end_imrow_arr) + 1 + img_spacing, 1048576) )
            imtext_cell_id = "{}{}".format(img_col_char, min(max(end_imrow_arr) + 1, 1048576) )
            final_row_id = min(max(end_imrow_arr) + img_spacing, 1048576) ### final_row_id is the final row_number with included img_spacing

            # print("Checking img exist process...: ", new_cell_id, end_imrow_arr)
            del i, end_imrow_arr

        else:
            if isinstance(img_cm_size, tuple) == True and len(img_cm_size) == 2:
                # print(img_cm_size, type(img_cm_size[0]), type(img_cm_size[1]))
                if (validate_unsign_int(img_cm_size[0]) == True or validate_unsign_float(img_cm_size[0]) == True)\
                and (validate_unsign_int(img_cm_size[1]) == True or validate_unsign_float(img_cm_size[1]) == True):

                    img_row_num = min(int(re.findall('(\\d+)', cell_id)[0]), 1048576)
                    height_cm = img_cm_size[1]

                    if img_row_num < 1048576: ## Maximum row number in Excel is 1048576
                        cusum_height = 0
                        iter_index = img_row_num
                        loop_break = False
                        while loop_break == False:
                            row_height, _ = self.xl_get_row_height(ws, iter_index)
                            row_height = round_float(np.multiply(row_height, self.__height_cm_xlunit_ratio), 2)

                            cusum_height = cusum_height + row_height
                            if (cusum_height + self.__xl_img_cm_tolerance) >= height_cm:
                                loop_break = True
                                break

                            if iter_index > 1048576: ## We have reached the final excel row.
                                loop_break = True
                                break

                            iter_index += 1
                        
                        del loop_break

                        end_row_num = iter_index
                    else:
                        end_row_num = 1048576

                    del new_cell_id
                    new_cell_id = "{}{}".format(img_col_char, min(int(end_row_num) + 1 + img_spacing, 1048576) )
                    imtext_cell_id = "{}{}".format(img_col_char, min(int(end_row_num) + 1, 1048576) )
                    final_row_id = min(int(end_row_num) + img_spacing, 1048576) ### final_row_id is the final row_number with included img_spacing

                else:
                    raise AttributeError("img_cm_size is a tuple containing img size in centimetres(cm), must be a positive float or int!")
            
            else:
                raise AttributeError("If img_exist_dict is not used, please include the img_cm_size.\nimg_cm_size is a tuple containing img size in centimetres(cm)")

        # print(new_cell_id)
        

        return new_cell_id, imtext_cell_id, final_row_id


    def cell_anchor_img(self, ws, img_obj, cell_id, imwidth, col_offset = float(0.2), row_offset = float(0)):
        ## Before 21-1-2022: col_offset = float(0.2), row_offset = float(0.2)
        ## After 21-1-2022: col_offset = float(0.2), row_offset = float(0)

        new_width, new_height = self.image_resize_compute(img_obj.width, img_obj.height, new_W = imwidth)
        # print('new_width, new_height: ', new_width, new_height)
        img_obj.height = new_height # insert image height in pixels as float or int (e.g. 305.5)
        img_obj.width = new_width # insert image width in pixels as float or int (e.g. 405.8)

        col_char = re.findall('(\\D+)',cell_id)[0]
        row_num = int(re.findall('(\\d+)',cell_id)[0])

        xl_imsize = XDRPositiveSize2D(p2e(new_width), p2e(new_height))

        col_width, _ = self.xl_get_column_width(ws, col_char)
        row_height, _ = self.xl_get_row_height(ws, row_num)

        col_cm = np.multiply(col_width, self.__width_cm_xlunit_ratio)
        row_cm = np.multiply(row_height, self.__height_cm_xlunit_ratio)

        coloffset = xl_cm_to_emu(col_cm, col_offset)
        rowoffset = xl_cm_to_emu(row_cm, row_offset)

        marker = AnchorMarker(col= xl_col_label_num(col_char) - 1, colOff=coloffset, row=row_num - 1, rowOff=rowoffset)
        # img_obj.anchor = start_col + str(int(row_num)) # where you want image to be anchored/start from
        img_obj.anchor = OneCellAnchor(_from=marker, ext=xl_imsize)


    def compute_next_anchor_img(self, ws, img_obj, cell_id, start_col, end_col = None, img_spacing = 0):
        ### img_obj must be a openpyxl.drawing.image.Image object.
        err_flag_arr = np.zeros((2,), dtype = bool)
        start_col_num = end_col_num = None

        if False == type_int(start_col):
            if type(start_col) == str:
                start_col_num = xl_col_label_num(start_col)

                if type_int(start_col_num) == False:
                    err_flag_arr[0] = True
                    raise ValueError("start_col must be a string corresponding to Excel column characters from 'A' to 'XFD'")

                elif type_int(start_col_num) == True:
                    if 1 <= start_col_num <= 16384:
                        pass
                    else:
                        err_flag_arr[0] = True
                        raise ValueError("start_col must be a string corresponding to Excel column characters from 'A' to 'XFD'")

            else:
                err_flag_arr[0] = True
                raise ValueError("start_col must be an int type or str type input")

        elif True == type_int(start_col):
            if 1 <= start_col <= 16384:
                start_col_num = copy.copy(start_col)
                start_col = xl_col_label_char(start_col_num)

            else:
                err_flag_arr[0] = True
                raise ValueError("start_col must be an integer from 1 to 16384")

        if False == type_int(end_col):
            if type(end_col) == str:
                end_col_num = xl_col_label_num(end_col)
                
                if type_int(end_col_num) == False:
                    err_flag_arr[1] = True
                    raise ValueError("end_col must be a string corresponding to Excel column characters from 'A' to 'XFD'")

                elif type_int(end_col_num) == True:
                    if 1 <= end_col_num <= 16384:
                        pass
                    else:
                        err_flag_arr[1] = True
                        raise ValueError("end_col must be a string corresponding to Excel column characters from 'A' to 'XFD'")

            elif end_col is None and err_flag_arr[0] == False:
                end_col_num = start_col_num

            else:
                err_flag_arr[1] = True
                raise ValueError("end_col must be an int type or str type input")

        elif True == type_int(end_col):
            if 1 <= end_col <= 16384:
                end_col_num = copy.copy(end_col)
                end_col = xl_col_label_char(end_col)

            else:
                err_flag_arr[1] = True
                raise ValueError("end_col must be an integer from 1 to 16384")


        if np.any(err_flag_arr) == False and type_int(start_col_num) == True and type_int(end_col_num) == True:
            if end_col_num < start_col_num:
                end_col = start_col

            # print('start_col: ', start_col)
            # print('end_col: ', end_col)

            imwidth = self.cell_compute_imwidth(ws, start_col, end_col)
            ### We assume that the image is nicely fit within the start_col and end_col

            new_height = self.image_resize_compute(img_obj.width, img_obj.height, new_W = imwidth)[1]
            # new_width, new_height = self.image_resize_compute(img_obj.width, img_obj.height, new_W = imwidth)
            # width_cm = round_float(np.divide(new_width, 37.795), 4)
            height_cm = round_float(np.divide(new_height, 37.795), 4)

            img_col_char = re.findall('(\\D+)', cell_id)[0]
            img_row_num = int(re.findall('(\\d+)', cell_id)[0])
            if img_row_num < 1048576: ## Maximum row number in Excel is 1048576
                cusum_height = 0
                iter_index = img_row_num
                loop_break = False
                while loop_break == False:
                    row_height, _ = self.xl_get_row_height(ws, iter_index)
                    row_height = round_float(np.multiply(row_height, self.__height_cm_xlunit_ratio), 2)

                    cusum_height = cusum_height + row_height
                    if (cusum_height + self.__xl_img_cm_tolerance) >= height_cm:
                        loop_break = True
                        break

                    if iter_index > 1048576:
                        loop_break = True
                        break

                    iter_index += 1

                end_row_num = iter_index
            else:
                end_row_num = 1048576
            
            new_cell_id = "{}{}".format(img_col_char, min(int(end_row_num) + 1 + img_spacing, 1048576) )
            imtext_cell_id = "{}{}".format(img_col_char, min(int(end_row_num) + 1, 1048576) )
            final_row_id = min(int(end_row_num) + img_spacing, 1048576) ### final_row_id is the final row_number with included img_spacing

            return new_cell_id, imtext_cell_id, final_row_id

        else:
            raise ValueError("Please check:\\n"
                + "1. start_col or end_col must be a string corresponding to Excel column characters from 'A' to 'XFD'"
                + "2. If start_col or end_col are int-type, ensure the range is from 1 to 16384")


    def cell_insert_data(self, ws, data, data_type, cell_location = None, border_bool = False
        , auto_fit_col = True, auto_fit_row = True, auto_fit_override = (False, False)
        , font_dict = dict(name='Arial',
            size=16,
            bold=False,
            italic=False, 
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')

        , alignment_dict = dict(wrapText = True
            , horizontal = 'center'
            , vertical = 'center') ):
        '''
        cell_insert_data(): auto_fit_override[0] = override for auto_fit_col
                            auto_fit_override[1] = override for auto_fit_row
        '''

        cell_id = None
        if cell_location is not None:
            if type(cell_location) == str:
                cell_id = cell_location

            elif type(cell_location) == tuple:
                if type(cell_location[0]) == int and type(cell_location[1]) == int:
                    cell_row, cell_col = cell_location
                    cell_id = str(self.worksheet.cell(row=cell_row, column=cell_col).column_letter) \
                    + str(self.worksheet.cell(row=cell_row,column=cell_col).row)

            try:
                ws[cell_id]
            except Exception:
                cell_id = None

            # print('cell_id: ', cell_id)

            if cell_id is not None:
                # print(re.findall('\\d+', cell_id)) #id number without the alphabet
                ws[cell_id].alignment = Alignment(**alignment_dict)
                ws[cell_id].font = Font(**font_dict)

                if border_bool == True:
                    ws[cell_id].border = Border(left = Side(border_style='thin', color='FF000000')
                                                , right = Side(border_style='thin', color='FF000000')
                                                , top = Side(border_style='thin', color='FF000000')
                                                , bottom = Side(border_style='thin', color='FF000000'))

                if data_type == 'image':
                    try:
                        img_obj = openpyxl.drawing.image.Image(data) #data can be a path to the img in string, or np.array img matrix
                    except Exception:# as e:
                        # print('Exception Error: ', e)
                        data = None
                        img_obj = None

                    if data is not None:
                        new_width, new_height = self.image_resize_compute(img_obj.width, img_obj.height ,new_W = 230)

                        # print('new_width, new_height: ', new_width, new_height)
                        img_obj.height = new_height # insert image height in pixels as float or int (e.g. 305.5)
                        img_obj.width = new_width # insert image width in pixels as float or int (e.g. 405.8)
                        img_obj.anchor = cell_id # where you want image to be anchored/start from
                        ws.add_image(img_obj) # adding in the image

                        adjust_col_width = None
                        adjust_row_height = None

                        if auto_fit_col == True:
                            adjust_col_width = self.fit_cell_column(pixel_width = new_width)
                        
                        if auto_fit_row == True:
                            adjust_row_height = self.fit_cell_row(pixel_height = new_height)

                        # print(re.split('(\\d+)',cell_id)[0])
                        # print(int(re.findall('(\\d+)',cell_id)[0]))

                        curr_width = ws.column_dimensions[re.findall('(\\D+)',cell_id)[0]].width 
                        curr_height = ws.row_dimensions[int(re.findall('(\\d+)',cell_id)[0])].height
                        # print('Data type: Image')
                        # print('curr_width, curr_height: ', curr_width, curr_height)
                        # print('adjust_col_width, adjust_row_height: ', adjust_col_width, adjust_row_height)

                        if curr_height is None:
                            curr_height = 15

                        if adjust_col_width is not None:
                            if auto_fit_override[0] == True:
                                ws.column_dimensions[re.findall('(\\D+)',cell_id)[0]].width = adjust_col_width
                            else:
                                if adjust_col_width > curr_width:
                                    ws.column_dimensions[re.findall('(\\D+)',cell_id)[0]].width = adjust_col_width

                        if adjust_row_height is not None:
                            if auto_fit_override[1] == True:
                                ws.row_dimensions[int(re.findall('(\\d+)',cell_id)[0])].height = adjust_row_height
                            else:
                                if adjust_row_height > curr_height:
                                    ws.row_dimensions[int(re.findall('(\\d+)',cell_id)[0])].height = adjust_row_height

                        elif adjust_row_height is None:
                            if curr_height < 21:
                                ws.row_dimensions[int(re.findall('(\\d+)',cell_id)[0])].height = 21
                            #Current this algorithm generates Excel with Reference to Arial Font 16. So default height must be set 21


                elif data_type == 'string':
                    ws[cell_id] = data

                    adjust_col_width = None
                    adjust_row_height = None

                    if auto_fit_col == True:
                        adjust_col_width = self.fit_cell_column(string = data)
                    
                    if auto_fit_row == True:
                        adjust_row_height = self.fit_cell_row(string = data)

                    # print(re.split('(\\d+)',cell_id)[0])
                    # print(int(re.findall('(\\d+)',cell_id)[0]))

                    curr_width = ws.column_dimensions[re.findall('(\\D+)',cell_id)[0]].width 
                    curr_height = ws.row_dimensions[int(re.findall('(\\d+)',cell_id)[0])].height

                    # print('Data type: String')
                    # print('curr_width, curr_height: ', curr_width, curr_height)
                    # print('adjust_col_width, adjust_row_height: ', adjust_col_width, adjust_row_height)

                    if curr_height is None:
                        curr_height = 15
                    # print(ws.column_dimensions[re.split('(\\d+)',cell_id)[0]].width)

                    if adjust_col_width is not None:
                        if auto_fit_override[0] == True:
                            ws.column_dimensions[re.findall('(\\D+)',cell_id)[0]].width = adjust_col_width
                        else:
                            if adjust_col_width > curr_width:
                                ws.column_dimensions[re.findall('(\\D+)',cell_id)[0]].width = adjust_col_width

                    if adjust_row_height is not None:
                        if auto_fit_override[1] == True:
                            ws.row_dimensions[int(re.findall('(\\d+)',cell_id)[0])].height = adjust_row_height
                        else:
                            if adjust_row_height > curr_height:
                                ws.row_dimensions[int(re.findall('(\\d+)',cell_id)[0])].height = adjust_row_height

                    elif adjust_row_height is None:
                        if curr_height < 21:
                            ws.row_dimensions[int(re.findall('(\\d+)',cell_id)[0])].height = 21
                        #Current this algorithm generates Excel with Reference to Arial Font 16. So default height must be set 21

    def fit_cell_row(self, string = None, pixel_height = None): #Currently this works if current font is default font
        try:
            pixel_height = float(pixel_height)
        except (ValueError, TypeError):
            pixel_height = None

        if (string is not None and type(string) == str) and (pixel_height is None):
            # print(string)
            newline_num = len(string.split('\n'))
            # cell_height = np.multiply(newline_num, 15)
            cell_height = np.multiply(newline_num, np.multiply(15, self.height_font_factor) )

            resize_cell_height = cell_height

            return resize_cell_height

        elif (string is None) and (pixel_height is not None and (type(pixel_height) == int or type(pixel_height) == float)):
            rel_pixel_height =  np.multiply(np.multiply(pixel_height, 0.0264583333), np.divide(14.4, 0.53)) + 8
            # print(rel_pixel_height)
            resize_cell_height = rel_pixel_height

            return resize_cell_height

        elif (string is not None and type(string) == str) and (pixel_height is not None and (type(pixel_height) == int or type(pixel_height) == float)):
            newline_num = len(string.split('\n'))
            # cell_height = np.multiply(newline_num, 15)
            # cell_height = np.multiply(newline_num, 20.40)
            cell_height = np.multiply(newline_num, np.multiply(15, self.height_font_factor) )

            rel_pixel_height =  np.multiply(np.multiply(pixel_height, 0.0264583333), np.divide(14.4, 0.53)) + 8

            if cell_height < rel_pixel_height:
                resize_cell_height = rel_pixel_height

            elif cell_height > rel_pixel_height:
                resize_cell_height = cell_height

            else:
                resize_cell_height = cell_height

            return resize_cell_height

        else:
            return None

    def fit_cell_column(self, string = None, pixel_width = None): #Currently this works if current font is default font
        if (string is not None and type(string) == str) and (pixel_width is None):
            newline_num = len(string.split('\n'))
            if newline_num > 1:
                string_list = string.split('\n')
                long_string = max(string_list, key = len)

                str_width = getTextWidth(self.ref_font, long_string, 16
                    , glyph_set = self.glyph_set, font_cmap = self.font_cmap, units_per_em = self.units_per_em)
                excel_char_unit = np.divide(str_width, self.unit_char_width)

                ## excel_char_unit = np.divide(self.ref_font.getsize(long_string)[0], self.ref_font.getsize('0')[0])
                char_width = np.multiply(excel_char_unit - 2, self.width_font_factor[1]) + np.multiply(2, self.width_font_factor[0])
                # print(long_string, len(long_string))
                # print(str_width)
                # print(excel_char_unit)
                
            else:
                str_width = getTextWidth(self.ref_font, string, 16
                    , glyph_set = self.glyph_set, font_cmap = self.font_cmap, units_per_em = self.units_per_em)
                excel_char_unit = np.divide(str_width, self.unit_char_width)

                ## excel_char_unit = np.divide(self.ref_font.getsize(string)[0], self.ref_font.getsize('0')[0])
                ## char_width = np.multiply(excel_char_unit, self.width_font_factor)
                char_width = np.multiply(excel_char_unit - 2, self.width_font_factor[1]) + np.multiply(2, self.width_font_factor[0])
                # print(string, len(string))
                # print(str_width)
                # print(excel_char_unit)

            resize_cell_width = char_width
            # print(resize_cell_width)
            # print('---------------')
            return resize_cell_width + self.width_font_factor[0]

        elif (string is None) and (pixel_width is not None and (type(pixel_width) == int or type(pixel_width) == float)):
            pixel_char_width = float( np.multiply(np.divide(pixel_width, 8), 1.1) )
            resize_cell_width = pixel_char_width

            return resize_cell_width + 0.56

        elif (string is not None and type(string) == str) and (pixel_width is not None and (type(pixel_width) == int or type(pixel_width) == float)):
            newline_num = len(string.split('\n'))
            if newline_num > 1:
                string_list = string.split('\n')
                long_string = max(string_list, key = len)
                # char_width = len(long_string)
                excel_char_unit = np.divide(self.ref_font.getsize(long_string)[0], self.ref_font.getsize('0')[0])
                ## char_width = np.multiply(excel_char_unit, self.width_font_factor)
                char_width = np.multiply(excel_char_unit - 2, self.width_font_factor[1]) + np.multiply(2, self.width_font_factor[0])
            
            else:
                # char_width = len(string)
                excel_char_unit = np.divide(self.ref_font.getsize(string)[0], self.ref_font.getsize('0')[0])
                ## char_width = np.multiply(excel_char_unit, self.width_font_factor)
                char_width = np.multiply(excel_char_unit - 2, self.width_font_factor[1]) + np.multiply(2, self.width_font_factor[0])


            pixel_char_width = float( np.multiply(np.divide(pixel_width, 8), 1.1) )

            if char_width < pixel_char_width:
                resize_cell_width = pixel_char_width

                return resize_cell_width + 0.56

            elif char_width > pixel_char_width:
                resize_cell_width = char_width
                return resize_cell_width + self.width_font_factor[0]

            else:
                resize_cell_width = char_width
                return resize_cell_width + self.width_font_factor[0]

        else:
            return None

    def image_resize_compute(self, ori_W, ori_H, new_W = None, new_H = None):
        if new_W is not None and new_H is None:
            return (new_W , np.multiply(np.divide(ori_H,ori_W), new_W))

        elif new_W is None and new_H is not None:
            return (np.multiply(np.divide(ori_W,ori_H), new_H), new_H)

        elif new_W is not None and new_H is not None:
            return (new_W , new_H)

        elif new_W is None and new_H is None:
            return (None, None)