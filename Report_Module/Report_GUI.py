import os
from os import path
import ntpath

import sys
import copy

from datetime import datetime

import tkinter as tk
from tkinter import filedialog

import threading
import re

from collections import OrderedDict

from PIL import ImageTk, Image, ImageDraw, ImageFont

import cv2

import numpy as np
from functools import partial

import io
import openpyxl
from openpyxl import load_workbook

from tk_openpyxl import XL_WorkBook, xl_read_worksheet, custom_col_char_dict, xl_col_label_num, set_outer_border, xl_unmerge_cell
import tk_opencsv

from Tk_MsgBox.custom_msgbox import Ask_Msgbox, Info_Msgbox, Error_Msgbox, Warning_Msgbox

from misc_module.os_create_folder import open_save_folder
from misc_module.image_resize import img_resize_dim, opencv_img_resize, pil_img_resize, open_pil_img
from misc_module.tk_img_module import to_tk_img
from misc_module.tool_tip import CreateToolTip

from Tk_Custom_Widget.img_preview_disp import CreatePreviewDisp
from Tk_Custom_Widget.ScrolledCanvas import ScrolledCanvas
from Tk_Custom_Widget.tk_custom_combobox import CustomBox
from Tk_Custom_Widget.tk_custom_singletext import CustomSingleText
from Tk_Custom_Widget.tk_custom_toplvl import CustomToplvl

from Tk_Custom_Widget.tk_custom_scrolltext import CustomScrollText

def widget_enable(*widgets):
    for widget in widgets:
        try:
            widget['state'] = 'normal'
        except AttributeError:
            pass

def widget_disable(*widgets):
    for widget in widgets:
        try:
            widget['state'] = 'disable'
        except AttributeError:
            pass

def widget_readonly(*widgets):
    for widget in widgets:
        try:
            widget['state'] = 'readonly'
        except AttributeError:
            pass

def move_element(odict, thekey, newpos):
    if thekey in odict:
        odict[thekey] = odict.pop(thekey)
        if newpos > (len(odict) - 1):
            newpos = (len(odict) - 1)

        # print("newpos: ", newpos)

        i = 0
        for key, value in odict.copy().items():
            if key != thekey and i >= newpos:
                odict[key] = odict.pop(key)
            i += 1
    else:
        raise KeyError("{} does not exist in {}".format(thekey, odict))
    return odict

def remove_key(ref_dict, del_key): #Removing a key & element from a dictionary-type object
    new_dict = {}
    key_pos = 0

    i = 0
    for key, val in ref_dict.items():
        if int(key) < del_key:
            new_dict[key] = val
        elif int(key) > del_key:
            new_dict[str(int(key)-1)] = val
        else: # key == del_key
            key_pos = i
            pass
            # continue

        i += 1

    ref_dict.clear()
    ref_dict.update(new_dict)
    del new_dict

    return ref_dict, key_pos

def character_limit(tk_str_var, limit):
    if len(tk_str_var.get()) > 0:
        tk_str_var.set(tk_str_var.get()[:limit])

def group_print(*args, **kwargs):
    for arg in args:
        print(arg)
    for key, arg in kwargs.items():
        print(key, arg)


class Text_Scroll(tk.Frame):
    def __init__(self, master, scroll_canvas_class = None, **kwargs):
        tk.Frame.__init__(self, master, **kwargs)
        self.scroll_canvas_class = scroll_canvas_class

        self.frame_w = kwargs['width']
        self.frame_h = kwargs['height']
        self.resize_event_w = kwargs['width'] #During init we set this equal to frame_w
        self.resize_event_h = kwargs['height'] #During init we set this equal to frame_h

        self.txt_frame = tk.Frame(self, width = self.frame_w, height = self.frame_h)
        self.txt_frame.place(relx = 0, rely = 0, anchor='nw')
        self.bind("<Configure>", self.on_resize) #auto resize self.txt_frame

        self.tk_txt = tk.Text(self.txt_frame, font = 'Helvetica 10', maxundo = -1, wrap = tk.WORD, relief = tk.GROOVE)
        self.tk_scrolly = tk.Scrollbar(self, command= self.tk_txt.yview)

        self.tk_txt.place(relx=0, rely=0, relheight=1, relwidth=1, width = -16, anchor = 'nw')
        self.tk_scrolly.place(relx=1, rely=0, relheight=1, anchor='ne')

        self.tk_txt.configure(yscrollcommand= self.tk_scrolly.set)

        if self.scroll_canvas_class is not None:
            try:
                self.tk_txt.bind('<Enter>', self._bound_to_mousewheel)
                self.tk_txt.bind('<Leave>', self.scroll_canvas_class._bound_to_mousewheel)
            except Exception:
                pass

    def on_resize(self, event):
        # print('self.frame_w, self.frame_h: ', self.frame_w, self.frame_h)
        # print('event.width, event.height: ', event.width, event.height)
        self.resize_event_h = int(event.height)
        self.resize_event_w = int(event.width)

        if self.frame_h <= self.resize_event_h:
            self.txt_frame['height'] = self.resize_event_h

        elif self.frame_h > self.resize_event_h:
            self.txt_frame['height'] = self.frame_h

        if self.frame_w <= self.resize_event_w:
            self.txt_frame['width'] = self.resize_event_w

        elif self.frame_w > self.resize_event_w:
            self.txt_frame['width'] = self.frame_w

    def insert_readonly(self, *_args):
        self.tk_txt['state'] = 'normal'
        self.tk_txt.insert(*_args)
        self.tk_txt['state'] = 'disable'

    def delete_txt(self):
        # print("self.tk_txt['state']: ", self.tk_txt['state'])
        if self.tk_txt['state'] == 'normal':
            self.tk_txt.delete('1.0', tk.END)

        elif self.tk_txt['state'] == 'disabled':
            self.tk_txt['state'] = 'normal'
            # print("self.tk_txt['state']: ", self.tk_txt['state'])
            self.tk_txt.delete('1.0', tk.END)
            self.tk_txt['state'] = 'disable'
        
        # print('Delete TextBox')

    def _bound_to_mousewheel(self,event):
        #print('Enter')
        self.tk_txt.bind_all("<MouseWheel>", self._on_mousewheel)   

    def _unbound_to_mousewheel(self, event):
        #print('Leave')
        self.tk_txt.unbind_all("<MouseWheel>")

    def _on_mousewheel(self, event):
        self.tk_txt.yview_scroll(int(-1*(event.delta/120)), "units")

class Report_GUI(tk.Frame):
    def __init__(self, master, scroll_canvas_class = None
        , gui_graphic = {}
        , **kwargs):
        tk.Frame.__init__(self, master, **kwargs)
        self.master = master
        self.scroll_canvas_class = scroll_canvas_class

        self.img_format_list = [("All Files","*.bmp *.jpg *.jpeg *.png *.tiff"),
                                ('BMP file', '*.bmp'),
                                ('JPG file', '*.jpg'),
                                ('JPEG file', '*.jpeg'),
                                ('PNG file', '*.png'),
                               ('TIFF file', '*.tiff')]

        self.report_format_list = [('XLSX file', '*.xlsx')]


        self.gui_graphic = dict(  toggle_ON_btn_img = None, toggle_OFF_btn_img = None
                                , save_icon = None, close_icon = None
                                , up_arrow_icon = None, down_arrow_icon = None
                                , refresh_icon = None
                                , text_icon = None
                                , window_icon = None
                                , folder_icon = None
                                , reset_icon = None
                                , clipboard_icon = None)

        for key, item in gui_graphic.items():
            if key in self.gui_graphic:
                self.gui_graphic[key] = item


        self.save_icon          = to_tk_img(pil_img_resize(self.gui_graphic['save_icon'], img_width = 20, img_height = 20))
        self.close_icon         = to_tk_img(pil_img_resize(self.gui_graphic['close_icon'], img_width = 20, img_height = 20))
        self.refresh_icon       = to_tk_img(pil_img_resize(self.gui_graphic['refresh_icon'], img_width = 25, img_height = 25))
        self.folder_icon        = to_tk_img(pil_img_resize(self.gui_graphic['folder_icon'], img_width = 25, img_height = 25))
        self.text_icon          = to_tk_img(pil_img_resize(self.gui_graphic['text_icon'], img_width = 20, img_height = 20))
        self.up_arrow_icon      = to_tk_img(pil_img_resize(self.gui_graphic['up_arrow_icon'], img_width = 20, img_height = 20))
        self.down_arrow_icon    = to_tk_img(pil_img_resize(self.gui_graphic['down_arrow_icon'], img_width = 20, img_height = 20))
        self.window_icon        = self.gui_graphic['window_icon']

        self.toggle_ON_btn_img  = to_tk_img(pil_img_resize(self.gui_graphic['toggle_ON_btn_img'], img_scale = 0.06))
        self.toggle_OFF_btn_img = to_tk_img(pil_img_resize(self.gui_graphic['toggle_OFF_btn_img'], img_scale = 0.06))
        self.reset_icon         = to_tk_img(pil_img_resize(self.gui_graphic['reset_icon'], img_width = 25, img_height = 25))
        self.clipboard_icon     = to_tk_img(pil_img_resize(self.gui_graphic['clipboard_icon'], img_width = 25, img_height = 25))


        self.thread_handle = None
        self.thread_event = threading.Event() #During Report Generation/Update, we can use this event flag to track the process & interrupt specific Tk Event(s).
        self.thread_event.clear()

        self.err_msgbox_handle_1 = None #Testing Criteria must be filled
        self.err_msgbox_flag_1 = False

        self.err_msgbox_handle_2 = None #File is Currently Opened in Microsoft Excel
        self.err_msgbox_flag_2 = False

        self.err_msgbox_handle_3 = None #Exception Error in the Algorithm
        self.err_msgbox_flag_3 = False
        self.__exception_msg = ''

        self.info_msgbox_handle = None
        self.info_msgbox_flag = False

        self.btn_blink_handle = None
        self.__btn_blink_state = False

        self.__curr_imtext_popout = None


        self.__edit_bool = False #To trace whether or not any data is being editted
        
        self.__im_curr_dir = os.path.join(os.environ['USERPROFILE'],  "TMS_Saved_Images")
        self.__report_curr_dir = os.path.join(os.environ['USERPROFILE'], "TMS_Saved_Reports")
        self.__csv_curr_dir = os.environ['USERPROFILE']

        self.xl_class = XL_WorkBook()
        self.xl_worksheet_name = 'Sheet'
        self.xl_load_bool = False
        self.xl_load_path = None
        self.xl_img_path = None

        self.edit_data_dict = {} #We Separate Text Data and Image Data... so Sample Detail & Sample Image is for sub-category 'Photo Sample'.
        """ Argument(s)/Parameter(s) for self.edit_data_dict:
            tuple_data[0]: dictionary key; must be a string type data
            tuple_data[1]: tracker bool arr id num; must be an integer type data where n > 0, self.track_edit_bool_arr
            tuple_data[2]: size of edit boolean array; must be an integer type data which is > 0, this array is used to check edit status for each sub-categories
            tuple_data[3]: edit data arr/list; must be an numpy-array type or list type data """

        edit_subcategory_list = [("Sample Detail", 0, 2, self.create_ref_edit_arr(2))
        , ("Sample Image", 0, 1, []), ("Lighting Drawing", 1, 1, [])
        , ("Testing Criteria", 2, 6, self.create_ref_edit_arr(6)), ("Lighting Details", 3, 4, self.create_ref_edit_arr(4))
        , ("Controller Details", 4, 4, self.create_ref_edit_arr(4)), ("Camera Details", 5, 10, self.create_ref_edit_arr(10))
        , ("Lens Details", 6, 6, self.create_ref_edit_arr(6)), ("Target Image", 7, 1, [])
        , ("Grayscale Image", 8, 1, []), ("Binary Image", 9, 1, []), ("Setup Image", 10, 1, [])]

        for tuple_data in edit_subcategory_list:
            self.create_edit_bool_dict(self.edit_data_dict, tuple_data = tuple_data)

        # print(self.edit_data_dict)

        self.track_edit_bool_arr = np.zeros((int(len(self.edit_data_dict) - 1),), dtype = bool) #We Minus 1 because 'Sample Detail' & 'Sample Image' shares the same tracker id
        # print(self.track_edit_bool_arr)
        self.loading_display = tk.Frame(self, bg = 'SystemButtonFace') #'SystemButtonFace' # 'DodgerBlue2'
        self.loading_tk_var = tk.StringVar()
        loading_tk_lb = tk.Label(self.loading_display, textvariable = self.loading_tk_var, bg = 'SystemButtonFace', font = 'Helvetica 16')
        loading_tk_lb.place(relx = 0.5, rely = 0.5, anchor = 'c')

        self.loading_tk_var.set('Processing')
        self.loading_display.place(x=0, y = 0, relx = 0, rely = 0, relheight = 1, relwidth = 0.55, width = -80, height = -15, anchor = 'nw')
        self.loading_display.place_forget()
        # self.loading_display.lower() ## opposite of .lower() is .lift()
        self.__loading_handle = None

        self.report_detail_panel()
        self.report_ctrl_panel()

    def create_ref_edit_arr(self, arr_size):
        return np.empty((arr_size,), dtype=object)

    def create_edit_bool_dict(self, bool_dict, tuple_data = (None, None, None, None)):
        err_flag = False
        if type(tuple_data) == tuple and len(tuple_data) == 4:
            try:
                id_num = int(tuple_data[1])
                if id_num < 0:
                    err_flag = True
            except Exception:
                err_flag = True

            try:
                arr_size = int(tuple_data[2])
                if arr_size < 0:
                    err_flag = True
            except Exception:
                err_flag = True

            if (isinstance(tuple_data[3], np.ndarray) == True) or type(tuple_data[3]) == list:
                pass
            else:
                err_flag = True

            if err_flag == False:
                dict_data_arr = np.empty((3,), dtype = object)
                dict_key = tuple_data[0]
                dict_data_arr[0] = tuple_data[1]
                dict_data_arr[1] = np.zeros((tuple_data[2],), dtype = bool)
                dict_data_arr[2] = tuple_data[3]
                bool_dict[dict_key] = dict_data_arr

            elif err_flag == True:
                raise AttributeError(
                      "'tuple_data' has to be a tuple with length of 4\n"
                    + "tuple_data[0]: dictionary key; must be a string type data\n"
                    + "tuple_data[1]: tracker bool arr id num; must be an integer type data where n > 0\n"
                    + "tuple_data[2]: size of edit boolean array; must be an integer type data which is > 0\n"
                    + "tuple_data[3]: edit data arr/list; must be an numpy-array type or list type data")

        else:
            err_flag = True
            raise AttributeError(
                      "'tuple_data' has to be a tuple with length of 4\n"
                    + "tuple_data[0]: dictionary key; must be a string type data\n"
                    + "tuple_data[1]: tracker bool arr id num; must be an integer type data where n > 0\n"
                    + "tuple_data[2]: size of edit boolean array; must be an integer type data where n > 0\n"
                    + "tuple_data[3]: edit data arr/list; must be an numpy-array type or list type data")

    def clear_ref_edit_list(self, *args):
        for list_item in args:
            list_item *= 0

    def clear_ref_edit_arr(self, *args):
        for arr_item in args:
            arr_item[:] = None

    def reset_edit_bool_arr(self, *args):
        for arr_item in args:
            arr_item[:] = False

    def reset_track_edit_bool_arr(self):
        self.track_edit_bool_arr[:] = False
        self.__edit_bool = np.any(self.track_edit_bool_arr)

    def update_edit_entry_data(self, edit_dict_key, new_data_list):
        if type(new_data_list) == list:
            if edit_dict_key in self.edit_data_dict:
                i = 0
                for entry_data in new_data_list:
                    self.edit_data_dict[edit_dict_key][2][i] = entry_data
                    i += 1

    def check_image_edit(self, edit_dict_key, img_data):
        ## CHECK IMAGE UPLOAD TO SEE IF IT IS EDITTED/CHANGED
        #it has to be in order
        ref_list = self.edit_data_dict[edit_dict_key][2]

        if type(ref_list) == list:
            if type(img_data) == dict or (isinstance(img_data, OrderedDict) == True):
                img_list = list(img_data.values())
                # group_print(img_list, ref_list)
                if img_list == ref_list:
                    self.edit_data_dict[edit_dict_key][1][0] = False
                else:
                    self.edit_data_dict[edit_dict_key][1][0] = True

                id_num = self.edit_data_dict[edit_dict_key][0]
                self.track_edit_bool_arr[id_num] = np.any(self.edit_data_dict[edit_dict_key][1])

            elif type(img_data) == list:
                img_list = img_data
                if img_list == ref_list:
                    self.edit_data_dict[edit_dict_key][1][0] = False
                else:
                    self.edit_data_dict[edit_dict_key][1][0] = True

                id_num = self.edit_data_dict[edit_dict_key][0]
                self.track_edit_bool_arr[id_num] = np.any(self.edit_data_dict[edit_dict_key][1])

            else:
                raise TypeError("'img_data' must be a dict-type data or list-type data.")
        else:
            raise AttributeError("data[2] in edit_data_dict is not a list-type. Please de-bug, ensure that for image edit tracking, data[2] is type-list")

        # print('check_image_edit: ', self.track_edit_bool_arr)
        self.__edit_bool = np.any(self.track_edit_bool_arr)

    def check_entry_edit(self, tk_event, ref_arr_index, tk_var, edit_dict_key = None):
        ## CHECK REPORT DETAIL ENTRIES TO SEE IF IT IS EDITTED/CHANGED
        try:
            if tk_event.keysym == 'Left' or tk_event.keysym == 'Right' or tk_event.keysym == 'Up' or tk_event.keysym == 'Down' or tk_event.keysym == 'Tab':
                return
        except Exception:
            pass

        if edit_dict_key in self.edit_data_dict:
            ref_data = self.edit_data_dict[edit_dict_key][2][ref_arr_index]
            if ref_data is not None:
                if tk_var.get() != ref_data:
                    self.edit_data_dict[edit_dict_key][1][ref_arr_index] = True

                elif tk_var.get() == ref_data:
                    self.edit_data_dict[edit_dict_key][1][ref_arr_index] = False
            else:
                if tk_var.get() == '':
                    self.edit_data_dict[edit_dict_key][1][ref_arr_index] = False
                else:
                    self.edit_data_dict[edit_dict_key][1][ref_arr_index] = True

            id_num = self.edit_data_dict[edit_dict_key][0]
            self.track_edit_bool_arr[id_num] = np.any(self.edit_data_dict[edit_dict_key][1])
            # print(self.track_edit_bool_arr, np.any(self.track_edit_bool_arr))
            # print('check_entry_edit: ', self.track_edit_bool_arr)

        self.__edit_bool = np.any(self.track_edit_bool_arr)

            # group_print(__edit_bool = self.__edit_bool)

    def check_combobox_edit(self, tk_event, ref_arr_index, tk_custombox, edit_dict_key = None):
        if edit_dict_key in self.edit_data_dict:
            ref_data = self.edit_data_dict[edit_dict_key][2][ref_arr_index]
            # print(self.edit_data_dict[edit_dict_key][2])
            if ref_data is not None:
                if tk_custombox.get() != ref_data:
                    self.edit_data_dict[edit_dict_key][1][ref_arr_index] = True

                elif tk_custombox.get() == ref_data:
                    self.edit_data_dict[edit_dict_key][1][ref_arr_index] = False
            else:
                if tk_custombox.get() == '':
                    self.edit_data_dict[edit_dict_key][1][ref_arr_index] = False
                else:
                    self.edit_data_dict[edit_dict_key][1][ref_arr_index] = True

            id_num = self.edit_data_dict[edit_dict_key][0]
            self.track_edit_bool_arr[id_num] = np.any(self.edit_data_dict[edit_dict_key][1])

        self.__edit_bool = np.any(self.track_edit_bool_arr)


    def report_mode_btn_state(self, active_button, inactive_button1 = None):
        orig_colour_bg = 'snow2'
        active_button['activeforeground'] = 'white'
        active_button['fg'] = 'white'
        active_button['activebackground'] = 'medium blue' #'blue'
        active_button['bg'] = 'medium blue' #'blue'
        active_button['font'] = 'Helvetica 11 bold'
        active_button['disabledforeground'] = 'white'
        

        if inactive_button1 != None:
            inactive_button1['activeforeground'] = 'black'
            inactive_button1['fg'] = 'black'
            inactive_button1['activebackground'] = 'DarkSlateGray1' #orig_colour_bg
            inactive_button1['bg'] = orig_colour_bg
            inactive_button1['font'] = 'Helvetica 11'
            inactive_button1['disabledforeground'] = 'gray'

    def combobox_sel_none(self, tk_combobox):
        if tk_combobox.get() == 'None':
            tk_combobox.set('')

    def custom_scroll_inner_bound(self, event, inner_scroll_class, outer_scroll_class):
        # print(event.type)
        # print(dir(event))
        # self.ivs_morph_scroll_class.canvas.bind_all("<MouseWheel>", self.custom_inner_scrolly)
        inner_scroll_class.canvas.bind_all("<MouseWheel>", lambda e: self.custom_inner_scrolly(e, inner_scroll_class, outer_scroll_class))

    def custom_inner_scrolly(self, event, inner_scroll_class, outer_scroll_class):
        # print(event , inner_scroll_class, outer_scroll_class, inner_scroll_class.scrolly_lock)

        if inner_scroll_class.scrolly_lock == False:
            inner_scroll_class.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            # print('custom_inner_scrolly: ',event.delta)
            y0_inner = float(inner_scroll_class.canvas.yview()[0])
            y1_inner = float(inner_scroll_class.canvas.yview()[1])
            # print(y0_inner, y1_inner)
            if 0 <= y1_inner < 1:
                if y0_inner == 0: #inner scroll: Start point
                    if event.delta > 0: #scroll up
                        outer_scroll_class.canvas.yview_scroll(int(-1*(event.delta/120)), "units")

            elif y1_inner == 1:
                if 0<= y0_inner < 1: #inner scroll: End point
                    if event.delta < 0: #scroll down
                        outer_scroll_class.canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        elif inner_scroll_class.scrolly_lock == True:
            outer_scroll_class.canvas.yview_scroll(int(-1*(event.delta/120)), "units")


    def report_detail_panel(self):
        self.left_main_gui = tk.Frame(self, bg = 'DodgerBlue2') #'SystemButtonFace'
        tk.Label(self.left_main_gui, text = 'Report Detail:', font = 'Helvetica 14 bold', fg = "white", bg = 'DodgerBlue2').place(x=5, y = 2)

        self.reload_worksheet_btn = tk.Button(self.left_main_gui, relief = tk.GROOVE, image = self.refresh_icon
            , activebackground = 'DarkSlateGray1')

        CreateToolTip(self.reload_worksheet_btn, 'Reload Worksheet'
            , 0, -25-2, width = 138, height = 25, font = 'Tahoma 12')

        self.reload_worksheet_btn.place(x = -25, y = 2, relx = 1, rely = 0, anchor = 'ne')
        widget_disable(self.reload_worksheet_btn)
        self.reload_worksheet_btn['command'] = self.reload_worksheet_data

        self.load_csv_btn = tk.Button(self.left_main_gui, relief = tk.GROOVE, image = self.clipboard_icon)
        CreateToolTip(self.load_csv_btn, 'Load CSV Data'
            , 0, -25-2, width = 138, height = 25, font = 'Tahoma 12')
        self.load_csv_btn['command'] = self.load_csv_data
        self.load_csv_btn.place(x = -25 - (25+15), y = 2, relx = 1, rely = 0, anchor = 'ne')

        self.clear_data_btn = tk.Button(self.left_main_gui, relief = tk.GROOVE, image = self.reset_icon)
        CreateToolTip(self.clear_data_btn, 'Clear All Data'
            , 0, -25-2, width = 138, height = 25, font = 'Tahoma 12')
        self.clear_data_btn['command'] = self.clear_btn_event
        self.clear_data_btn.place(x = -25 - (25+15) - (25+15), y = 2, relx = 1, rely = 0, anchor = 'ne')

        self.left_main_gui.place(x=0, y = 0, relx = 0, rely = 0, relheight = 1, relwidth = 0.55, width = -80, height = -15, anchor = 'nw')

        self.ctrl_scroll_class = ScrolledCanvas(self.left_main_gui, frame_w = 470, frame_h = 2000, canvas_x = 0, canvas_y = 35, bg = 'DodgerBlue2')
        self.ctrl_scroll_class.show()

        parent = self.ctrl_scroll_class.window_fr

        ctrl_parent_y = 0

        #1. Photo Sample
        self.photo_sample_parent = tk.Frame(parent, bg = 'SystemButtonFace')
        self.photo_sample_parent['height'] = 230
        self.photo_sample_parent.place(x=0, y=ctrl_parent_y, relx=0, rely=0, relwidth = 1, anchor = 'nw')

        sub_parent_y = 0

        self.photo_sample_tk_lb = tk.Label(self.photo_sample_parent, text = '1. Photo Sample', font = 'Helvetica 13', bg = 'SystemButtonFace')
        self.photo_sample_tk_lb.place(x=0, y=0, relx=0, rely=0, anchor = 'nw')

        self.photo_sample_tk_lb.update_idletasks()
        sub_parent_y = sub_parent_y + self.photo_sample_tk_lb.winfo_height() + 5

        self.photo_dim_tk_lb = tk.Label(self.photo_sample_parent, text = 'Dimension(mm):', font = 'Helvetica 12', bg = 'SystemButtonFace')
        self.photo_dim_tk_lb.place(x=10, y=sub_parent_y, relx=0, rely=0, anchor = 'nw')
        self.photo_dim_var = tk.StringVar()
        self.photo_dim_entry = CustomSingleText(self.photo_sample_parent, relief = tk.GROOVE, font  = 'Helvetica 12'
            , highlightbackground="black", highlightthickness=1, width = 20, height = 1, wrap = tk.NONE, undo = True
            , autoseparators=True, maxundo=-1, textvariable = self.photo_dim_var)

        self.photo_dim_entry.bind("<KeyRelease>", partial(self.check_entry_edit, ref_arr_index = 0, tk_var = self.photo_dim_var
            , edit_dict_key = "Sample Detail"))

        self.photo_dim_tk_lb.update_idletasks()
        self.photo_dim_entry.place(x= self.photo_dim_tk_lb.winfo_width() + 15, y = sub_parent_y, relx = 0, rely = 0, anchor = 'nw')

        sub_parent_y = sub_parent_y + self.photo_dim_tk_lb.winfo_height() + 5

        self.photo_bg_tk_lb = tk.Label(self.photo_sample_parent, text = 'Background:', font = 'Helvetica 12', bg = 'SystemButtonFace')
        self.photo_bg_tk_lb.place(x=10, y = sub_parent_y, relx=0, rely=0, anchor = 'nw')

        self.photo_bg_tk_lb.update_idletasks()

        self.photo_bg_combobox = CustomBox(self.photo_sample_parent, width=20, state='readonly', font = 'Helvetica 12')
        self.photo_bg_list = ['', 'White', 'Black']
        self.photo_bg_combobox['value'] = self.photo_bg_list
        self.photo_bg_combobox.bind("<<ComboboxSelected>>", partial(self.check_combobox_edit, ref_arr_index = 1, tk_custombox = self.photo_bg_combobox
            , edit_dict_key = "Sample Detail"))

        #lambda event: self.worksheet_name_select(event))
        # self.photo_bg_combobox.current(0)
        self.photo_bg_combobox.place(x= self.photo_bg_tk_lb.winfo_width() + 15, y = sub_parent_y, relx = 0, rely = 0, anchor = 'nw')

        sub_parent_y = sub_parent_y + self.photo_bg_tk_lb.winfo_height() + 5 + 10

        self.photo_sample_entry_list = list((self.photo_dim_entry, self.photo_bg_combobox))

        self.photo_upload_btn = tk.Button(self.photo_sample_parent, relief = tk.GROOVE, text = 'Upload', font = 'Helvetica 12'
            , activebackground = 'DarkSlateGray1')

        self.photo_upload_btn.place(x=50, y=sub_parent_y, relx=0, rely=0, anchor = 'nw')

        self.photo_upload_tk_fr = tk.Frame(self.photo_sample_parent, bg = 'SystemButtonFace')
        # self.photo_upload_tk_fr['width'] = 300
        self.photo_upload_tk_fr['height'] = 150+15
        # self.photo_upload_tk_fr.place(x=120, y = 55, relx = 0, rely = 0, anchor = 'nw')
        self.photo_upload_tk_fr.place(x=120, y = sub_parent_y, relx = 0, rely = 0, relwidth = 1, width = -170, anchor = 'nw')

        self.photo_upload_log_dict = OrderedDict() #{}
        self.photo_upload_img_dict = OrderedDict()

        # print(isinstance(self.photo_upload_img_dict, OrderedDict))

        self.photo_upload_col_tracker = [] 
        #To check track the column length (which row has the longest length) to implement adaptive upload log window size.
        self.photo_upload_log_class = ScrolledCanvas(self.photo_upload_tk_fr, frame_w = 300, frame_h = 150, canvas_x = 0, canvas_y = 0, bg = 'DarkSlateGray3')
        self.photo_upload_log_class.show()

        self.photo_upload_log_class.canvas.bind('<Enter>', lambda e: self.custom_scroll_inner_bound(e, self.photo_upload_log_class, self.ctrl_scroll_class))
        self.photo_upload_log_class.canvas.bind('<Leave>', self.ctrl_scroll_class._bound_to_mousewheel)

        self.photo_upload_log = self.photo_upload_log_class.window_fr
        self.photo_upload_btn['command'] = lambda : self.upload_btn_callback(self.photo_upload_log, self.photo_upload_log_dict, self.photo_upload_img_dict
                , self.photo_upload_log_class, self.photo_upload_col_tracker
                , edit_dict_key = "Sample Image"
                , root_path = self.__im_curr_dir)

        sub_parent_y = sub_parent_y + self.photo_upload_tk_fr['height'] + 5

        self.photo_sample_parent['height'] = sub_parent_y

        ctrl_parent_y = ctrl_parent_y + self.photo_sample_parent['height'] + 5
        ##########################################################################################################################################################################################
        self.light_setup_parent = tk.Frame(parent, bg = 'SystemButtonFace')

        self.light_setup_parent['height'] = 200
        self.light_setup_parent.place(x=0, y=ctrl_parent_y, relx=0, rely=0, relwidth = 1, anchor = 'nw')
        self.light_setup_tk_lb = tk.Label(self.light_setup_parent, text = '2. Lighting Drawing', font = 'Helvetica 13', bg = 'SystemButtonFace')
        self.light_setup_tk_lb.place(x=0, y=0, relx=0, rely=0, anchor = 'nw')

        self.light_setup_upload_btn = tk.Button(self.light_setup_parent, relief = tk.GROOVE, text = 'Upload', font = 'Helvetica 12'
            , activebackground = 'DarkSlateGray1')
        self.light_setup_upload_btn.place(x=50, y=30, relx=0, rely=0, anchor = 'nw')

        self.light_setup_tk_fr = tk.Frame(self.light_setup_parent, bg = 'SystemButtonFace')
        self.light_setup_tk_fr['height'] = 150+15
        self.light_setup_tk_fr.place(x=120, y = 30, relx = 0, rely = 0, relwidth = 1, width = -170, anchor = 'nw')

        self.light_setup_log_dict = OrderedDict() #{}
        self.light_setup_img_dict = OrderedDict() #{} #Light Drawing Information Images

        self.light_setup_col_tracker = [] 
        #To check track the column length (which row has the longest length) to implement adaptive upload log window size.
        self.light_setup_log_class = ScrolledCanvas(self.light_setup_tk_fr, frame_w = 300, frame_h = 150, canvas_x = 0, canvas_y = 0, bg = 'DarkSlateGray3')
        self.light_setup_log_class.show()

        self.light_setup_log_class.canvas.bind('<Enter>', lambda e: self.custom_scroll_inner_bound(e, self.light_setup_log_class, self.ctrl_scroll_class))
        self.light_setup_log_class.canvas.bind('<Leave>', self.ctrl_scroll_class._bound_to_mousewheel)

        self.light_setup_log = self.light_setup_log_class.window_fr
        self.light_setup_upload_btn['command'] = lambda : self.upload_btn_callback(self.light_setup_log, self.light_setup_log_dict, self.light_setup_img_dict
                , self.light_setup_log_class, self.light_setup_col_tracker
                , edit_dict_key = "Lighting Drawing"
                # , root_path = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__))) + '\\report src'
                , root_path = os.path.dirname(os.path.dirname(__file__)) + '\\report src'
                , track_dir_bool = False)

        ctrl_parent_y = ctrl_parent_y + self.light_setup_parent['height'] + 5
        ##########################################################################################################################################################################################
        self.test_criteria_parent = tk.Frame(parent, bg = 'SystemButtonFace')
        # self.test_criteria_parent['height'] = 210
        # self.test_criteria_parent.place(x=0, y=225 + 105, relx=0, rely=0, relwidth = 1, anchor = 'nw')
        self.test_criteria_parent.place(x=0, y=ctrl_parent_y, relx=0, rely=0, relwidth = 1, anchor = 'nw')
        self.test_criteria_tk_lb = tk.Label(self.test_criteria_parent, text = '3. Testing Criteria', font = 'Helvetica 13', bg = 'SystemButtonFace')
        self.test_criteria_tk_lb.place(x=0, y=0, relx=0, rely=0, anchor = 'nw')


        lb_name_list = ['Light Dimension (mm):', 'Lighting Thickness(mm):', 'Lighting Inner Diameter(mm):'
        , 'Lighting Working Distance(mm):', 'Lens Working Distance(mm):', 'Field of View(mm x mm):']
        tk_lb_list, tk_entry_list, self.test_criteria_parent['height'] = self.entry_and_lb_gen(self.test_criteria_parent, lb_name_list, start_y = 30
            , edit_dict_key = "Testing Criteria")
        # print(tk_lb_list, tk_entry_list)

        self.light_dim_tk_lb = tk_lb_list[0]
        self.light_dim_tk_var, self.light_dim_entry = tk_entry_list[0][0], tk_entry_list[0][1]

        self.light_thickness_tk_lb = tk_lb_list[1]
        self.light_thickness_tk_var, self.light_thickness_entry = tk_entry_list[1][0], tk_entry_list[1][1]

        self.light_inn_diam_tk_lb = tk_lb_list[2]
        self.light_inn_diam_tk_var, self.light_inn_diam_entry = tk_entry_list[2][0], tk_entry_list[2][1]

        self.light_wd_tk_lb = tk_lb_list[3]
        self.light_wd_tk_var, self.light_wd_entry = tk_entry_list[3][0], tk_entry_list[3][1]

        self.lens_wd_tk_lb = tk_lb_list[4]
        self.lens_wd_tk_var, self.lens_wd_entry = tk_entry_list[4][0], tk_entry_list[4][1]

        self.test_fov_tk_lb = tk_lb_list[5]
        self.test_fov_tk_var, self.test_fov_entry = tk_entry_list[5][0], tk_entry_list[5][1]

        self.test_criteria_var_list = list((self.light_dim_tk_var, self.light_thickness_tk_var, self.light_inn_diam_tk_var
            , self.light_wd_tk_var, self.lens_wd_tk_var, self.test_fov_tk_var))
        self.test_criteria_entry_list = list((self.light_dim_entry, self.light_thickness_entry, self.light_inn_diam_entry

            , self.light_wd_entry, self.lens_wd_entry, self.test_fov_entry))

        del lb_name_list, tk_lb_list, tk_entry_list

        ctrl_parent_y = ctrl_parent_y + self.test_criteria_parent['height'] + 5
        ##########################################################################################################################################################################################
        self.light_detail_parent = tk.Frame(parent, bg = 'SystemButtonFace')
        # self.light_detail_parent['height'] = 210
        self.light_detail_parent.place(x=0, y=ctrl_parent_y, relx=0, rely=0, relwidth = 1, anchor = 'nw')
        self.light_detail_tk_lb = tk.Label(self.light_detail_parent, text = '4. Lighting Details', font = 'Helvetica 13', bg = 'SystemButtonFace')
        self.light_detail_tk_lb.place(x=0, y=0, relx=0, rely=0, anchor = 'nw')

        lb_name_list = ['Lighting Model:', 'Lighting Color:'
        , 'Accessories\n(Polarizer/Diffuser/Extension\nCable/Mounting Bracket):'
        , 'Lighting Voltage/Current (mA):']
        tk_lb_list, tk_entry_list, self.light_detail_parent['height'] = self.entry_and_lb_gen(self.light_detail_parent, lb_name_list, start_y = 30
            , edit_dict_key = "Lighting Details")

        self.light_model_tk_lb = tk_lb_list[0]
        self.light_model_tk_var, self.light_model_entry = tk_entry_list[0][0], tk_entry_list[0][1]

        self.light_color_tk_lb = tk_lb_list[1]
        self.light_color_tk_var, self.light_color_entry = tk_entry_list[1][0], tk_entry_list[1][1]

        self.light_accessory_tk_lb = tk_lb_list[2]
        self.light_accessory_tk_var, self.light_accessory_entry = tk_entry_list[2][0], tk_entry_list[2][1]

        self.light_voltage_tk_lb = tk_lb_list[3]
        self.light_voltage_tk_var, self.light_voltage_entry = tk_entry_list[3][0], tk_entry_list[3][1]
        
        self.light_detail_var_list = list((self.light_model_tk_var, self.light_color_tk_var, self.light_accessory_tk_var, self.light_voltage_tk_var))
        self.light_detail_entry_list = list((self.light_model_entry, self.light_color_entry, self.light_accessory_entry, self.light_voltage_entry))

        del lb_name_list, tk_lb_list, tk_entry_list

        ctrl_parent_y = ctrl_parent_y + self.light_detail_parent['height'] + 5
        ##########################################################################################################################################################################################
        self.ctrl_detail_parent = tk.Frame(parent, bg = 'SystemButtonFace')

        self.ctrl_detail_parent.place(x=0, y=ctrl_parent_y, relx=0, rely=0, relwidth = 1, anchor = 'nw')
        self.ctrl_detail_tk_lb = tk.Label(self.ctrl_detail_parent, text = '5. Controller Details', font = 'Helvetica 13', bg = 'SystemButtonFace')
        self.ctrl_detail_tk_lb.place(x=0, y=0, relx=0, rely=0, anchor = 'nw')

        lb_name_list = ['Controller Model:', 'Controller Mode\n(Constant/Strobe/Trigger):'
        , 'Current Multiplier:', 'Intensity:']
        tk_lb_list, tk_entry_list, self.ctrl_detail_parent['height'] = self.entry_and_lb_gen(self.ctrl_detail_parent, lb_name_list, start_y = 30
            , edit_dict_key = "Controller Details")
        
        self.ctrl_model_tk_lb = tk_lb_list[0]
        self.ctrl_model_tk_var, self.ctrl_model_entry = tk_entry_list[0][0], tk_entry_list[0][1]

        self.ctrl_mode_type_tk_lb = tk_lb_list[1]
        self.ctrl_mode_type_tk_var, self.ctrl_mode_type_entry = tk_entry_list[1][0], tk_entry_list[1][1]

        self.ctrl_current_multi_tk_lb = tk_lb_list[2]
        self.ctrl_current_multi_tk_var, self.ctrl_current_multi_entry = tk_entry_list[2][0], tk_entry_list[2][1]

        self.ctrl_intensity_tk_lb = tk_lb_list[3]
        self.ctrl_intensity_tk_var, self.ctrl_intensity_entry = tk_entry_list[3][0], tk_entry_list[3][1]
        
        self.ctrl_detail_var_list = list((self.ctrl_model_tk_var, self.ctrl_mode_type_tk_var, self.ctrl_current_multi_tk_var, self.ctrl_intensity_tk_var))

        self.ctrl_detail_entry_list = list((self.ctrl_model_entry, self.ctrl_mode_type_entry, self.ctrl_current_multi_entry, self.ctrl_intensity_entry))

        del lb_name_list, tk_lb_list, tk_entry_list

        ctrl_parent_y = ctrl_parent_y + self.ctrl_detail_parent['height'] + 5

        ##########################################################################################################################################################################################
        self.cam_detail_parent = tk.Frame(parent, bg = 'SystemButtonFace')

        self.cam_detail_parent.place(x=0, y=ctrl_parent_y, relx=0, rely=0, relwidth = 1, anchor = 'nw')
        self.cam_detail_tk_lb = tk.Label(self.cam_detail_parent, text = '6. Camera Details', font = 'Helvetica 13', bg = 'SystemButtonFace')
        self.cam_detail_tk_lb.place(x=0, y=0, relx=0, rely=0, anchor = 'nw')

        lb_name_list = ['Camera Model:', 'Sensor Size:', 'Resolution (pixel):', 'Megapixel:'
        , 'Sensor type:', 'Pixel size (' + chr(ord('\u03BC')) +'M):', 'Shutter type:', 'Mono/Color:'
        , 'Y axis (mm):', 'X axis (mm):']
        tk_lb_list, tk_entry_list, self.cam_detail_parent['height'] = self.entry_and_lb_gen(self.cam_detail_parent, lb_name_list, start_y = 30
            , edit_dict_key = "Camera Details")
        
        self.cam_model_tk_lb = tk_lb_list[0]
        self.cam_model_tk_var, self.cam_model_entry = tk_entry_list[0][0], tk_entry_list[0][1]

        self.cam_sensor_size_tk_lb = tk_lb_list[1]
        self.cam_sensor_size_tk_var, self.cam_sensor_size_entry = tk_entry_list[1][0], tk_entry_list[1][1]

        self.cam_resolution_tk_lb = tk_lb_list[2]
        self.cam_resolution_tk_var, self.cam_resolution_entry = tk_entry_list[2][0], tk_entry_list[2][1]

        self.cam_megapix_tk_lb = tk_lb_list[3]
        self.cam_megapix_tk_var, self.cam_megapix_entry = tk_entry_list[3][0], tk_entry_list[3][1]

        self.cam_sensor_type_tk_lb = tk_lb_list[4]
        self.cam_sensor_type_tk_var, self.cam_sensor_type_entry = tk_entry_list[4][0], tk_entry_list[4][1]

        self.cam_pixel_size_tk_lb = tk_lb_list[5]
        self.cam_pixel_size_tk_var, self.cam_pixel_size_entry = tk_entry_list[5][0], tk_entry_list[5][1]

        self.cam_shutter_type_tk_lb = tk_lb_list[6]
        self.cam_shutter_type_tk_var, self.cam_shutter_type_entry = tk_entry_list[6][0], tk_entry_list[6][1]

        self.cam_color_tk_lb = tk_lb_list[7]
        self.cam_color_tk_var, self.cam_color_entry = tk_entry_list[7][0], tk_entry_list[7][1]

        self.cam_y_axis_tk_lb = tk_lb_list[8]
        self.cam_y_axis_tk_var, self.cam_y_axis_entry = tk_entry_list[8][0], tk_entry_list[8][1]

        self.cam_x_axis_tk_lb = tk_lb_list[9]
        self.cam_x_axis_tk_var, self.cam_x_axis_entry = tk_entry_list[9][0], tk_entry_list[9][1]
        
        self.cam_detail_var_list = list((self.cam_model_tk_var, self.cam_sensor_size_tk_var, self.cam_resolution_tk_var, self.cam_megapix_tk_var,
            self.cam_sensor_type_tk_var, self.cam_pixel_size_tk_var, self.cam_shutter_type_tk_var, self.cam_color_tk_var, self.cam_y_axis_tk_var, 
            self.cam_x_axis_tk_var))

        self.cam_detail_entry_list = list((self.cam_model_entry, self.cam_sensor_size_entry, self.cam_resolution_entry, self.cam_megapix_entry,
            self.cam_sensor_type_entry, self.cam_pixel_size_entry, self.cam_shutter_type_entry, self.cam_color_entry, self.cam_y_axis_entry, 
            self.cam_x_axis_entry))

        del lb_name_list, tk_lb_list, tk_entry_list

        ctrl_parent_y = ctrl_parent_y + self.cam_detail_parent['height'] + 5

        ##########################################################################################################################################################################################
        self.lens_detail_parent = tk.Frame(parent, bg = 'SystemButtonFace')

        self.lens_detail_parent.place(x=0, y=ctrl_parent_y, relx=0, rely=0, relwidth = 1, anchor = 'nw')
        self.lens_detail_tk_lb = tk.Label(self.lens_detail_parent, text = '7. Lens Details', font = 'Helvetica 13', bg = 'SystemButtonFace')
        self.lens_detail_tk_lb.place(x=0, y=0, relx=0, rely=0, anchor = 'nw')

        lb_name_list = ['Lens Model:','Focal Length (mm):', 'Mount:', 'Sensor Size (max.):'
        , 'TV distortion (%):', 'F.O.V (mm):']
        tk_lb_list, tk_entry_list, self.lens_detail_parent['height'] = self.entry_and_lb_gen(self.lens_detail_parent, lb_name_list, start_y = 30
            , edit_dict_key = "Lens Details")
        
        self.lens_model_tk_lb = tk_lb_list[0]
        self.lens_model_tk_var, self.lens_model_entry = tk_entry_list[0][0], tk_entry_list[0][1]

        self.lens_focal_length_tk_lb = tk_lb_list[1]
        self.lens_focal_length_tk_var, self.lens_focal_length_entry = tk_entry_list[1][0], tk_entry_list[1][1]

        self.lens_mount_tk_lb = tk_lb_list[2]
        self.lens_mount_tk_var, self.lens_mount_entry = tk_entry_list[2][0], tk_entry_list[2][1]

        self.lens_sensor_size_tk_lb = tk_lb_list[3]
        self.lens_sensor_size_tk_var, self.lens_sensor_size_entry = tk_entry_list[3][0], tk_entry_list[3][1]

        self.lens_tv_distortion_tk_lb = tk_lb_list[4]
        self.lens_tv_distortion_tk_var, self.lens_tv_distortion_entry = tk_entry_list[4][0], tk_entry_list[4][1]

        self.lens_fov_tk_lb = tk_lb_list[5]
        self.lens_fov_tk_var, self.lens_fov_entry = tk_entry_list[5][0], tk_entry_list[5][1]
        
        self.lens_detail_var_list = list((self.lens_model_tk_var, self.lens_focal_length_tk_var, self.lens_mount_tk_var
            , self.lens_sensor_size_tk_var, self.lens_tv_distortion_tk_var, self.lens_fov_tk_var))

        self.lens_detail_entry_list = list((self.lens_model_entry, self.lens_focal_length_entry, self.lens_mount_entry
            , self.lens_sensor_size_entry, self.lens_tv_distortion_entry, self.lens_fov_entry))

        del lb_name_list, tk_lb_list, tk_entry_list

        ctrl_parent_y = ctrl_parent_y + self.lens_detail_parent['height'] + 5

        ##########################################################################################################################################################################################
        self.target_sample_parent = tk.Frame(parent, bg = 'SystemButtonFace')
        self.target_sample_parent['height'] = 200
        self.target_sample_parent.place(x=0, y=ctrl_parent_y, relx=0, rely=0, relwidth = 1, anchor = 'nw')

        self.target_sample_tk_lb = tk.Label(self.target_sample_parent, text = '8. Targeted Image', font = 'Helvetica 13', bg = 'SystemButtonFace')
        self.target_sample_tk_lb.place(x=0, y=0, relx=0, rely=0, anchor = 'nw')
        
        self.target_upload_btn = tk.Button(self.target_sample_parent, relief = tk.GROOVE, text = 'Upload', font = 'Helvetica 12'
            , activebackground = 'DarkSlateGray1')

        self.target_upload_btn.place(x=50, y=30, relx=0, rely=0, anchor = 'nw')

        self.target_upload_tk_fr = tk.Frame(self.target_sample_parent, bg = 'SystemButtonFace')
        # self.target_upload_tk_fr['width'] = 300
        self.target_upload_tk_fr['height'] = 150+15
        # self.target_upload_tk_fr.place(x=120, y = 55, relx = 0, rely = 0, anchor = 'nw')
        self.target_upload_tk_fr.place(x=120, y = 30, relx = 0, rely = 0, relwidth = 1, width = -170, anchor = 'nw')

        self.target_upload_log_dict = OrderedDict() #{}
        self.target_upload_img_dict = OrderedDict() #{}

        self.target_upload_col_tracker = [] 
        #To check track the column length (which row has the longest length) to implement adaptive upload log window size.
        self.target_upload_log_class = ScrolledCanvas(self.target_upload_tk_fr, frame_w = 300, frame_h = 150, canvas_x = 0, canvas_y = 0, bg = 'DarkSlateGray3')
        self.target_upload_log_class.show()

        self.target_upload_log_class.canvas.bind('<Enter>', lambda e: self.custom_scroll_inner_bound(e, self.target_upload_log_class, self.ctrl_scroll_class))
        self.target_upload_log_class.canvas.bind('<Leave>', self.ctrl_scroll_class._bound_to_mousewheel)

        self.target_upload_log = self.target_upload_log_class.window_fr
        self.target_upload_btn['command'] = lambda : self.upload_btn_callback(self.target_upload_log, self.target_upload_log_dict, self.target_upload_img_dict
                , self.target_upload_log_class, self.target_upload_col_tracker
                , edit_dict_key = "Target Image"
                , root_path = self.__im_curr_dir)

        ctrl_parent_y = ctrl_parent_y + self.target_sample_parent['height'] + 5

        ##########################################################################################################################################################################################
        self.gray_sample_parent = tk.Frame(parent, bg = 'SystemButtonFace')
        self.gray_sample_parent['height'] = 200
        self.gray_sample_parent.place(x=0, y=ctrl_parent_y, relx=0, rely=0, relwidth = 1, anchor = 'nw')

        self.gray_sample_tk_lb = tk.Label(self.gray_sample_parent, text = '9. Grayscale Image', font = 'Helvetica 13', bg = 'SystemButtonFace')
        self.gray_sample_tk_lb.place(x=0, y=0, relx=0, rely=0, anchor = 'nw')
        
        self.gray_upload_btn = tk.Button(self.gray_sample_parent, relief = tk.GROOVE, text = 'Upload', font = 'Helvetica 12'
            , activebackground = 'DarkSlateGray1')

        self.gray_upload_btn.place(x=50, y=30, relx=0, rely=0, anchor = 'nw')

        self.gray_upload_tk_fr = tk.Frame(self.gray_sample_parent, bg = 'SystemButtonFace')
        # self.gray_upload_tk_fr['width'] = 300
        self.gray_upload_tk_fr['height'] = 150+15
        # self.gray_upload_tk_fr.place(x=120, y = 55, relx = 0, rely = 0, anchor = 'nw')
        self.gray_upload_tk_fr.place(x=120, y = 30, relx = 0, rely = 0, relwidth = 1, width = -170, anchor = 'nw')

        self.gray_upload_log_dict = OrderedDict() #{}
        self.gray_upload_img_dict = OrderedDict() #{}

        self.gray_upload_col_tracker = [] 
        #To check track the column length (which row has the longest length) to implement adaptive upload log window size.
        self.gray_upload_log_class = ScrolledCanvas(self.gray_upload_tk_fr, frame_w = 300, frame_h = 150, canvas_x = 0, canvas_y = 0, bg = 'DarkSlateGray3')
        self.gray_upload_log_class.show()

        self.gray_upload_log_class.canvas.bind('<Enter>', lambda e: self.custom_scroll_inner_bound(e, self.gray_upload_log_class, self.ctrl_scroll_class))
        self.gray_upload_log_class.canvas.bind('<Leave>', self.ctrl_scroll_class._bound_to_mousewheel)

        self.gray_upload_log = self.gray_upload_log_class.window_fr
        self.gray_upload_btn['command'] = lambda : self.upload_btn_callback(self.gray_upload_log, self.gray_upload_log_dict, self.gray_upload_img_dict
                , self.gray_upload_log_class, self.gray_upload_col_tracker
                , edit_dict_key = "Grayscale Image"
                , root_path = self.__im_curr_dir)

        ctrl_parent_y = ctrl_parent_y + self.gray_sample_parent['height'] + 5

        ##########################################################################################################################################################################################
        self.binary_sample_parent = tk.Frame(parent, bg = 'SystemButtonFace')
        self.binary_sample_parent['height'] = 200
        self.binary_sample_parent.place(x=0, y=ctrl_parent_y, relx=0, rely=0, relwidth = 1, anchor = 'nw')

        self.binary_sample_tk_lb = tk.Label(self.binary_sample_parent, text = '10. Binary Image', font = 'Helvetica 13', bg = 'SystemButtonFace')
        self.binary_sample_tk_lb.place(x=0, y=0, relx=0, rely=0, anchor = 'nw')
        
        self.binary_upload_btn = tk.Button(self.binary_sample_parent, relief = tk.GROOVE, text = 'Upload', font = 'Helvetica 12'
            , activebackground = 'DarkSlateGray1')

        self.binary_upload_btn.place(x=50, y=30, relx=0, rely=0, anchor = 'nw')

        self.binary_upload_tk_fr = tk.Frame(self.binary_sample_parent, bg = 'SystemButtonFace')
        # self.binary_upload_tk_fr['width'] = 300
        self.binary_upload_tk_fr['height'] = 150+15
        # self.binary_upload_tk_fr.place(x=120, y = 55, relx = 0, rely = 0, anchor = 'nw')
        self.binary_upload_tk_fr.place(x=120, y = 30, relx = 0, rely = 0, relwidth = 1, width = -170, anchor = 'nw')

        self.binary_upload_log_dict = OrderedDict() #{}
        self.binary_upload_img_dict = OrderedDict() #{}
        self.binary_upload_imgtxt_dict = {}

        self.binary_upload_col_tracker = [] 
        #To check track the column length (which row has the longest length) to implement adaptive upload log window size.
        self.binary_upload_log_class = ScrolledCanvas(self.binary_upload_tk_fr, frame_w = 300, frame_h = 150, canvas_x = 0, canvas_y = 0, bg = 'DarkSlateGray3')
        self.binary_upload_log_class.show()

        self.binary_upload_log_class.canvas.bind('<Enter>', lambda e: self.custom_scroll_inner_bound(e, self.binary_upload_log_class, self.ctrl_scroll_class))
        self.binary_upload_log_class.canvas.bind('<Leave>', self.ctrl_scroll_class._bound_to_mousewheel)

        self.binary_upload_log = self.binary_upload_log_class.window_fr
        self.binary_upload_btn['command'] = lambda : self.upload_btn_callback(self.binary_upload_log, self.binary_upload_log_dict, self.binary_upload_img_dict
                , self.binary_upload_log_class, self.binary_upload_col_tracker
                , edit_dict_key = "Binary Image"
                , root_path = self.__im_curr_dir)

        ctrl_parent_y = ctrl_parent_y + self.binary_sample_parent['height'] + 5

        ##########################################################################################################################################################################################
        self.setup_sample_parent = tk.Frame(parent, bg = 'SystemButtonFace')
        self.setup_sample_parent['height'] = 200
        self.setup_sample_parent.place(x=0, y=ctrl_parent_y, relx=0, rely=0, relwidth = 1, anchor = 'nw')

        self.setup_sample_tk_lb = tk.Label(self.setup_sample_parent, text = '11. Setup Image', font = 'Helvetica 13', bg = 'SystemButtonFace')
        self.setup_sample_tk_lb.place(x=0, y=0, relx=0, rely=0, anchor = 'nw')
        
        self.setup_upload_btn = tk.Button(self.setup_sample_parent, relief = tk.GROOVE, text = 'Upload', font = 'Helvetica 12'
            , activebackground = 'DarkSlateGray1')

        self.setup_upload_btn.place(x=50, y=30, relx=0, rely=0, anchor = 'nw')

        self.setup_upload_tk_fr = tk.Frame(self.setup_sample_parent, bg = 'SystemButtonFace')
        self.setup_upload_tk_fr['height'] = 150+15
        self.setup_upload_tk_fr.place(x=120, y = 30, relx = 0, rely = 0, relwidth = 1, width = -170, anchor = 'nw')

        self.setup_upload_log_dict = OrderedDict() #{}
        self.setup_upload_img_dict = OrderedDict() #{}

        self.setup_upload_col_tracker = [] 
        #To check track the column length (which row has the longest length) to implement adaptive upload log window size.
        self.setup_upload_log_class = ScrolledCanvas(self.setup_upload_tk_fr, frame_w = 300, frame_h = 150, canvas_x = 0, canvas_y = 0, bg = 'DarkSlateGray3')
        self.setup_upload_log_class.show()

        self.setup_upload_log_class.canvas.bind('<Enter>', lambda e: self.custom_scroll_inner_bound(e, self.setup_upload_log_class, self.ctrl_scroll_class))
        self.setup_upload_log_class.canvas.bind('<Leave>', self.ctrl_scroll_class._bound_to_mousewheel)

        self.setup_upload_log = self.setup_upload_log_class.window_fr
        self.setup_upload_btn['command'] = lambda : self.upload_btn_callback(self.setup_upload_log, self.setup_upload_log_dict, self.setup_upload_img_dict
                , self.setup_upload_log_class, self.setup_upload_col_tracker
                , edit_dict_key = "Setup Image"
                , root_path = self.__im_curr_dir)

        ctrl_parent_y = ctrl_parent_y + self.setup_sample_parent['height'] + 5
        
        self.ctrl_scroll_class.resize_frame(height = ctrl_parent_y + 10)


    def entry_and_lb_gen(self, parent, lb_name_list = [], start_x = 0, start_y = 0, lb_bg = 'SystemButtonFace', edit_dict_key = None):
        return_lb_list = []
        return_entry_list = []

        resultant_height = 0

        if type(lb_name_list) == list:
            i = 0
            curr_tk_lb_height = 0

            for lb_name in lb_name_list:
                if type(lb_name) == str:
                    lb_place_x = start_x + 10

                    lb_place_y = start_y + curr_tk_lb_height + np.multiply(i, 5)

                    tk_lb = tk.Label(parent, text = lb_name, font = 'Helvetica 13', bg = lb_bg, justify = tk.LEFT)
                    tk_lb.place(x=lb_place_x, y=lb_place_y, relx=0, rely=0, anchor = 'nw')

                    tk_lb.update_idletasks()
                    # print(tk_lb.winfo_width(), tk_lb.winfo_height())

                    entry_place_x = start_x + int(tk_lb.winfo_width()) + 15

                    entry_place_y = start_y + curr_tk_lb_height + np.multiply(i, 5) +  np.divide(tk_lb.winfo_height(),2) #anchor = 'w', WEST

                    curr_tk_lb_height = curr_tk_lb_height + tk_lb.winfo_height()

                    tk_var = tk.StringVar()
                    tk_entry = CustomSingleText(parent, relief = tk.GROOVE, font  = 'Helvetica 12'
                        , highlightbackground="black", highlightthickness=1, width = 20, height = 1, wrap = tk.NONE, undo = True
                        , autoseparators=True, maxundo=-1, textvariable = tk_var)

                    
                    tk_entry.bind("<KeyRelease>", partial(self.check_entry_edit, ref_arr_index = i, tk_var = tk_var
                        , edit_dict_key = edit_dict_key))

                    tk_entry.place(x=entry_place_x, y=entry_place_y, relx=0, rely=0, anchor = 'w')

                    return_lb_list.append(tk_lb)
                    return_entry_list.append((tk_var, tk_entry))

                    i += 1

            resultant_height = curr_tk_lb_height + np.multiply(i, 5) + start_y
            # print('resultant_height: ', resultant_height)

        else:
            raise TypeError("User must provide 'lb_name_list', which is a list-type data, containing str-type data to generate"
                + '\n Tkinter Label with respective Entry')

        return return_lb_list, return_entry_list, resultant_height

    def report_ctrl_panel(self):
        self.right_main_gui = tk.Frame(self, bg = 'SystemButtonFace')
        # self.right_main_gui.place(x=0, y = 35, relx = 0, rely = 0, relheight = 1, relwidth = 1, width = -(300 +150) - 80, height = -35 -15, anchor = 'nw')
        # self.right_main_gui.place(x=-70, y = 35, relx = 0.55, rely = 0, relheight = 1, relwidth = 0.45, width = 50, height = -35 -15, anchor = 'nw')
        self.right_main_gui.place(x=-75, y = 5, relx = 0.55, rely = 0, relheight = 1, relwidth = 0.45, width = 50, height = -5 - 15, anchor = 'nw')

        gen_report_tk_fr = tk.Frame(self.right_main_gui, bg = 'DodgerBlue2')
        event_log_tk_fr = tk.Frame(self.right_main_gui, bg = 'SystemButtonFace')

        # gen_report_tk_fr.place(x=0, y = 0, relx = 0, rely = 0, relheight = 1, relwidth = 1, width = 0, height = -255, anchor = 'nw')
        # event_log_tk_fr.place(x=0, y = -255, relx = 0, rely = 1, relheight = 0, relwidth = 1, width = 0, height = 255, anchor = 'nw') #since the height of event log is 220 and placed at y = 30
        
        gen_report_tk_fr.place(x=0, y = 0, relx = 0, rely = 0, relheight = 0.5, relwidth = 1, width = 0, height = -6, anchor = 'nw')
        event_log_tk_fr.place(x=0, y = -2, relx = 0, rely = 0.5, relheight = 0.5, relwidth = 1, width = 0, height = -2, anchor = 'nw') 

        parent1 = gen_report_tk_fr
        parent2 = event_log_tk_fr

        tk.Label(parent1, text = 'Generate Report:', font = 'Helvetica 14 bold', bg = 'DodgerBlue2', fg = 'white').place(x=10, y = 0)

        tk.Label(parent2, text = 'Event Logs:', font = 'Helvetica 12 italic').place(x=0, y =0)
        self.event_log_txtbox = Text_Scroll(parent2, scroll_canvas_class = self.scroll_canvas_class, width = 450, height = 220)
        self.event_log_txtbox.tk_txt['state'] = 'disable'
        # self.event_log_txtbox.tk_txt['wrap'] = tk.CHAR

        # self.event_log_txtbox.place(x=0, y =30, relwidth = 1, anchor = 'nw')
        self.event_log_txtbox.place(x=0, y =30, relwidth = 1, relheight = 1, height = -30, anchor = 'nw')

        self.event_log_clr_btn = tk.Button(parent2, relief=tk.GROOVE, text = 'Clear Log', font = 'Helvetica 11'
            , activebackground = 'DarkSlateGray1')
        self.event_log_clr_btn['width'] = 10
        self.event_log_clr_btn['command'] = self.clear_event_log
        self.event_log_clr_btn.place(x=-20, y = 0, relx = 1, rely = 0, anchor = 'ne')


        self.new_report_btn = tk.Button(parent1, relief=tk.GROOVE, text = 'New Report', width = 11)
        self.new_report_btn['command'] = self.new_report_init
        self.new_report_btn.place(x=10, y = 35)
        # self.new_report_btn.place(relx=0, rely = 1, x= 10, y = -180, anchor = 'sw')

        self.load_report_btn = tk.Button(parent1, relief=tk.GROOVE, text = 'Load Report', width = 11)
        self.load_report_btn['command'] = self.load_existing_report
        self.load_report_btn.place(x=130, y = 35)
        # self.load_report_btn.place(relx=0, rely = 1, x= 130, y = -180, anchor = 'sw')

        self.folder_dir_btn = tk.Button(parent1, relief = tk.FLAT)#, relief = tk.GROOVE)#, width = 2, height = 1)
        self.folder_dir_btn['bg'] = 'gold'
        self.folder_dir_btn['command'] = partial(open_save_folder, folder_path = os.path.join(os.environ['USERPROFILE'],  "TMS_Saved_Reports"), create_bool = True)
        self.folder_dir_btn['image'] = self.folder_icon
        CreateToolTip(self.folder_dir_btn, 'Open Save Folder'
            , 32, -5, font = 'Tahoma 11')
        self.folder_dir_btn.place(x=250, y = 35)

        self.report_mode_btn_state(self.new_report_btn, self.load_report_btn)

        self.curr_report_file_tk_var = tk.StringVar()
        tk.Label(parent1, textvariable = self.curr_report_file_tk_var, font = 'Helvetica 10 bold', bg = 'DodgerBlue2', fg = 'white').place(x=10, y = 75)
        # .place(relx=0, rely = 1, x= 10, y = -155, anchor = 'sw')
        # .place(x=10, y = 70)
        self.curr_report_file_tk_var.set('Current Report File: ' + (str(self.xl_load_path).split('\\'))[-1])

        self.curr_worksheet_tk_var = tk.StringVar()
        tk.Label(parent1, textvariable = self.curr_worksheet_tk_var, font = 'Helvetica 10 bold', bg = 'DodgerBlue2', fg = 'white').place(x=10, y = 105)
        # place(relx=0, rely = 1, x= 10, y = -125, anchor = 'sw')
        #.place(x=10, y = 100)
        self.curr_worksheet_tk_var.set('Current Worksheet: ' + 'None')

        tk.Label(parent1, text = 'Worksheet Name:', font = 'Helvetica 11 bold', bg = 'DodgerBlue2', fg = 'white').place(relx=0, rely = 0, x= 10, y = 145, anchor = 'nw')

        self.worksheet_name_dropbox = CustomBox(parent1, width=20, state='readonly', font = 'Helvetica 11')
        self.worksheet_name_dropbox.bind("<<ComboboxSelected>>", lambda event: self.worksheet_name_select(event))
        self.worksheet_name_dropbox.place(relx=0, rely = 0, x= 145, y = 146, anchor = 'nw')

        self.worksheet_name_var = tk.StringVar()
        # self.worksheet_name_entry = tk.Entry(parent1, textvariable = self.worksheet_name_var, relief = tk.GROOVE, font  = 'Helvetica 12'
        #     , highlightbackground="black", highlightthickness=1, width = 20)

        self.worksheet_name_entry = CustomSingleText(parent1, relief = tk.GROOVE, font  = 'Helvetica 12'
            , highlightbackground="black", highlightthickness=1, width = 20, height = 1, wrap = tk.NONE, undo = True
            , autoseparators=True, maxundo=-1, textvariable = self.worksheet_name_var)

        self.worksheet_name_var.trace("w", lambda *args: character_limit(self.worksheet_name_var, 31))
        # self.worksheet_name_var.set(self.xl_worksheet_name)
        self.worksheet_name_entry.set_text(self.xl_worksheet_name, reset_undo_stack = True)

        self.worksheet_name_entry.place(relx=0, rely = 0, x= 10, y = 175, anchor = 'nw')

        self.create_worksheet_bool = tk.IntVar(value = 1)
        self.create_worksheet_checkbtn = tk.Checkbutton(parent1, text = 'Create Worksheet', variable=self.create_worksheet_bool
            , onvalue=1, offvalue=0, font = 'Helvetica 10 bold', indicatoron = 0
            , disabledforeground = 'white'
            , bg = 'medium blue', fg = 'white', selectcolor = 'medium blue'
            , activebackground = 'medium blue'
            , image=self.toggle_OFF_btn_img, selectimage=self.toggle_ON_btn_img, compound = tk.LEFT, bd = 0)

        # self.create_worksheet_checkbtn = tk.Checkbutton(parent1, text = 'Create Worksheet', variable=self.create_worksheet_bool
        #     , onvalue=1, offvalue=0, font = 'Helvetica 10 bold', indicatoron = 0
        #     , disabledforeground = 'white'
        #     , bg = 'DodgerBlue2', fg = 'white', selectcolor = 'medium blue'
        #     , activebackground = 'DodgerBlue2')

        self.create_worksheet_checkbtn['command'] = self.enable_create_worksheet
        self.create_worksheet_checkbtn.place(relx=0, rely = 0, x= 10 + 190 + 75, y = 173, anchor = 'nw')
        # self.create_worksheet_checkbtn.place(relx=0, rely = 0, x= 10 + 190 + 75, y = 175 + 2, anchor = 'nw')

        # self.create_worksheet_checkbtn.place(relx=0, rely = 1, x= 10 + 140, y = -57, anchor = 'sw')

        self.rename_worksheet_btn = tk.Button(parent1, relief = tk.GROOVE, text = 'Rename', font = 'Helvetica 11'
            , activebackground = 'DarkSlateGray1')
        self.rename_worksheet_btn['command'] = self.rename_worksheet_func
        self.rename_worksheet_btn.place(relx=0, rely = 0, x= 10 + 190, y = 173, anchor = 'nw')

        widget_disable(self.worksheet_name_dropbox, self.rename_worksheet_btn, self.create_worksheet_checkbtn)

        self.update_report_bool = tk.IntVar(value = 1)
        self.update_report_checkbtn = tk.Checkbutton(parent1, text = 'Update Report', variable=self.update_report_bool
            , onvalue=1, offvalue=0, font = 'Helvetica 10 bold', indicatoron = 0
            , disabledforeground = 'white'
            , bg = 'medium blue', fg = 'white', selectcolor = 'medium blue'
            , activebackground = 'medium blue', activeforeground = 'white'
            , image=self.toggle_OFF_btn_img, selectimage=self.toggle_ON_btn_img, compound = tk.LEFT, bd = 0)

        # self.update_report_checkbtn = tk.Checkbutton(parent1, text = 'Update Report', variable=self.update_report_bool
        #     , onvalue=1, offvalue=0, font = 'Helvetica 10 bold', indicatoron = 0
        #     , disabledforeground = 'white'
        #     , bg = 'DodgerBlue2', fg = 'white', selectcolor = 'medium blue'
        #     , activebackground = 'DodgerBlue2', activeforeground = 'white')

        self.update_report_checkbtn['command'] = self.enable_update_generate_report
        self.update_report_checkbtn.place(relx=0, rely = 0, x= 10 + 140, y = 207, anchor = 'nw')
        # self.update_report_checkbtn.place(relx=0, rely = 0, x= 10 + 140, y = 207 + 5, anchor = 'nw')

        self.update_report_btn = tk.Button(parent1, relief=tk.GROOVE, text = 'Update Report', font = 'Helvetica 12'
            , activebackground = 'DarkSlateGray1')
        self.update_report_btn['command'] = self.btn_event_update_report #self.update_report_func
        
        self.generate_report_btn = tk.Button(parent1, relief = tk.GROOVE, text = 'Generate Report', font = 'Helvetica 12'
            , activebackground = 'DarkSlateGray1')
        self.generate_report_btn['command'] = self.btn_event_gen_report
        self.generate_report_btn.place(relx=0, rely = 0, x= 10, y = 207, anchor = 'nw')

        self.btn_blink_start()


    def ctrl_panel_shift_y(self, y_shift, *widgets):
        for widget in widgets:
            tk_place_info = widget.place_info()
            widget.place(x = tk_place_info['x']
                , y = int(tk_place_info['y']) + y_shift
                , relx = tk_place_info['relx']
                , rely = tk_place_info['rely']
                , relwidth = tk_place_info['relwidth']
                , relheight = tk_place_info['relheight']
                , width = tk_place_info['width']
                , height = tk_place_info['height']
                , anchor = tk_place_info['anchor']
                )

    def clear_upload_log(self, scroll_class, log_dict, img_dict):
        log_dict.clear()
        img_dict.clear()
        parent = scroll_class.window_fr

        for widget in parent.winfo_children():
            widget.destroy()

        scroll_class.resize_frame(width = 300, height = 150)

    def clear_all_upload_log(self):
        self.clear_upload_log(self.photo_upload_log_class, self.photo_upload_log_dict, self.photo_upload_img_dict)
        self.clear_upload_log(self.target_upload_log_class, self.target_upload_log_dict, self.target_upload_img_dict)
        self.clear_upload_log(self.gray_upload_log_class, self.gray_upload_log_dict, self.gray_upload_img_dict)
        self.clear_upload_log(self.binary_upload_log_class, self.binary_upload_log_dict, self.binary_upload_img_dict)
        self.clear_upload_log(self.setup_upload_log_class, self.setup_upload_log_dict, self.setup_upload_img_dict)

        self.clear_upload_log(self.light_setup_log_class, self.light_setup_log_dict, self.light_setup_img_dict)

    def clear_all_report_detail(self):
        self.photo_dim_entry.set_text('', reset_undo_stack = True)
        self.photo_bg_combobox.set('')

        for tk_widget in self.test_criteria_entry_list:
            tk_widget.set_text('', reset_undo_stack = True)

        for tk_widget in self.light_detail_entry_list:
            tk_widget.set_text('', reset_undo_stack = True)

        for tk_widget in self.ctrl_detail_entry_list:
            tk_widget.set_text('', reset_undo_stack = True)

        for tk_widget in self.cam_detail_entry_list:
            tk_widget.set_text('', reset_undo_stack = True)

        for tk_widget in self.lens_detail_entry_list:
            tk_widget.set_text('', reset_undo_stack = True)


    def img_txt_popout_gen(self, parent, imgtxt_dict = None, toplvl_title = 'Toplevel', min_w = None, min_h = None, icon_img = None
        , topmost_bool = False
        , *args, **kwargs):

        tk_toplvl = CustomToplvl(parent, toplvl_title = toplvl_title, min_w = min_w, min_h = min_h, icon_img = icon_img
            , topmost_bool = topmost_bool
            , *args, **kwargs)

        tk_toplvl.protocol("WM_DELETE_WINDOW", partial(self.img_txt_popout_close, tk_toplvl = tk_toplvl))

        tk_scroll_txt = CustomScrollText(tk_toplvl)
        tk_scroll_txt.tk_txt['font'] = 'Helvetica 11'

        # tk_scroll_txt.place(x=0, y=50, relx=0, rely=0, relwidth = 1, relheight = 1, height = -50, anchor = 'nw')
        tk_scroll_txt.place(x=0, y=10, relx=0, rely=0, relwidth = 1, relheight = 1, height = -10, anchor = 'nw')
        # tk_scroll_txt.tk_str_var.trace("w",  lambda var_name, var_index, operation: self.img_txt_trace(tk_str_var = tk_scroll_txt.tk_str_var))

        # print(isinstance(tk_scroll_txt.tk_str_var, tk.StringVar))
        return tk_toplvl, tk_scroll_txt.tk_str_var

    def img_txt_trace(self, tk_str_var):
        # print(tk_str_var.get())
        pass

    def img_txt_popout_open(self, tk_toplvl):
        if False == tk_toplvl.check_open():
            tk_toplvl.open()
            if self.__curr_imtext_popout != tk_toplvl:
                self.img_txt_popout_set(tk_toplvl)
            self.__curr_imtext_popout = tk_toplvl

            screen_width = tk_toplvl.winfo_screenwidth()
            screen_height = tk_toplvl.winfo_screenheight()
            x_coordinate = int((screen_width/2) - (tk_toplvl.winfo_width()/2))
            y_coordinate = int((screen_height/2) - (tk_toplvl.winfo_height()/2))
            tk_toplvl.geometry("{}x{}+{}+{}".format(tk_toplvl.winfo_width(), tk_toplvl.winfo_height(), x_coordinate, y_coordinate))
        else:
            tk_toplvl.show()
            if self.__curr_imtext_popout != tk_toplvl:
                self.img_txt_popout_set(tk_toplvl)
            self.__curr_imtext_popout = tk_toplvl

    def img_txt_popout_close(self, tk_toplvl):
        tk_toplvl.close()

    def img_txt_popout_close_all(self):
        log_dict_list = [self.photo_upload_log_dict, self.light_setup_log_dict
                        , self.target_upload_log_dict, self.gray_upload_log_dict
                        , self.binary_upload_log_dict, self.setup_upload_log_dict]
        for log_dict in log_dict_list:
            for widget_list in log_dict.values():
                if len(widget_list) > 7:
                    if isinstance(widget_list[6], CustomToplvl):
                        widget_list[6].close()

        del log_dict_list

    def img_txt_popout_set(self, tk_toplvl):
        ### When we open 1 img_text_popout, the other existing popout must close...
        # print("Closing other imtext popout...")
        log_dict_list = [self.photo_upload_log_dict, self.light_setup_log_dict
                        , self.target_upload_log_dict, self.gray_upload_log_dict
                        , self.binary_upload_log_dict, self.setup_upload_log_dict]

        for log_dict in log_dict_list:
            for widget_list in log_dict.values():
                if len(widget_list) > 7:
                    if isinstance(widget_list[6], CustomToplvl) and widget_list[6] != tk_toplvl:
                        widget_list[6].close()

        del log_dict_list


    def upload_log_callback(self, parent, log_dict, img_dict, img_file = None, scroll_class = None, col_tracker_list = None, event = None
        , update_only = False, edit_dict_key = None):
        i = len(log_dict)

        widget_list = []
        place_y = int(5 + np.multiply(30, int(i)))
        col_width = int(0)
        ##############################################################################################

        save_tk_btn = tk.Button(parent, relief = tk.GROOVE, image = self.save_icon)#, width = 2, height = 1)
        save_tk_btn.place(x=col_width + 2,y=place_y - 2, relx=0, rely=0, anchor = 'nw')
        widget_list.append(save_tk_btn)

        save_tk_btn.update_idletasks()
        col_width = col_width + int(2 + save_tk_btn.winfo_width())
        ##############################################################################################

        img_lbl_str = 'Image ' + str(i + 1)

        img_tk_lbl = tk.Label(parent, text = img_lbl_str, bg = 'white', font = 'Helvetica 11')

        img_tk_lbl.place(x=col_width + 7, y=place_y, relx=0, rely=0, anchor = 'nw')
        widget_list.append(img_tk_lbl)

        img_tk_lbl.update_idletasks()
        col_width = col_width +  int(7 + img_tk_lbl.winfo_width())
        ##############################################################################################

        remove_tk_btn = tk.Button(parent, relief = tk.GROOVE, image = self.close_icon)#, width = 2, height = 1, bg = 'red')
        
        remove_tk_btn.place(x= col_width + 7,y=place_y - 2, relx=0, rely=0, anchor = 'nw')
        widget_list.append(remove_tk_btn)

        remove_tk_btn.update_idletasks()
        col_width = col_width +  int(7 + remove_tk_btn.winfo_width())
        ##############################################################################################

        up_tk_btn = tk.Button(parent, relief = tk.GROOVE, image = self.up_arrow_icon)#, width = 2, height = 1)
        up_tk_btn.place(x=col_width + 7,y=place_y - 2, relx=0, rely=0, anchor = 'nw')
        widget_list.append(up_tk_btn)

        up_tk_btn.update_idletasks()
        col_width = col_width +  int(7 + up_tk_btn.winfo_width())
        ##############################################################################################

        down_tk_btn = tk.Button(parent, relief = tk.GROOVE, image = self.down_arrow_icon)#, width = 2, height = 1)
        down_tk_btn.place(x=col_width + 7,y=place_y - 2, relx=0, rely=0, anchor = 'nw')
        widget_list.append(down_tk_btn)

        down_tk_btn.update_idletasks()

        col_width = col_width +  int(7 + down_tk_btn.winfo_width())
        ##############################################################################################

        txt_tk_btn = tk.Button(parent, relief = tk.GROOVE, image = self.text_icon)#, width = 2, height = 1)
        txt_tk_btn['bg'] = 'DodgerBlue2'
        txt_tk_btn['activebackground'] = 'navy'
        txt_tk_btn.place(x=col_width + 7,y=place_y - 2, relx=0, rely=0, anchor = 'nw')
        widget_list.append(txt_tk_btn)

        txt_tk_btn.update_idletasks()
        col_width = col_width +  int(7 + txt_tk_btn.winfo_width())

        tk_toplvl, tk_txt_var = self.img_txt_popout_gen(parent, toplvl_title = 'Image {} Description'.format(i + 1), min_w = 300, min_h = 225, topmost_bool = True
            , width = 300, height = 225)
        tk_toplvl.resizable(0,0)

        # tk_toplvl, tk_txt_var = self.img_txt_popout_gen(parent, toplvl_title = 'Image {} Text'.format(i + 1), min_w = 300, min_h = 225, topmost_bool = True
        #     , width = 300, height = 225)
        
        tk_toplvl, tk_txt_var = self.img_txt_popout_gen(parent, toplvl_title = '{} {} Text'.format(edit_dict_key, i + 1), min_w = 300, min_h = 225, topmost_bool = True
            , width = 300, height = 225, icon_img = self.window_icon)

        tk_toplvl.resizable(0,0)

        if edit_dict_key == 'Grayscale Image':
            if i == 0:
                tk_txt_var.set('Original Image')
            elif i == 1:
                tk_txt_var.set('Red Channel Image')
            elif i == 2:
                tk_txt_var.set('Green Channel Image')
            elif i == 3:
                tk_txt_var.set('Blue Channel Image')

        widget_list.append(tk_toplvl)
        widget_list.append(tk_txt_var)

        ### widget_list index; 0: save_tk_btn, 1: img_tk_lbl, 2: remove_tk_btn, 3: up_tk_btn, 4: down_tk_btn, 5: txt_tk_btn, 6: tk_toplvl, 7: tk_txt_var
        ##############################################################################################

        log_dict[str(i)] = widget_list

        if update_only == False:
            img_dict[str(i)] = img_file

        if len(img_dict) > 0:
            CreatePreviewDisp(img_tk_lbl, img_dict[str(i)], shift_x_disp = int(np.divide(img_tk_lbl.winfo_width(),2)), shift_y_disp = -75-1, width = 100, height = 75)

        save_tk_btn['command'] = partial(self.save_img_func, img_dict[str(i)])

        remove_tk_btn['command'] = partial(self.remove_upload_log_callback, dict_id=i, log_dict=log_dict, img_dict = img_dict, scroll_class = scroll_class
            , col_tracker_list = col_tracker_list, edit_dict_key = edit_dict_key)

        up_tk_btn['command'] =  partial(self.shift_upload_img_order, log_dict, img_dict, str(i), pos_shift = -1, edit_dict_key = edit_dict_key)
        down_tk_btn['command'] = partial(self.shift_upload_img_order, log_dict, img_dict, str(i), pos_shift = 1, edit_dict_key = edit_dict_key)
        txt_tk_btn['command'] = partial(self.img_txt_popout_open, tk_toplvl)

        row_height = int(5 + np.multiply(30, int(i+1)))

        if col_tracker_list is not None and type(col_tracker_list) == list:
            col_tracker_list.append(col_width)

        if scroll_class is not None:
            # print(col_width, scroll_class.frame_w - 20)
            if col_width > scroll_class.frame_w - 20: #-20 because of scroll_bar
                scroll_class.resize_frame(width = max(col_width + 20, 300))
                # scroll_class.invoke_resize()

            if row_height > scroll_class.frame_h - 15:
                scroll_class.resize_frame(height = max(row_height + 15, 150))
                # scroll_class.invoke_resize()

        del widget_list

    def ordered_list_get_index(self, ordered_dict, thekey):
        ref_list = list(ordered_dict.keys())
        return ref_list.index(thekey)

    def shift_upload_img_order(self, log_dict, img_dict, thekey, pos_shift, edit_dict_key = None):
        # print(img_dict)
        # print(thekey, type(thekey))
        # print('--------------------------------------')
        currpos = self.ordered_list_get_index(img_dict, thekey)
        # print("currpos: ", currpos)
        newpos = int(currpos + pos_shift)
        # print("newpos: ", newpos)
        if newpos >= 0:
            log_dict = move_element(log_dict, thekey, newpos)
            img_dict = move_element(img_dict, thekey, newpos)

            i = 0
            for _, widget_list in log_dict.items():
                if newpos == currpos:
                    pass

                elif newpos != currpos:
                    #### if newpos < currpos, this means we are shifting the target img UP-WARD
                    #### if newpos > currpos, this means we are shifting the target img DOWN-WARD
                    if i >= min(int(newpos), int(currpos)):
                        place_y = int(5 + np.multiply(30, int(i) ))
                        col_width = int(0)
                        
                        widget_list[0].place(x= col_width + 2,y=place_y - 2, relx=0, rely=0, anchor = 'nw')
                        widget_list[0].update_idletasks()
                        col_width = col_width + int(2 + widget_list[0].winfo_width())

                        widget_list[1].place(x= col_width + 7, y=place_y, relx=0, rely=0, anchor = 'nw')
                        widget_list[1].update_idletasks()
                        col_width = col_width +  int(7 + widget_list[1].winfo_width())

                        widget_list[2].place(x= col_width + 7, y=place_y - 2, relx=0, rely=0, anchor = 'nw')
                        widget_list[2].update_idletasks()
                        col_width = col_width +  int(7 + widget_list[2].winfo_width())

                        widget_list[3].place(x= col_width + 7, y=place_y - 2, relx=0, rely=0, anchor = 'nw')
                        widget_list[3].update_idletasks()
                        col_width = col_width +  int(7 + widget_list[3].winfo_width())

                        widget_list[4].place(x= col_width + 7, y=place_y - 2, relx=0, rely=0, anchor = 'nw')
                        widget_list[4].update_idletasks()
                        col_width = col_width +  int(7 + widget_list[4].winfo_width())

                        widget_list[5].place(x= col_width + 7, y=place_y - 2, relx=0, rely=0, anchor = 'nw')
                        widget_list[5].update_idletasks()
                        col_width = col_width +  int(7 + widget_list[5].winfo_width())

                        del col_width
                i += 1

            if edit_dict_key in self.edit_data_dict:
                self.check_image_edit(edit_dict_key, img_dict)

        else:
            if edit_dict_key in self.edit_data_dict:
                self.check_image_edit(edit_dict_key, img_dict)
            pass


    def upload_btn_callback(self, upload_log, upload_log_dict, upload_img_dict, scroll_class, col_tracker_list
        , edit_dict_key = None, root_path = None, track_dir_bool = True):

        if root_path is not None and path.isdir(root_path):
            # print(path.isdir(root_path))
            file = filedialog.askopenfilenames(initialdir = root_path, title="Select file", filetypes=self.img_format_list)

        else:
            # print("initialdir not Given")
            file = filedialog.askopenfilenames(title="Select file", filetypes=self.img_format_list)

        # print(file, type(file))
        if file == '':
            #print('empty directory')
            pass
        else:
            if type(file) == str:
                self.upload_log_callback(upload_log, upload_log_dict, upload_img_dict, img_file = file
                    , scroll_class = scroll_class, col_tracker_list = col_tracker_list, edit_dict_key = edit_dict_key)

                if track_dir_bool == True:
                    self.__im_curr_dir = os.path.join(file, os.pardir)

                if edit_dict_key in self.edit_data_dict:
                    self.check_image_edit(edit_dict_key, upload_img_dict)

            elif type(file) == tuple:
                for imfile in file:
                    self.upload_log_callback(upload_log, upload_log_dict, upload_img_dict, img_file = imfile
                        , scroll_class = scroll_class, col_tracker_list = col_tracker_list, edit_dict_key = edit_dict_key)

                if track_dir_bool == True:
                    self.__im_curr_dir = os.path.join(imfile, os.pardir)

                if edit_dict_key in self.edit_data_dict:
                    self.check_image_edit(edit_dict_key, upload_img_dict)

            # print(self.photo_upload_log_dict, self.photo_upload_col_tracker)

    def remove_upload_log_callback(self, dict_id, log_dict, img_dict, scroll_class = None, col_tracker_list = None, edit_dict_key = None, event = None):
        # print('dict_id: ', dict_id)

        for widget in log_dict[str(dict_id)]:
            try:
                widget.destroy()
            except(AttributeError, tk.TclError):
                del widget

        log_dict, _ = remove_key(log_dict, dict_id)
        img_dict, key_pos = remove_key(img_dict, dict_id) #key_pos is the position of the element for the removed key

        del col_tracker_list[dict_id]

        if edit_dict_key in self.edit_data_dict:
            self.check_image_edit(edit_dict_key, img_dict)

        i = 0
        for tk_id, widget_list in log_dict.copy().items():
            # print('tk_id, i, key_pos: ', tk_id, i, key_pos)
            widget_list[0]['command'] = partial(self.save_img_func, img_dict[tk_id])

            update_text = 'Image {}'.format(int(tk_id) + 1)
            widget_list[1]['text'] = update_text

            widget_list[2]['command'] = partial(self.remove_upload_log_callback, dict_id=int(tk_id), log_dict=log_dict, img_dict = img_dict, scroll_class = scroll_class
                , col_tracker_list = col_tracker_list, edit_dict_key = edit_dict_key)

            widget_list[3]['command'] = partial(self.shift_upload_img_order, log_dict, img_dict, tk_id, pos_shift = -1, edit_dict_key = edit_dict_key)
            widget_list[4]['command'] = partial(self.shift_upload_img_order, log_dict, img_dict, tk_id, pos_shift =  1, edit_dict_key = edit_dict_key)

            # widget_list[6].title('Image {} Description'.format(int(tk_id) + 1))
            widget_list[6].title('{} {}'.format(edit_dict_key, int(tk_id) + 1))

            if i >= key_pos:
                place_y = int(5 + np.multiply(30, int(i) ))
                col_width = int(0)

                widget_list[0].place(x= col_width + 2,y=place_y - 2, relx=0, rely=0, anchor = 'nw')
                widget_list[0].update_idletasks()
                col_width = col_width + int(2 + widget_list[0].winfo_width())

                widget_list[1].place(x= col_width + 7, y=place_y, relx=0, rely=0, anchor = 'nw')
                widget_list[1].update_idletasks()
                col_width = col_width +  int(7 + widget_list[1].winfo_width())

                widget_list[2].place(x= col_width + 7, y=place_y - 2, relx=0, rely=0, anchor = 'nw')
                widget_list[2].update_idletasks()
                col_width = col_width +  int(7 + widget_list[2].winfo_width())

                widget_list[3].place(x= col_width + 7, y=place_y - 2, relx=0, rely=0, anchor = 'nw')
                widget_list[3].update_idletasks()
                col_width = col_width +  int(7 + widget_list[3].winfo_width())

                widget_list[4].place(x= col_width + 7, y=place_y - 2, relx=0, rely=0, anchor = 'nw')
                widget_list[4].update_idletasks()
                col_width = col_width +  int(7 + widget_list[4].winfo_width())

                widget_list[5].place(x= col_width + 7, y=place_y - 2, relx=0, rely=0, anchor = 'nw')
                widget_list[5].update_idletasks()
                col_width = col_width +  int(7 + widget_list[5].winfo_width())

                del col_width
            i += 1

        if len(col_tracker_list) > 0:
            col_width = max(col_tracker_list)
            if scroll_class is not None:
                scroll_class.resize_frame(width = max(col_width + 20, 300))
                # scroll_class.invoke_resize()

        elif len(col_tracker_list) == 0:
            if scroll_class is not None:
                scroll_class.resize_frame(width = 300)
                # scroll_class.invoke_resize()

        row_height = int(5 + np.multiply(30, int(len(log_dict))) )
        scroll_class.resize_frame(height = max(row_height + 15, 150))

    def save_img_func(self, imdata):
        f = filedialog.asksaveasfile(initialdir = self.__im_curr_dir
            , mode='w', defaultextension = self.img_format_list, filetypes = self.img_format_list)
        if f is None: # asksaveasfile return `None` if dialog closed with "cancel".
            return

        else:
            abs_path = os.path.abspath(f.name)
            if type(imdata) == str and (path.exists(imdata)) == True:
                im_np = cv2.imread(imdata, -1)
                cv2.imwrite(abs_path, im_np)
                del im_np
                tkinter.messagebox.showinfo('Success','The Red Channel Image is Saved')

                Info_Msgbox(message = 'Save Image Success!' + '\n\n' + abs_path, title = 'Image Save'
                    , font = 'Helvetica 10', width = 400)

            elif isinstance(imdata, np.ndarray) == True:
                cv2.imwrite(abs_path, im_np)
                del im_np
                Info_Msgbox(message = 'Save Image Success!' + '\n\n' + abs_path, title = 'Image Save'
                    , font = 'Helvetica 10', width = 400)

            elif isinstance(imdata, Image.Image) == True:
                im_pil.save(abs_path)
                del im_pil

                Info_Msgbox(message = 'Save Image Success!' + '\n\n' + abs_path, title = 'Image Save'
                    , font = 'Helvetica 10', width = 400)

            elif isinstance(imdata, openpyxl.drawing.image.Image) == True:
                im_pil = Image.open(io.BytesIO(imdata._data()))
                im_pil.save(abs_path)
                del im_pil

                Info_Msgbox(message = 'Save Image Success!' + '\n\n' + abs_path, title = 'Image Save'
                    , font = 'Helvetica 10', width = 400)

    def reset_init_state(self):
        self.xl_load_bool = False
        self.xl_load_path = None

        self.xl_worksheet_name = 'Sheet'
        # self.worksheet_name_var.set(self.xl_worksheet_name)
        self.worksheet_name_entry.set_text(self.xl_worksheet_name, reset_undo_stack = True)

        self.create_worksheet_bool.set(1)

        self.worksheet_name_dropbox['value'] = []
        self.worksheet_name_dropbox.set('')
        widget_disable(self.worksheet_name_dropbox, self.rename_worksheet_btn, self.create_worksheet_checkbtn
            , self.reload_worksheet_btn)

        self.curr_report_file_tk_var.set('Current Report File: ' + (str(self.xl_load_path).split('/'))[-1])
        self.curr_worksheet_tk_var.set('Current Worksheet: ' + 'None')

        widget_enable(self.update_report_btn, self.generate_report_btn, self.update_report_checkbtn
            , self.worksheet_name_entry)

    def reset_worksheet_data(self):
        for edit_data in self.edit_data_dict.values():
            if type(edit_data[2]) == list:
                self.clear_ref_edit_list(edit_data[2])
            elif (isinstance(edit_data[2], np.ndarray)) == True:
                self.clear_ref_edit_arr(edit_data[2])
            
            self.reset_edit_bool_arr(edit_data[1])

        self.clear_all_upload_log()
        self.reset_track_edit_bool_arr()
        self.clear_all_report_detail()

    def clear_btn_event(self):
        ask_msgbox = Ask_Msgbox(message = "Unsaved changes will be cleared!" + "\n\nAre you sure?"
            , parent = self.master, title = 'Clear All', message_anchor = 'w', ask_OK = False, mode = 'warning')

        if ask_msgbox.ask_result() == True:
            self.reset_worksheet_data()

    def new_report_init(self):
        self.report_mode_btn_state(self.new_report_btn, self.load_report_btn)
        ask_msgbox = Ask_Msgbox(message = "You have selected 'New Report'.\nAny unsaved changes will be cleared!" + "\n\nAre you sure?"
            , parent = self.master, title = 'New Report', message_anchor = 'w', ask_OK = False, height = 200)

        if ask_msgbox.ask_result() == True:
            self.xl_load_bool = False
            self.xl_load_path = None
            # self.__edit_bool = False
            widget_disable(self.reload_worksheet_btn)

            self.xl_worksheet_name = 'Sheet'
            # self.worksheet_name_var.set(self.xl_worksheet_name)
            self.worksheet_name_entry.set_text(self.xl_worksheet_name, reset_undo_stack = True)
            self.update_report_btn.place_forget()

            self.update_report_checkbtn.place(relx=0, rely = 0, x= 10 + 140, y = 207, anchor = 'nw')
            # self.update_report_checkbtn.place(relx=0, rely = 0, x= 10 + 140, y = 207 + 5, anchor = 'nw')
            self.generate_report_btn.place(relx=0, rely = 0, x= 10, y = 207, anchor = 'nw')

            self.create_worksheet_bool.set(1)

            self.worksheet_name_dropbox['value'] = []
            self.worksheet_name_dropbox.set('')

            widget_disable(self.worksheet_name_dropbox, self.rename_worksheet_btn, self.create_worksheet_checkbtn)

            self.reset_worksheet_data()

            self.curr_report_file_tk_var.set('Current Report File: ' + (str(self.xl_load_path).split('/'))[-1])
            self.curr_worksheet_tk_var.set('Current Worksheet: ' + 'None')

        else:
            if self.xl_load_path is not None and self.xl_load_bool == True:
                self.report_mode_btn_state(self.load_report_btn, self.new_report_btn)


    def load_existing_report(self):
        self.report_mode_btn_state(self.load_report_btn, self.new_report_btn)
        file = filedialog.askopenfilename(initialdir = self.__report_curr_dir, parent = self, title="Select file", filetypes = self.report_format_list, multiple = False)
        err_ret = False

        if file == '':
            if self.xl_load_path is not None and self.xl_load_bool == True:
                self.report_mode_btn_state(self.load_report_btn, self.new_report_btn)
            else:
                self.report_mode_btn_state(self.new_report_btn, self.load_report_btn)
            return
        else:
            _file_name = (re.findall(r'[^\\/]+|[\\/]', file))[-1]
            # print(file, _file_name)
            save_folder = re.sub(_file_name, '', file)
            # print(save_folder)
            try: 
                os.rename(save_folder + ntpath.basename(file), save_folder + 'open_tempfile.xlsx')
                os.rename(save_folder + 'open_tempfile.xlsx', save_folder + ntpath.basename(file))
            except OSError:# as e:
                # print(e)
                err_ret = True

            if err_ret == True:
                Error_Msgbox(message = "Load Error!\n\n"
                    + "Report File is Currently Opened in Microsoft Excel, please close Microsoft Excel and try to Load again.", title = 'Error'
                        , message_anchor = 'w', width = 370)

                self.event_log_custom_error(file, 'Report File is Currently Opened in Microsoft Excel, please close Microsoft Excel and try to Load again.', error_type = 'Load ERROR')
                if self.xl_load_bool is False:
                    self.new_report_init()

            elif err_ret == False:
                self.__report_curr_dir = os.path.join(file, os.pardir)
                self.xl_load_path = file
                # self.xl_class.xl_update_workbook(self.xl_load_path)
                self.xl_class.xl_open_workbook(self.xl_load_path)
                
                ###########################################################################
                # Have to Set this Default Width for Some of these Columns....
                self.xl_class.xl_set_column_width(self.xl_class.worksheet, 'B', 'F')
                self.xl_class.xl_set_column_width(self.xl_class.worksheet, 'G', 'L')
                self.xl_class.xl_set_column_width(self.xl_class.worksheet, 'Q', 'V')
                self.xl_class.xl_set_column_width(self.xl_class.worksheet, 'W', 'AB')
                self.xl_class.xl_set_column_width(self.xl_class.worksheet, 'AC', 'AH')
                self.xl_class.xl_set_column_width(self.xl_class.worksheet, 'AI', 'AN')
                ###########################################################################

                self.reset_worksheet_data()

                self.load_worksheet_data()
                self.load_worksheet_image()

                self.update_worksheet_dropbox()
                self.xl_worksheet_name = self.xl_class.worksheet.title
                # self.worksheet_name_var.set(self.xl_worksheet_name)
                self.worksheet_name_entry.set_text(self.xl_worksheet_name, reset_undo_stack = True)

                widget_enable(self.reload_worksheet_btn)

                self.worksheet_name_dropbox['state'] = 'readonly'

                widget_enable(self.rename_worksheet_btn, self.create_worksheet_checkbtn)

                self.xl_class.xl_close_workbook(self.xl_load_path)

                self.xl_load_bool = True

                # self.update_report_btn.place(relx=0, rely = 0, x= 10, y = 410, anchor = 'nw')
                self.update_report_btn.place(relx=0, rely = 0, x= 10, y = 207, anchor = 'nw')
                self.generate_report_btn.place_forget()
                self.update_report_checkbtn.place_forget()

                self.create_worksheet_bool.set(0)


                self.curr_report_file_tk_var.set('Current Report File: ' + (str(self.xl_load_path).split('/'))[-1])
                self.curr_worksheet_tk_var.set('Current Worksheet: ' + self.xl_worksheet_name)

                self.event_log_load_report()

    def reload_worksheet_data(self):
        if self.__edit_bool == False:
            self.reset_worksheet_data()
            self.load_worksheet_data()
            self.load_worksheet_image()

        elif self.__edit_bool == True:
            ask_msgbox = Ask_Msgbox(message = "Current Edit/Changes on UI will be replaced if you Reload Worksheet!" + "\n\nAre you sure?"
            , parent = self.master, title = 'Warning', mode = 'warning', message_anchor = 'w', width = 350, ask_OK = False)
            if ask_msgbox.ask_result() == True:
                self.reset_worksheet_data()
                self.load_worksheet_data()
                self.load_worksheet_image()
                print(self.photo_upload_img_dict)
            else:
                pass

    def load_worksheet_data(self):
        read_result = xl_read_worksheet(ws = self.xl_class.worksheet, min_row=3, min_col='B', values_only=True)
        
        for read_data in read_result:
            if read_data is not None:
                self.photo_dim_entry.set_text(read_data, reset_undo_stack = True)
            elif read_data is None:
                self.photo_dim_entry.set_text('', reset_undo_stack = True)

            self.edit_data_dict["Sample Detail"][2][0] = read_data
            break ### read_result from B3 has only 1 result. So, should random error occur (random additional data), we break after 1 iteration as failsafe


        read_result = xl_read_worksheet(ws = self.xl_class.worksheet, min_row=4, min_col='B', values_only=True)
        for read_data in read_result:
            if read_data in self.photo_bg_list:
                self.photo_bg_combobox.current(self.photo_bg_list.index(read_data))
            else:
                # print("Background Category Empty...")
                self.photo_bg_combobox.set('')

            self.edit_data_dict["Sample Detail"][2][1] = read_data
            break ### read_result from B4 has only 1 result. So, should random error occur (random additional data), we break after 1 iteration as failsafe
        
        # print(self.edit_data_dict["Sample Detail"][2])

        read_result = xl_read_worksheet(ws = self.xl_class.worksheet, min_row=5, max_row=10, min_col='O', values_only=True)
        i = 0
        for i, read_data in enumerate(read_result):
            if read_data is not None:
                self.test_criteria_entry_list[i].set_text(read_data, reset_undo_stack = True)
            elif read_data is None:
                self.test_criteria_entry_list[i].set_text('', reset_undo_stack = True)

            self.edit_data_dict["Testing Criteria"][2][i] = read_data

        read_result = xl_read_worksheet(ws = self.xl_class.worksheet, min_row=12, max_row=15, min_col='N', values_only=True)
        i = 0
        for i, read_data in enumerate(read_result):
            if read_data is not None:
                self.light_detail_entry_list[i].set_text(read_data, reset_undo_stack = True)
            elif read_data is None:
                self.light_detail_entry_list[i].set_text('', reset_undo_stack = True)

            self.edit_data_dict["Lighting Details"][2][i] = read_data

        read_result = xl_read_worksheet(ws = self.xl_class.worksheet, min_row=17, max_row=20, min_col='N', values_only=True)
        i = 0
        for i, read_data in enumerate(read_result):
            if read_data is not None:
                self.ctrl_detail_entry_list[i].set_text(read_data, reset_undo_stack = True)
            elif read_data is None:
                self.ctrl_detail_entry_list[i].set_text('', reset_undo_stack = True)

            self.edit_data_dict["Controller Details"][2][i] = read_data

        read_result = xl_read_worksheet(ws = self.xl_class.worksheet, min_row=22, max_row=31, min_col='N', values_only=True)
        i = 0
        for i, read_data in enumerate(read_result):
            if read_data is not None:
                self.cam_detail_entry_list[i].set_text(read_data, reset_undo_stack = True)
            elif read_data is None:
                self.cam_detail_entry_list[i].set_text('', reset_undo_stack = True)

            self.edit_data_dict["Camera Details"][2][i] = read_data

        read_result = xl_read_worksheet(ws = self.xl_class.worksheet, min_row=33, max_row=38, min_col='N', values_only=True)
        i = 0
        for i, read_data in enumerate(read_result):
            if read_data is not None:
                self.lens_detail_entry_list[i].set_text(read_data, reset_undo_stack = True)
            elif read_data is None:
                self.lens_detail_entry_list[i].set_text('', reset_undo_stack = True)

            self.edit_data_dict["Lens Details"][2][i] = read_data
    
    def load_csv_data(self):
        csv_file = filedialog.askopenfilename(initialdir = self.__csv_curr_dir, parent = self, title="Select file", filetypes = [('CSV file', '*.csv')], multiple = False)
        if csv_file == '':
            return
        else:
            self.reset_worksheet_data()
            self.__csv_curr_dir = os.path.join(csv_file, os.pardir)

            ret, csv_data, csv_img = tk_opencsv.read_csv(csv_file)
            if ret == 0:
                sample_detail   = csv_data[0]
                test_criteria   = csv_data[1]
                light_detail    = csv_data[2]
                ctrl_detail     = csv_data[3]
                cam_detail      = csv_data[4]
                lens_detail     = csv_data[5]

                sample_bg   = sample_detail['Background'] 
                sample_dim  = sample_detail['Product dimension']
                if sample_dim is not None:
                    self.photo_dim_entry.set_text(str(sample_dim), reset_undo_stack = True)

                elif sample_dim is None:
                    self.photo_dim_entry.set_text('', reset_undo_stack = True)

                if sample_bg in self.photo_bg_list:
                    self.photo_bg_combobox.current(self.photo_bg_list.index(sample_bg))
                else:
                    self.photo_bg_combobox.set('')
            
            i = 0
            for i, read_data in enumerate(test_criteria.values()):
                if read_data is not None:
                    self.test_criteria_entry_list[i].set_text(read_data, reset_undo_stack = True)
                elif read_data is None:
                    self.test_criteria_entry_list[i].set_text('', reset_undo_stack = True)
            
            i = 0
            for i, read_data in enumerate(light_detail.values()):
                if read_data is not None:
                    self.light_detail_entry_list[i].set_text(read_data, reset_undo_stack = True)
                elif read_data is None:
                    self.light_detail_entry_list[i].set_text('', reset_undo_stack = True)
            
            i = 0
            for i, read_data in enumerate(ctrl_detail.values()):
                if read_data is not None:
                    self.ctrl_detail_entry_list[i].set_text(read_data, reset_undo_stack = True)
                elif read_data is None:
                    self.ctrl_detail_entry_list[i].set_text('', reset_undo_stack = True)
            
            i = 0
            for i, read_data in enumerate(cam_detail.values()):
                if read_data is not None:
                    self.cam_detail_entry_list[i].set_text(read_data, reset_undo_stack = True)
                elif read_data is None:
                    self.cam_detail_entry_list[i].set_text('', reset_undo_stack = True)
            
            i = 0
            for i, read_data in enumerate(lens_detail.values()):
                if read_data is not None:
                    self.lens_detail_entry_list[i].set_text(read_data, reset_undo_stack = True)
                elif read_data is None:
                    self.lens_detail_entry_list[i].set_text('', reset_undo_stack = True)

            ### Check Entry for any Edit/Changes
            for i, tk_var in enumerate(self.test_criteria_var_list):
                self.check_entry_edit(tk_event = None, ref_arr_index = i, tk_var = tk_var, edit_dict_key = "Testing Criteria")

            for i, tk_var in enumerate(self.light_detail_var_list):
                self.check_entry_edit(tk_event = None, ref_arr_index = i, tk_var = tk_var, edit_dict_key = "Lighting Details")

            for i, tk_var in enumerate(self.ctrl_detail_var_list):
                self.check_entry_edit(tk_event = None, ref_arr_index = i, tk_var = tk_var, edit_dict_key = "Controller Details")

            for i, tk_var in enumerate(self.cam_detail_var_list):
                self.check_entry_edit(tk_event = None, ref_arr_index = i, tk_var = tk_var, edit_dict_key = "Camera Details")

            for i, tk_var in enumerate(self.lens_detail_var_list):
                self.check_entry_edit(tk_event = None, ref_arr_index = i, tk_var = tk_var, edit_dict_key = "Lens Details")

            #############################################################################################################################################
            ###load csv images
            start_id_arr = np.zeros(2, dtype=np.uint8) #Determine the start image id when loading image from worksheet. If current dictionary(ies) is/are empty, we start at id = 0.
            start_id_arr[0] = (0 if len(self.photo_upload_img_dict) == 0 else int(list(self.photo_upload_img_dict.keys())[-1]) + 1)
            start_id_arr[1] = (0 if len(self.target_upload_img_dict) == 0 else int(list(self.target_upload_img_dict.keys())[-1]) + 1)

            start_id_copy = start_id_arr.copy() ### A copy to recall what were the start_id values when we update the GUI using self.upload_log_callback

            for np_img in csv_img[0].values():
                self.photo_upload_img_dict[str(start_id_arr[0])] = np_img
                start_id_arr[0] = start_id_arr[0] + 1

            for np_img in csv_img[1].values():
                self.target_upload_img_dict[str(start_id_arr[1])] = np_img
                start_id_arr[1] = start_id_arr[1] + 1

            for _ in range(start_id_copy[0], len(self.photo_upload_img_dict)):
                self.upload_log_callback(parent = self.photo_upload_log, log_dict = self.photo_upload_log_dict, img_dict = self.photo_upload_img_dict
                    , scroll_class = self.photo_upload_log_class, col_tracker_list = self.photo_upload_col_tracker, update_only = True, edit_dict_key = "Sample Image")
            for _ in range(start_id_copy[1], len(self.target_upload_img_dict)):
                self.upload_log_callback(parent = self.target_upload_log, log_dict = self.target_upload_log_dict, img_dict = self.target_upload_img_dict
                    , scroll_class = self.target_upload_log_class, col_tracker_list = self.target_upload_col_tracker, update_only = True, edit_dict_key = "Target Image")

            del start_id_copy, start_id_arr

            ### Check Image Upload for any Edit/Changes
            self.check_image_edit("Sample Image", self.photo_upload_img_dict)
            self.check_image_edit("Target Image", self.target_upload_img_dict)


    def retrieve_imtext_data(self, ws, img_obj, cell_id, start_col, end_col):

        _, imtext_cell_id, _ = self.xl_class.compute_next_anchor_img(ws = ws, img_obj = img_obj, cell_id = cell_id
            , start_col = start_col, end_col = end_col)
        # print(imtext_cell_id)
        
        col_char = re.findall('(\\D+)', imtext_cell_id)[0]
        row_num = int(re.findall('(\\d+)', imtext_cell_id)[0])

        ### We just use row_num to find the image text because reports generated via this software has image spacing of + 1
        imtext_data = xl_read_worksheet(ws = ws, min_row=row_num, min_col=col_char, values_only=True)

        ### imtext_data is generator-type object from yield()

        return imtext_data

    def load_worksheet_image(self):
        start_id_arr = np.zeros(6, dtype=np.uint8) #Determine the start image id when loading image from worksheet. If current dictionary(ies) is/are empty, we start at id = 0.
        start_id_arr[0] = (0 if len(self.photo_upload_img_dict) == 0 else int(list(self.photo_upload_img_dict.keys())[-1]) + 1)
        start_id_arr[1] = (0 if len(self.light_setup_img_dict) == 0 else int(list(self.light_setup_img_dict.keys())[-1]) + 1)
        start_id_arr[2] = (0 if len(self.target_upload_img_dict) == 0 else int(list(self.target_upload_img_dict.keys())[-1]) + 1)
        start_id_arr[3] = (0 if len(self.gray_upload_img_dict) == 0 else int(list(self.gray_upload_img_dict.keys())[-1]) + 1)
        start_id_arr[4] = (0 if len(self.binary_upload_img_dict) == 0 else int(list(self.binary_upload_img_dict.keys())[-1]) + 1)
        start_id_arr[5] = (0 if len(self.setup_upload_img_dict) == 0 else int(list(self.setup_upload_img_dict.keys())[-1]) + 1)
        # print(start_id_arr)
        start_id_copy = start_id_arr.copy() ### A copy to recall what were the start_id values when we update the GUI using self.upload_log_callback

        xl_imloader = self.xl_class.xl_read_image(ws = self.xl_class.worksheet, sort_col_bool = True) #will sort according to columns then rows (top-to-bottom)...
        # print(xl_imloader)
        imtext_memory_arr = np.empty((6), dtype=object)
        for i in range(0, len(imtext_memory_arr)):
            imtext_memory_arr[i] = [] ###array containg image text or descriptions from the excel report when user load excel report.

        ### Extracting image(s) from excel and inserting it into respective img_dictionary(ies)
        print(xl_imloader._images.items(), start_id_arr)
        for cell_id, img_obj in xl_imloader._images.items():
            col_char = re.split('(\\d+)',cell_id)[0]
            if col_char == 'A':
                self.photo_upload_img_dict[str(start_id_arr[0])] = img_obj
                start_id_arr[0] = start_id_arr[0] + 1
                imtext_data = self.retrieve_imtext_data(ws = self.xl_class.worksheet, img_obj = img_obj, cell_id = cell_id, start_col = 'A', end_col = 'F')
                imtext_memory_arr[0].append(imtext_data)

            elif col_char == 'G':
                self.light_setup_img_dict[str(start_id_arr[1])] = img_obj
                start_id_arr[1] = start_id_arr[1] + 1
                imtext_data = self.retrieve_imtext_data(ws = self.xl_class.worksheet, img_obj = img_obj, cell_id = cell_id, start_col = 'G', end_col = 'L')
                imtext_memory_arr[1].append(imtext_data)

            elif col_char == 'Q':
                self.target_upload_img_dict[str(start_id_arr[2])] = img_obj
                start_id_arr[2] = start_id_arr[2] + 1
                imtext_data = self.retrieve_imtext_data(ws = self.xl_class.worksheet, img_obj = img_obj, cell_id = cell_id, start_col = 'Q', end_col = 'V')
                imtext_memory_arr[2].append(imtext_data)

            elif col_char == 'W':
                self.gray_upload_img_dict[str(start_id_arr[3])] = img_obj
                start_id_arr[3] = start_id_arr[3] + 1
                imtext_data = self.retrieve_imtext_data(ws = self.xl_class.worksheet, img_obj = img_obj, cell_id = cell_id, start_col = 'W', end_col = 'AB')
                imtext_memory_arr[3].append(imtext_data)

            elif col_char == 'AC':
                self.binary_upload_img_dict[str(start_id_arr[4])] = img_obj
                start_id_arr[4] = start_id_arr[4] + 1
                imtext_data = self.retrieve_imtext_data(ws = self.xl_class.worksheet, img_obj = img_obj, cell_id = cell_id, start_col = 'AC', end_col = 'AH')
                imtext_memory_arr[4].append(imtext_data)

            elif col_char == 'AI':
                self.setup_upload_img_dict[str(start_id_arr[5])] = img_obj
                start_id_arr[5] = start_id_arr[5] + 1
                imtext_data = self.retrieve_imtext_data(ws = self.xl_class.worksheet, img_obj = img_obj, cell_id = cell_id, start_col = 'AI', end_col = 'AN')
                imtext_memory_arr[5].append(imtext_data)

            del img_obj

        del xl_imloader

        ### Create corresponding widgets in each upload User-Interface base on the amount of image(s) in the image dictionary(ies)
        for _ in range(start_id_copy[0], len(self.photo_upload_img_dict)):
            self.upload_log_callback(parent = self.photo_upload_log, log_dict = self.photo_upload_log_dict, img_dict = self.photo_upload_img_dict
                , scroll_class = self.photo_upload_log_class, col_tracker_list = self.photo_upload_col_tracker, update_only = True, edit_dict_key = "Sample Image")

        for _ in range(start_id_copy[1], len(self.light_setup_img_dict)):
            self.upload_log_callback(parent = self.light_setup_log, log_dict = self.light_setup_log_dict, img_dict = self.light_setup_img_dict
                , scroll_class = self.light_setup_log_class, col_tracker_list = self.light_setup_col_tracker, update_only = True, edit_dict_key = "Lighting Drawing")

        for _ in range(start_id_copy[2], len(self.target_upload_img_dict)):
            self.upload_log_callback(parent = self.target_upload_log, log_dict = self.target_upload_log_dict, img_dict = self.target_upload_img_dict
                , scroll_class = self.target_upload_log_class, col_tracker_list = self.target_upload_col_tracker, update_only = True, edit_dict_key = "Target Image")

        for _ in range(start_id_copy[3], len(self.gray_upload_img_dict)):
            self.upload_log_callback(parent = self.gray_upload_log, log_dict = self.gray_upload_log_dict, img_dict = self.gray_upload_img_dict
                , scroll_class = self.gray_upload_log_class, col_tracker_list = self.gray_upload_col_tracker, update_only = True, edit_dict_key = "Grayscale Image")

        for _ in range(start_id_copy[4], len(self.binary_upload_img_dict)):
            self.upload_log_callback(parent = self.binary_upload_log, log_dict = self.binary_upload_log_dict, img_dict = self.binary_upload_img_dict
                , scroll_class = self.binary_upload_log_class, col_tracker_list = self.binary_upload_col_tracker, update_only = True, edit_dict_key = "Binary Image")

        for _ in range(start_id_copy[5], len(self.setup_upload_img_dict)):
            self.upload_log_callback(parent = self.setup_upload_log, log_dict = self.setup_upload_log_dict, img_dict = self.setup_upload_img_dict
                , scroll_class = self.setup_upload_log_class, col_tracker_list = self.setup_upload_col_tracker, update_only = True, edit_dict_key = "Setup Image")

        del start_id_copy, start_id_arr

        self.edit_data_dict["Sample Image"][2] = list(self.photo_upload_img_dict.values())
        self.edit_data_dict["Lighting Drawing"][2] = list(self.light_setup_img_dict.values())
        self.edit_data_dict["Target Image"][2] = list(self.target_upload_img_dict.values())
        self.edit_data_dict["Grayscale Image"][2] = list(self.gray_upload_img_dict.values())
        self.edit_data_dict["Binary Image"][2] = list(self.binary_upload_img_dict.values())
        self.edit_data_dict["Setup Image"][2] = list(self.setup_upload_img_dict.values())

        # group_print(self.edit_data_dict["Sample Image"][2]
        #     , self.edit_data_dict["Lighting Drawing"][2]
        #     , self.edit_data_dict["Target Image"][2]
        #     , self.edit_data_dict["Grayscale Image"][2]
        #     , self.edit_data_dict["Binary Image"][2]
        #     , self.edit_data_dict["Setup Image"][2])

        #############################################################################################################################################################
        ### Load Image Text(s) into each imtext popout GUI(s). We use upload_log_dict to access the tk.StringVar() Tkinter object.
        self.load_imtext_popout(log_dict = self.photo_upload_log_dict, imtext_memory = imtext_memory_arr[0])
        self.load_imtext_popout(log_dict = self.light_setup_log_dict,  imtext_memory = imtext_memory_arr[1])
        self.load_imtext_popout(log_dict = self.target_upload_log_dict,imtext_memory = imtext_memory_arr[2])
        self.load_imtext_popout(log_dict = self.gray_upload_log_dict,  imtext_memory = imtext_memory_arr[3])
        self.load_imtext_popout(log_dict = self.binary_upload_log_dict,imtext_memory = imtext_memory_arr[4])
        self.load_imtext_popout(log_dict = self.setup_upload_log_dict, imtext_memory = imtext_memory_arr[5])

        del imtext_memory_arr

    def load_imtext_popout(self, log_dict, imtext_memory):
        for i, widget_list in enumerate(log_dict.values()):
            if 0 <= i <= len(imtext_memory) - 1:
                imtext_str = ''
                for imtext in imtext_memory[i]:
                    if isinstance(imtext, str) and len(imtext) > 0:
                        imtext_str = imtext_str + '{} '.format(imtext)

                if len(imtext_str) > 0:
                    imtext_str = "".join(imtext_str.rstrip())
            widget_list[7].set(imtext_str)


    def enable_update_generate_report(self):
        if self.update_report_bool.get() == 1:
            ask_msgbox = Ask_Msgbox(message = "Enabling this will cause 'Generate Report' button to update new values"
                + " to the newly generated report." + "\n\nAre you sure?"
            , parent = self.master, title = "Enable Update", message_anchor = 'w', height = 170)
            
            if ask_msgbox.ask_result() == True:
                self.update_report_bool.set(1)
            else:
                self.update_report_bool.set(0)


        elif self.update_report_bool.get() == 0:
            ask_msgbox = Ask_Msgbox(message = "Disabling this will cause 'Generate Report' button to generate a new file."
                + "\n\nAre you sure?"
            , parent = self.master, title = "Disable Update", message_anchor = 'w', height = 170)

            if ask_msgbox.ask_result() == True:
                self.update_report_bool.set(0)
            else:
                self.update_report_bool.set(1)

    def enable_create_worksheet(self):
        if self.create_worksheet_bool.get() == 1:
            ask_msgbox = Ask_Msgbox(message = "Enabling this will generate a new WorkSheet in the current Excel file." 
                + "\n\nAre you sure?"
            , parent = self.master, title = "Enable Update", message_anchor = 'w', height = 170)
            
            if ask_msgbox.ask_result() == True:
                self.create_worksheet_bool.set(1)
            else:
                self.create_worksheet_bool.set(0)


        elif self.create_worksheet_bool.get() == 0:
            ask_msgbox = Ask_Msgbox(message = "Disabling this will update the current WorkSheet in the current Excel file.\n"+
                "This will DELETE any existing data in the worksheet and REPLACED with updated values!"
                + "\n\nAre you sure?"
            , parent = self.master, title = "Disable Update", message_anchor = 'w', mode = 'warning', width = 450, height = 170)

            if ask_msgbox.ask_result() == True:
                self.create_worksheet_bool.set(0)
            else:
                self.create_worksheet_bool.set(1)


    def btn_event_gen_report(self):
        self.thread_generate_report()

    def btn_event_update_report(self):
        self.thread_update_report()
        
    def thread_generate_report(self):
        if not self.thread_event.isSet():
            # print('self.xl_data: ', self.xl_data)
            self.img_txt_popout_close_all()

            self.thread_handle = threading.Thread(target=self.generate_report_func, daemon = True)
            self.thread_handle.start()
            # self.img_arr = None

    def thread_update_report(self):
        if not self.thread_event.isSet():
            self.img_txt_popout_close_all()

            self.thread_handle = threading.Thread(target=self.update_report_func, daemon = True)
            self.thread_handle.start()
            # self.img_arr = None

    def load_worksheet_name(self):
        if self.worksheet_name_var.get() == '':
            # self.worksheet_name_var.set(self.xl_worksheet_name)
            self.worksheet_name_entry.set_text(self.xl_worksheet_name, reset_undo_stack = False)
        else:
            self.xl_worksheet_name = self.worksheet_name_var.get()

    def rename_worksheet_func(self):
        self.thread_event.set()

        open_err_ret = False
        save_err_ret = False

        err_msg = None

        widget_disable(self.update_report_btn, self.generate_report_btn, self.update_report_checkbtn, self.rename_worksheet_btn, self.create_worksheet_checkbtn)

        _file_name = (re.findall(r'[^\\/]+|[\\/]', self.xl_load_path))[-1]
        # print(file, _file_name)
        save_folder = re.sub(_file_name, '', self.xl_load_path)
        # print(save_folder)

        try: 
            os.rename(save_folder + ntpath.basename(self.xl_load_path), save_folder + 'rename_tempfile.xlsx')
            os.rename(save_folder + 'rename_tempfile.xlsx', save_folder + ntpath.basename(self.xl_load_path))
        except OSError:
            open_err_ret = True

        if open_err_ret == True:
            Error_Msgbox(message = "Rename Error!\n\n"
                + "Report File is Currently Opened in Microsoft Excel, please close Microsoft Excel and try again.", title = 'Error'
                    , message_anchor = 'w', width = 370)
            self.event_log_custom_error(ntpath.basename(self.xl_load_path), 'Report File is Currently Opened in Microsoft Excel, please close Microsoft Excel and try again.', error_type = 'Rename ERROR')

        elif open_err_ret == False:
            self.xl_class.xl_open_workbook(self.xl_load_path, self.xl_worksheet_name)
            prev_worksheet_name = copy.copy(self.xl_worksheet_name)

            self.load_worksheet_name()
            _, self.xl_worksheet_name = self.xl_class.xl_rename_worksheet(self.xl_worksheet_name, override = True)
            self.xl_class.xl_select_worksheet(self.xl_worksheet_name)
            self.update_worksheet_dropbox()
            close_err_ret = self.xl_class.xl_close_workbook(self.xl_load_path)
            # print(self.xl_class.worksheet.title)

            if close_err_ret == True:
                self.xl_worksheet_name = prev_worksheet_name
                self.xl_class.xl_select_worksheet(self.xl_worksheet_name)
                self.update_worksheet_dropbox()

                Error_Msgbox(message = "Rename Error!\nPossible Problem(s):\n\n"
                    + "1) If Report File is Currently Opened in Microsoft Excel, please close Microsoft Excel and try again.\n\n"
                    + "2) If Rename Error still occur, report bug to TMS Technical Support.", title = 'Error'
                        , message_anchor = 'w', width = 370, height = 220)   

                self.event_log_custom_error(ntpath.basename(self.xl_load_path), 'Please Ensure The File Is Currently NOT Opened! If Error still occur, please report bug.', error_type = 'Rename ERROR')

            elif close_err_ret == False:
                _file_name = str((re.findall(r'[^\\/]+|[\\/]', self.xl_load_path))[-1])
                log_msg = "Worksheet: " + "'{}'".format(prev_worksheet_name) + " in " + _file_name + " is changed to " + "'{}'.".format(self.xl_worksheet_name) 
                self.event_log_custom_msg(msg_str = "Rename Success! " + log_msg)

                Info_Msgbox(message = "Worksheet Rename Success!", title = 'Info'
                        , message_anchor = 'w')

                del _file_name, log_msg

        widget_enable(self.update_report_btn, self.generate_report_btn, self.update_report_checkbtn, self.rename_worksheet_btn, self.create_worksheet_checkbtn)
        self.thread_event.clear()

        self.curr_worksheet_tk_var.set('Current Worksheet: ' + self.xl_worksheet_name)


    def update_worksheet_dropbox(self):
        self.worksheet_name_dropbox['value'] = self.xl_class.workbook.sheetnames
        self.worksheet_name_dropbox.current(self.xl_class.workbook.sheetnames.index(self.xl_class.worksheet.title))

    def worksheet_name_select(self, event):
        if self.xl_worksheet_name == self.worksheet_name_dropbox.get():
            # print("Same Worksheet..")
            pass
        else:
            if self.__edit_bool == True:
                ask_msgbox = Ask_Msgbox(message = "Current Edit/Changes on UI will be replaced if you Change Worksheet!"
                    + "\n\nAre you sure?"
                , parent = self.master, title = 'Warning', mode = 'warning', message_anchor = 'w', width = 350, ask_OK = False)
                if ask_msgbox.ask_result() == True:
                    self.xl_class.xl_select_worksheet(self.worksheet_name_dropbox.get())
                    self.xl_worksheet_name = self.xl_class.worksheet.title
                    # self.worksheet_name_var.set(self.xl_worksheet_name)
                    self.worksheet_name_entry.set_text(self.xl_worksheet_name, reset_undo_stack = False)

                    for edit_data in self.edit_data_dict.values():
                        if type(edit_data[2]) == list:
                            self.clear_ref_edit_list(edit_data[2])
                        elif (isinstance(edit_data[2], np.ndarray)) == True:
                            self.clear_ref_edit_arr(edit_data[2])

                        self.reset_edit_bool_arr(edit_data[1])

                    self.load_worksheet_data()

                    self.clear_all_upload_log()

                    self.load_worksheet_image()

                    self.reset_track_edit_bool_arr()

                    self.curr_worksheet_tk_var.set('Current Worksheet: ' + self.xl_worksheet_name)

                else:
                    self.xl_class.xl_select_worksheet(self.xl_worksheet_name)
                    self.worksheet_name_dropbox.current(self.xl_class.workbook.sheetnames.index(self.xl_worksheet_name))

                    self.__edit_bool = np.any(self.track_edit_bool_arr)

            elif self.__edit_bool == False:
                self.xl_class.xl_select_worksheet(self.worksheet_name_dropbox.get())
                self.xl_worksheet_name = self.xl_class.worksheet.title
                # self.worksheet_name_var.set(self.xl_worksheet_name)
                self.worksheet_name_entry.set_text(self.xl_worksheet_name, reset_undo_stack = False)

                for edit_data in self.edit_data_dict.values():
                    if type(edit_data[2]) == list:
                        self.clear_ref_edit_list(edit_data[2])
                    elif (isinstance(edit_data[2], np.ndarray)) == True:
                        self.clear_ref_edit_arr(edit_data[2])

                    self.reset_edit_bool_arr(edit_data[1])

                self.load_worksheet_data()

                self.clear_all_upload_log()

                self.load_worksheet_image()

                self.reset_track_edit_bool_arr()

                self.curr_worksheet_tk_var.set('Current Worksheet: ' + self.xl_worksheet_name)


    def insert_data_list(self, data_list, str_data):
        if str_data == '':
            data_list.append('')
        else:
            check_data = (str_data).split()
            if len(check_data) > 0:
                del check_data
                data_list.append(str_data)
            else:
                del check_data
                data_list.append('')

        return data_list

    def insert_imtext_data(self, imtext_cell_list, log_dict, merge_end_cell_id):
        ''' log_dict: upload log dict which has the list of all the tk_widget(s) from upload callback. tk_widget_list[7] is the tk.StringVar()
            imtext_cell_list: the list acquired from xl_class.cell_insert_data(). the list contain the cell id locations where imtext can be inserted.
            merge_end_cell_id: when we insert imtext, we must merge all the column(s) where the imtext resides. so, we provide the end column id for merge process.
        '''
        if type(imtext_cell_list) == list and len(imtext_cell_list) > 0:
            imtext_data_list = []

            for tk_widget_list in log_dict.values():
                imtext_data_list.append(tk_widget_list[7].get())

            text_cell_id_list = imtext_cell_list
            for i in range(0, min(len(imtext_data_list), len(text_cell_id_list)) ):
                merge_cell_id = "{}:{}{}".format(text_cell_id_list[i], merge_end_cell_id, int(re.findall('(\\d+)', text_cell_id_list[i])[0]))
                self.xl_class.worksheet.merge_cells(merge_cell_id)

                ### auto_fit_row is True because we want to auto-adjust row height if \n is introduced in the text.
                ### Furthermore, alignment(wrapText) is also set to True so Excel itself will set a \n line if text is too long.
                self.xl_class.cell_insert_data(self.xl_class.worksheet, imtext_data_list[i]
                    , 'string', text_cell_id_list[i], auto_fit_col = False, auto_fit_row = True
                    , alignment_dict = dict(wrapText=True, horizontal = 'right', vertical = 'center'))


    def insert_report_data(self):
        event_input_err = False

        photo_sample_data_list = []
        for tk_widget in self.photo_sample_entry_list:
            if isinstance(tk_widget, CustomSingleText):
                photo_sample_data_list = self.insert_data_list(photo_sample_data_list, tk_widget.get_text())
            elif isinstance(tk_widget, CustomBox):
                photo_sample_data_list = self.insert_data_list(photo_sample_data_list, tk_widget.get())

        test_criteria_data_list = []
        for tk_widget in self.test_criteria_entry_list:
            test_criteria_data_list = self.insert_data_list(test_criteria_data_list, tk_widget.get_text())

        cam_detail_data_list = []
        for tk_widget in self.cam_detail_entry_list:
            cam_detail_data_list = self.insert_data_list(cam_detail_data_list, tk_widget.get_text())

        lens_detail_data_list = []
        for tk_widget in self.lens_detail_entry_list:
            lens_detail_data_list = self.insert_data_list(lens_detail_data_list, tk_widget.get_text())

        ctrl_detail_data_list = []
        for tk_widget in self.ctrl_detail_entry_list:
            ctrl_detail_data_list = self.insert_data_list(ctrl_detail_data_list, tk_widget.get_text())

        light_detail_data_list = []
        for tk_widget in self.light_detail_entry_list:
            light_detail_data_list = self.insert_data_list(light_detail_data_list, tk_widget.get_text())


        if ('' in test_criteria_data_list) or ('' in cam_detail_data_list) or ('' in lens_detail_data_list):
            event_input_err = True # True # False

        if event_input_err == False:
            self.xl_class.xl_image_clear_all() ### Clear all the image(s) to re-sort and place the image(s) again
            ws_max_row = self.xl_class.worksheet.max_row
            max_row_arr = np.zeros(6, dtype=np.uint8) ### To compute the new max row, to adjust the bottom borders
            imtext_cell_list_arr = np.empty(6, dtype=object)

            xl_unmerge_cell(self.xl_class.worksheet, 'A', 5, 'F', ws_max_row
                , del_value = True
                , border_type = 'left-right')

            xl_unmerge_cell(self.xl_class.worksheet, 'G', 3, 'L', ws_max_row
                , del_value = True
                , border_type = 'left-right')

            xl_unmerge_cell(self.xl_class.worksheet, 'Q', 3, 'V', ws_max_row
                , del_value = True
                , border_type = 'left-right')

            xl_unmerge_cell(self.xl_class.worksheet, 'W', 3, 'AB', ws_max_row
                , del_value = True
                , border_type = 'left-right')

            xl_unmerge_cell(self.xl_class.worksheet, 'AC', 3, 'AH', ws_max_row
                , del_value = True
                , border_type = 'left-right')

            xl_unmerge_cell(self.xl_class.worksheet, 'AI', 3, 'AN', ws_max_row
                , del_value = True
                , border_type = 'left-right')

            #### inserting in cell B3, we prevent auto-fit because B3 is a merge-cell (more than enough space). Auto-fitting in B3 will cause problems in image(s) column A-F
            i = 3
            for str_data in photo_sample_data_list:
                self.xl_class.cell_insert_data(self.xl_class.worksheet, str_data
                    , 'string', 'B' + str(i), auto_fit_col = False, auto_fit_row = False
                    , alignment_dict = dict(wrapText=True, horizontal = 'left', vertical = 'center'))
                i += 1

            new_max_row, imtext_cell_list_arr[0] = self.xl_class.cell_insert_img(self.xl_class.worksheet, self.photo_upload_img_dict, 6, 'A', 'F', check_col_char = 'A', check_row_num = 5
                , get_text_cell_id = True)
            max_row_arr[0] = new_max_row
            self.insert_imtext_data(imtext_cell_list = imtext_cell_list_arr[0], log_dict = self.photo_upload_log_dict,  merge_end_cell_id = 'F')

            i = 5
            for entry_data in test_criteria_data_list:
                self.xl_class.cell_insert_data(self.xl_class.worksheet, entry_data
                    , 'string', 'O' + str(i) 
                    , alignment_dict = dict(wrapText=True, horizontal = 'right', vertical = 'center'))
                i += 1

            i = 22
            for entry_data in cam_detail_data_list:
                self.xl_class.cell_insert_data(self.xl_class.worksheet, entry_data
                    , 'string', 'N' + str(i) 
                    , alignment_dict = dict(wrapText=True, horizontal = 'right', vertical = 'center'))
                i += 1

            i = 33
            for entry_data in lens_detail_data_list:
                self.xl_class.cell_insert_data(self.xl_class.worksheet, entry_data
                    , 'string', 'N' + str(i) 
                    , alignment_dict = dict(wrapText=True, horizontal = 'right', vertical = 'center'))
                i += 1

            i = 17
            for entry_data in ctrl_detail_data_list:
                self.xl_class.cell_insert_data(self.xl_class.worksheet, entry_data
                    , 'string', 'N' + str(i) 
                    , alignment_dict = dict(wrapText=True, horizontal = 'right', vertical = 'center'))
                i += 1

            i = 12
            for entry_data in light_detail_data_list:
                self.xl_class.cell_insert_data(self.xl_class.worksheet, entry_data
                    , 'string', 'N' + str(i) 
                    , alignment_dict = dict(wrapText=True, horizontal = 'right', vertical = 'center'))
                i += 1

            del i

            new_max_row, imtext_cell_list_arr[1] = self.xl_class.cell_insert_img(self.xl_class.worksheet, self.light_setup_img_dict, 5, 'G', 'L', check_col_char = 'G', check_row_num = 5
                , get_text_cell_id = True)
            max_row_arr[1] = new_max_row
            self.insert_imtext_data(imtext_cell_list = imtext_cell_list_arr[1], log_dict = self.light_setup_log_dict,   merge_end_cell_id = 'L')

            new_max_row, imtext_cell_list_arr[2] = self.xl_class.cell_insert_img(self.xl_class.worksheet, self.target_upload_img_dict, 4, 'Q', 'V', check_col_char = 'Q', check_row_num = 4
                , get_text_cell_id = True)
            max_row_arr[2] = new_max_row
            self.insert_imtext_data(imtext_cell_list = imtext_cell_list_arr[2], log_dict = self.target_upload_log_dict, merge_end_cell_id = 'V')

            new_max_row, imtext_cell_list_arr[3] = self.xl_class.cell_insert_img(self.xl_class.worksheet, self.gray_upload_img_dict, 4, 'W', 'AB', check_col_char = 'W', check_row_num = 4
                , get_text_cell_id = True)
            max_row_arr[3] = new_max_row
            self.insert_imtext_data(imtext_cell_list = imtext_cell_list_arr[3], log_dict = self.gray_upload_log_dict,   merge_end_cell_id = 'AB')

            new_max_row, imtext_cell_list_arr[4] = self.xl_class.cell_insert_img(self.xl_class.worksheet, self.binary_upload_img_dict, 4, 'AC', 'AH', check_col_char = 'AC', check_row_num = 4
                , get_text_cell_id = True)
            max_row_arr[4] = new_max_row
            self.insert_imtext_data(imtext_cell_list = imtext_cell_list_arr[4], log_dict = self.binary_upload_log_dict, merge_end_cell_id = 'AH')
            
            new_max_row, imtext_cell_list_arr[5] = self.xl_class.cell_insert_img(self.xl_class.worksheet, self.setup_upload_img_dict, 4, 'AI', 'AN', check_col_char = 'AI', check_row_num = 4
                , get_text_cell_id = True)
            max_row_arr[5] = new_max_row
            self.insert_imtext_data(imtext_cell_list = imtext_cell_list_arr[5], log_dict = self.setup_upload_log_dict,  merge_end_cell_id = 'AN')

            # print(imtext_cell_list_arr)
            # group_print(ws_max_row, max(max_row_arr))

            del imtext_cell_list_arr

            ##### Update the Excel Borders
            # group_print(ws_max_row, max_row_arr)
            if ws_max_row >= 39:
                set_outer_border(ws = self.xl_class.worksheet, min_row = 39, max_row = ws_max_row, min_col = 'A', max_col='AN'
                    , border_type = None)

            if max(max_row_arr) > 39:
                set_outer_border(ws = self.xl_class.worksheet, min_row = 39, max_row = max(max_row_arr), min_col = 'A', max_col='AN'
                    , border_style = 'medium'
                    , col_separator = ['F', 'L', 'P', 'V', 'AB', 'AH'], border_type = 'bottom')

            else:
                set_outer_border(ws = self.xl_class.worksheet, min_row = 39, min_col = 'A', max_col='AN'
                    , border_style = 'medium'
                    , col_separator = ['F', 'L', 'P', 'V', 'AB', 'AH'], border_type = 'bottom')

            del max_row_arr


        report_data_list = [test_criteria_data_list, cam_detail_data_list, lens_detail_data_list, ctrl_detail_data_list, light_detail_data_list]

        del test_criteria_data_list
        del cam_detail_data_list
        del lens_detail_data_list
        del ctrl_detail_data_list
        del light_detail_data_list

        return event_input_err, report_data_list

    def generate_report_func(self):
        # print('xl_data: ', xl_data)
        self.loading_disp_start()
        self.thread_event.set()

        self.err_msgbox_flag_1 = False ### Error Flag: If user left out important report details

        event_error = False
        self.err_msgbox_flag_2 = False ### Error Flag: If user attempt to generate/update report when report is opened in Excel.

        self.err_msgbox_flag_3 = False ### Exception Flag: When Code/Algorithm have an error

        event_reset = False

        prev_worksheet_name = None

        self.report_error_msgbox_1()
        self.report_error_msgbox_2()
        self.report_error_msgbox_3()
        self.report_info_msgbox()

        widget_disable(self.update_report_btn, self.generate_report_btn, self.update_report_checkbtn
            , self.worksheet_name_dropbox, self.worksheet_name_entry
            , self.rename_worksheet_btn, self.create_worksheet_checkbtn
            , self.reload_worksheet_btn)

        if self.xl_load_path is not None and self.update_report_bool.get() == 1:
            try:
                if self.create_worksheet_bool.get() == 0:
                    self.xl_class.xl_update_workbook(self.xl_load_path, sheet_name = self.xl_worksheet_name)

                elif self.create_worksheet_bool.get() == 1:
                    self.load_worksheet_name()
                    prev_worksheet_name = copy.copy(self.xl_worksheet_name)

                    self.xl_worksheet_name = self.xl_class.xl_update_workbook(self.xl_load_path
                        , sheet_name = self.xl_worksheet_name, new_ws_bool = True)

            except Exception:
                self.load_worksheet_name()
                self.xl_worksheet_name = self.xl_class.xl_new_workbook(self.xl_worksheet_name)
                self.xl_class.xl_init_workbook(init_row_num = 2, init_col_char = 'A')
        else:
            self.load_worksheet_name()
            self.xl_worksheet_name = self.xl_class.xl_new_workbook(self.xl_worksheet_name)
            self.xl_class.xl_init_workbook(init_row_num = 2, init_col_char = 'A')

        try:
            event_input_err, report_data_list = self.insert_report_data()
        except Exception as e:
            self.thread_event.clear()
            self.__exception_msg = str(e)
            self.err_msgbox_flag_3 = True
            self.clear_info_msgbox()
            self.clear_error_msgbox_1()
            self.clear_error_msgbox_2()
            self.loading_disp_stop()
            self.worksheet_name_dropbox['state'] = 'readonly'

            widget_enable(self.update_report_btn, self.generate_report_btn, self.update_report_checkbtn
                , self.worksheet_name_entry
                , self.rename_worksheet_btn, self.create_worksheet_checkbtn
                , self.reload_worksheet_btn)
            self.xl_class.xl_close_workbook()
            if self.xl_load_path is not None:
                self.xl_class.xl_open_workbook(self.xl_load_path)
            
            return
        
        self.clear_error_msgbox_3()
        if self.xl_load_path is not None and self.update_report_bool.get() == 1:
            # print("Updating...")
            if event_input_err == True:
                self.clear_info_msgbox()
                self.clear_error_msgbox_2()

                if self.create_worksheet_bool.get() == 1:
                    self.xl_class.xl_remove_worksheet(self.xl_worksheet_name)
                    _, _ = self.xl_class.xl_save_workbook(file_path = self.xl_load_path)

                    if type(prev_worksheet_name) == str:
                        self.xl_class.xl_select_worksheet(prev_worksheet_name)
                        self.xl_worksheet_name = prev_worksheet_name

                self.err_msgbox_flag_1 = True
                

            elif event_input_err == False:
                self.clear_error_msgbox_1()
                _, event_error = self.xl_class.xl_save_workbook(file_path = self.xl_load_path)
                if event_error == False:
                    self.clear_error_msgbox_2()
                    self.xl_class.xl_close_workbook()

                    self.update_edit_entry_data("Sample Detail", [self.photo_dim_var.get()])
                    self.update_edit_entry_data("Testing Criteria", report_data_list[0])
                    self.update_edit_entry_data("Camera Details", report_data_list[1])
                    self.update_edit_entry_data("Lens Details", report_data_list[2])
                    self.update_edit_entry_data("Controller Details", report_data_list[3])
                    self.update_edit_entry_data("Lighting Details", report_data_list[4])

                    self.edit_data_dict["Sample Image"][2] = list(self.photo_upload_img_dict.values())
                    self.edit_data_dict["Lighting Drawing"][2] = list(self.light_setup_img_dict.values())
                    self.edit_data_dict["Target Image"][2] = list(self.target_upload_img_dict.values())
                    self.edit_data_dict["Grayscale Image"][2] = list(self.gray_upload_img_dict.values())
                    self.edit_data_dict["Binary Image"][2] = list(self.binary_upload_img_dict.values())
                    self.edit_data_dict["Setup Image"][2] = list(self.setup_upload_img_dict.values())

                    for edit_data in self.edit_data_dict.values():
                        self.reset_edit_bool_arr(edit_data[1])

                    self.reset_track_edit_bool_arr()

                    self.info_msgbox_flag = True
                    self.event_log_update_report()

                elif event_error == True:
                    self.clear_info_msgbox()
                    self.err_msgbox_flag_2 = True
                    self.event_log_error_report()

            self.curr_report_file_tk_var.set('Current Report File: ' + str((re.findall(r'[^\\/]+|[\\/]', self.xl_load_path))[-1]) )
            # self.worksheet_name_var.set(self.xl_worksheet_name)
            self.worksheet_name_entry.set_text(self.xl_worksheet_name, reset_undo_stack = False)
            self.worksheet_name_entry.xview_moveto('1')
            self.update_worksheet_dropbox()
            self.curr_worksheet_tk_var.set('Current Worksheet: ' + self.xl_worksheet_name)
                
        else:
            if event_input_err == True:
                event_reset = True
                self.reset_init_state() #self.xl_load_path is None

                self.err_msgbox_flag_1 = True
                self.clear_info_msgbox()
                self.clear_error_msgbox_2()

            elif event_input_err == False:
                self.clear_error_msgbox_1()

                save_path, event_error = self.xl_class.xl_save_workbook()
                self.xl_load_path = save_path
                if event_error == False:
                    self.clear_error_msgbox_2()
                    self.xl_class.xl_close_workbook()
                    
                    self.update_edit_entry_data("Sample Detail", [self.photo_dim_var.get()])
                    self.update_edit_entry_data("Testing Criteria", report_data_list[0])
                    self.update_edit_entry_data("Camera Details", report_data_list[1])
                    self.update_edit_entry_data("Lens Details", report_data_list[2])
                    self.update_edit_entry_data("Controller Details", report_data_list[3])
                    self.update_edit_entry_data("Lighting Details", report_data_list[4])

                    self.edit_data_dict["Sample Image"][2] = list(self.photo_upload_img_dict.values())
                    self.edit_data_dict["Lighting Drawing"][2] = list(self.light_setup_img_dict.values())
                    self.edit_data_dict["Target Image"][2] = list(self.target_upload_img_dict.values())
                    self.edit_data_dict["Grayscale Image"][2] = list(self.gray_upload_img_dict.values())
                    self.edit_data_dict["Binary Image"][2] = list(self.binary_upload_img_dict.values())
                    self.edit_data_dict["Setup Image"][2] = list(self.setup_upload_img_dict.values())

                    for edit_data in self.edit_data_dict.values():
                        self.reset_edit_bool_arr(edit_data[1])

                    self.reset_track_edit_bool_arr()

                    self.info_msgbox_flag = True
                    self.event_log_new_report()

                    self.curr_report_file_tk_var.set('Current Report File: ' + str((re.findall(r'[^\\/]+|[\\/]', self.xl_load_path))[-1]) )
                    # self.worksheet_name_var.set(self.xl_worksheet_name)
                    self.worksheet_name_entry.set_text(self.xl_worksheet_name, reset_undo_stack = False)
                    self.worksheet_name_entry.xview_moveto('1')
                    self.update_worksheet_dropbox()
                    self.curr_worksheet_tk_var.set('Current Worksheet: ' + self.xl_worksheet_name)

                elif event_error == True:
                    self.clear_info_msgbox()
                    os.remove(self.xl_load_path) #This else statement is true when a NEW Report is generated... Therefore, we delete the file.
                    event_reset = True
                    self.reset_init_state() #self.xl_load_path is None

                    self.err_msgbox_flag_2 = True
                    self.event_log_error_report()
                    

        self.thread_event.clear()

        del report_data_list

        if event_reset == False:
            self.worksheet_name_dropbox['state'] = 'readonly'

            widget_enable(self.update_report_btn, self.generate_report_btn, self.update_report_checkbtn
                , self.worksheet_name_entry
                , self.rename_worksheet_btn, self.create_worksheet_checkbtn
                , self.reload_worksheet_btn)

        self.loading_disp_stop()

    def update_report_func(self):
        if self.xl_load_bool == True:
            self.loading_disp_start()
            self.thread_event.set()

            self.err_msgbox_flag_1 = False ### Error Flag: If user left out important report details

            event_error = False
            self.err_msgbox_flag_2 = False ### Error Flag: If user attempt to generate/update report when report is opened in Excel.

            self.err_msgbox_flag_3 = False ### Exception Flag: When Code/Algorithm have an error

            prev_worksheet_name = None

            self.report_info_msgbox()
            self.report_error_msgbox_1()
            self.report_error_msgbox_2()
            self.report_error_msgbox_3()

            widget_disable(self.update_report_btn, self.generate_report_btn, self.update_report_checkbtn
                , self.worksheet_name_dropbox, self.worksheet_name_entry
                , self.rename_worksheet_btn, self.create_worksheet_checkbtn
                , self.reload_worksheet_btn)                

            if self.create_worksheet_bool.get() == 0:
                self.xl_class.xl_update_workbook(self.xl_load_path, sheet_name = self.xl_worksheet_name)

            elif self.create_worksheet_bool.get() == 1:
                self.load_worksheet_name()
                prev_worksheet_name = copy.copy(self.xl_worksheet_name)

                self.xl_worksheet_name = self.xl_class.xl_update_workbook(self.xl_load_path
                    , sheet_name = self.xl_worksheet_name, new_ws_bool = True)

            # print('Worksheet: ', self.xl_class.worksheet.title)

            try:
                event_input_err, report_data_list = self.insert_report_data()
            except Exception as e:
                self.thread_event.clear()
                self.__exception_msg = str(e)
                self.err_msgbox_flag_3 = True
                self.clear_info_msgbox()
                self.clear_error_msgbox_1()
                self.clear_error_msgbox_2()
                self.loading_disp_stop()
                self.worksheet_name_dropbox['state'] = 'readonly'

                widget_enable(self.update_report_btn, self.generate_report_btn, self.update_report_checkbtn
                    , self.worksheet_name_entry
                    , self.rename_worksheet_btn, self.create_worksheet_checkbtn
                    , self.reload_worksheet_btn)
                self.xl_class.xl_close_workbook()
                if self.xl_load_path is not None:
                    self.xl_class.xl_open_workbook(self.xl_load_path)
                return
            
            self.clear_error_msgbox_3()

            if event_input_err == True:
                if self.create_worksheet_bool.get() == 1:
                    self.xl_class.xl_remove_worksheet(self.xl_worksheet_name)
                    _, _ = self.xl_class.xl_save_workbook(file_path = self.xl_load_path)

                    if type(prev_worksheet_name) == str:
                        self.xl_class.xl_select_worksheet(prev_worksheet_name)
                        self.xl_worksheet_name = prev_worksheet_name

                self.err_msgbox_flag_1 = True
                self.clear_info_msgbox()
                self.clear_error_msgbox_2()

            elif event_input_err == False:
                self.clear_error_msgbox_1()
                _, event_error = self.xl_class.xl_save_workbook(file_path = self.xl_load_path)
                if event_error == False:
                    self.clear_error_msgbox_2()
                    self.xl_class.xl_close_workbook()

                    self.update_edit_entry_data("Sample Detail", [self.photo_dim_var.get()])
                    self.update_edit_entry_data("Testing Criteria", report_data_list[0])
                    self.update_edit_entry_data("Camera Details", report_data_list[1])
                    self.update_edit_entry_data("Lens Details", report_data_list[2])
                    self.update_edit_entry_data("Controller Details", report_data_list[3])
                    self.update_edit_entry_data("Lighting Details", report_data_list[4])

                    self.edit_data_dict["Sample Image"][2] = list(self.photo_upload_img_dict.values())
                    self.edit_data_dict["Lighting Drawing"][2] = list(self.light_setup_img_dict.values())
                    self.edit_data_dict["Target Image"][2] = list(self.target_upload_img_dict.values())
                    self.edit_data_dict["Grayscale Image"][2] = list(self.gray_upload_img_dict.values())
                    self.edit_data_dict["Binary Image"][2] = list(self.binary_upload_img_dict.values())
                    self.edit_data_dict["Setup Image"][2] = list(self.setup_upload_img_dict.values())

                    for edit_data in self.edit_data_dict.values():
                        self.reset_edit_bool_arr(edit_data[1])
                    
                    self.reset_track_edit_bool_arr()

                    self.info_msgbox_flag = True
                    self.event_log_update_report()

                elif event_error == True:
                    self.clear_info_msgbox()
                    self.err_msgbox_flag_2 = True
                    self.event_log_error_report()

            self.curr_report_file_tk_var.set('Current Report File: ' + str((re.findall(r'[^\\/]+|[\\/]', self.xl_load_path))[-1]) )
            # self.worksheet_name_var.set(self.xl_worksheet_name)
            self.worksheet_name_entry.set_text(self.xl_worksheet_name, reset_undo_stack = False)
            self.worksheet_name_entry.xview_moveto('1')
            self.update_worksheet_dropbox()
            self.curr_worksheet_tk_var.set('Current Worksheet: ' + self.xl_worksheet_name)
            # print('Update REPORT event_error: ', event_error)

            self.thread_event.clear()

            del report_data_list

            self.worksheet_name_dropbox['state'] = 'readonly'

            widget_enable(self.update_report_btn, self.generate_report_btn, self.update_report_checkbtn
                , self.worksheet_name_entry
                , self.rename_worksheet_btn, self.create_worksheet_checkbtn
                , self.reload_worksheet_btn)

        self.loading_disp_stop()

    def event_log_custom_msg(self, msg_str):
        event_str = str(datetime.now().strftime("(%d-%m-%Y, %H:%M:%S) ")) + str(msg_str) + '\n\n'
        # self.event_log_txtbox.insert_readonly(tk.END, event_str)
        self.event_log_txtbox.insert_readonly("0.0", event_str)

    def event_log_new_report(self):
        if self.xl_load_path is not None and type(self.xl_load_path) == str:
            event_str = str(datetime.now().strftime("(%d-%m-%Y, %H:%M:%S) ")) + 'New Report Generated.' +'\n'\
            + 'Saved At: ' +self.xl_load_path + '\n\n'
            # self.event_log_txtbox.insert_readonly(tk.END, event_str)
            self.event_log_txtbox.insert_readonly("0.0", event_str)

    def event_log_update_report(self):
        if self.xl_load_path is not None and type(self.xl_load_path) == str:

            event_str = str(datetime.now().strftime("(%d-%m-%Y, %H:%M:%S) ")) + 'Report UPDATED.' +'\n'\
            + 'Saved At: ' +self.xl_load_path + '\n\n'
            self.event_log_txtbox.insert_readonly("0.0", event_str)

    def event_log_load_report(self):
        if self.xl_load_path is not None and type(self.xl_load_path) == str:
            _file_name = (re.findall(r'[^\\/]+|[\\/]', self.xl_load_path))[-1]

            event_str = str(datetime.now().strftime("(%d-%m-%Y, %H:%M:%S) ")) + 'File Loaded: ' + \
            str(_file_name) + '\n\n'
            self.event_log_txtbox.insert_readonly("0.0", event_str)

    def event_log_error_report(self):
        if self.xl_load_path is not None and type(self.xl_load_path) == str:
            _file_name = (re.findall(r'[^\\/]+|[\\/]', self.xl_load_path))[-1]
            event_str = str(datetime.now().strftime("(%d-%m-%Y, %H:%M:%S) ")) + 'Report ERROR: ' + \
            str(_file_name) + '. Please Ensure The File Is Currently NOT Opened! If Error still occur, please report bug.' +'\n\n'
            self.event_log_txtbox.insert_readonly("0.0", event_str)

    def event_log_custom_error(self, file_path, error_msg, error_type = 'Report ERROR'):
        _file_name = (re.findall(r'[^\\/]+|[\\/]', file_path))[-1]
        event_str = str(datetime.now().strftime("(%d-%m-%Y, %H:%M:%S) ")) + str(error_type) + ': ' + \
        str(_file_name) + '. '+ str(error_msg) +'\n\n'
        self.event_log_txtbox.insert_readonly("0.0", event_str)

    def clear_event_log(self):
        ask_msgbox = Ask_Msgbox(message = "Are you sure?"
        , parent = self.master, title = "Clear Event Log", message_anchor = 'w')

        if ask_msgbox.ask_result() == True:
            self.event_log_txtbox.delete_txt()
        else:
            pass

    def report_error_msgbox_1(self):
        # print('Looping')
        if self.err_msgbox_flag_1 == False:
            self.err_msgbox_handle_1 = self.after(100, self.report_error_msgbox_1)

        else:
            self.clear_error_msgbox_1()
            Error_Msgbox(message = "*All details in 'Testing Criteria', 'Camera Details', & 'Lens Details' must be filled!", title = 'Error'
                    , message_anchor = 'w', width = 370)

    def clear_error_msgbox_1(self):
        if self.err_msgbox_handle_1 is not None:
            self.after_cancel(self.err_msgbox_handle_1)
            del self.err_msgbox_handle_1
            self.err_msgbox_handle_1 = None
            self.err_msgbox_flag_1 = False

    def report_error_msgbox_2(self):
        # print('Looping')
        if self.err_msgbox_flag_2 == False:
            self.err_msgbox_handle_2 = self.after(100, self.report_error_msgbox_2)

        else:
            self.clear_error_msgbox_2()
            # Error_Msgbox(message = "Save Error!\nReport File is Currently Opened in Microsoft Excel!\nPlease close Microsoft Excel and try again.", title = 'Error'
            #         , message_anchor = 'w', width = 370)
            Error_Msgbox(message = "Save Error!\nPossible Problem(s):\n\n"
                + "1) If Report File is Currently Opened in Microsoft Excel, please close Microsoft Excel and try again.\n\n"
                + "2) If Save Error still occur, report bug to TMS Technical Support.", title = 'Error'
                    , message_anchor = 'w', width = 370, height = 220)

    def clear_error_msgbox_2(self):
        if self.err_msgbox_handle_2 is not None:
            self.after_cancel(self.err_msgbox_handle_2)
            del self.err_msgbox_handle_2
            self.err_msgbox_handle_2 = None
            self.err_msgbox_flag_2 = False

    def report_error_msgbox_3(self):
        # print('Looping')
        if self.err_msgbox_flag_3 == False:
            self.err_msgbox_handle_3 = self.after(100, self.report_error_msgbox_3)

        else:
            self.clear_error_msgbox_3()
            # Error_Msgbox(message = "Save Error!\nReport File is Currently Opened in Microsoft Excel!\nPlease close Microsoft Excel and try again.", title = 'Error'
            #         , message_anchor = 'w', width = 370)
            Error_Msgbox(message = "Exception Found During Report Generation!\n\n"
                + "Error: {}".format(self.__exception_msg), title = 'Exception Error'
                    , message_anchor = 'w', width = 370, height = 200)

    def clear_error_msgbox_3(self):
        if self.err_msgbox_handle_3 is not None:
            self.after_cancel(self.err_msgbox_handle_3)
            del self.err_msgbox_handle_3
            self.err_msgbox_handle_3 = None
            self.err_msgbox_flag_3 = False


    def report_info_msgbox(self):
        # print('Looping report info msgbox')
        if self.info_msgbox_flag == False:
            self.info_msgbox_handle = self.after(100, self.report_info_msgbox)

        else:
            self.clear_info_msgbox()
            Info_Msgbox(message = "Report Generation Complete!", title = 'Info'
                , message_anchor = 'w')

    def clear_info_msgbox(self):
        if self.info_msgbox_handle is not None:
            self.after_cancel(self.info_msgbox_handle)
            del self.info_msgbox_handle
            self.info_msgbox_handle = None
            self.info_msgbox_flag = False

    def btn_blink_start(self):
        if self.btn_blink_handle is None:
            self.__btn_blink_state = True

        if not self.thread_event.isSet():
            if self.__edit_bool == True:
                if self.__btn_blink_state == True:
                    if self.btn_blink_handle is not None:
                        self.after_cancel(self.btn_blink_handle)
                    self.btn_blink_handle = self.after(800, self.btn_blink_start)
                    self.generate_report_btn['bg'] = 'gold' #'DarkGoldenrod1'
                    self.generate_report_btn['fg'] = 'black'
                    self.update_report_btn['bg'] = 'gold' #'DarkGoldenrod1'
                    self.update_report_btn['fg'] = 'black'
                    self.__btn_blink_state = False

                else:
                    if self.btn_blink_handle is not None:
                        self.after_cancel(self.btn_blink_handle)
                    self.btn_blink_handle = self.after(650, self.btn_blink_start)
                    self.generate_report_btn['bg'] = 'SystemButtonFace'
                    self.generate_report_btn['fg'] = 'black'
                    self.update_report_btn['bg'] = 'SystemButtonFace'
                    self.update_report_btn['fg'] = 'black'
                    self.__btn_blink_state = True

            else:
                if self.btn_blink_handle is not None:
                    self.after_cancel(self.btn_blink_handle)
                self.btn_blink_handle = self.after(150, self.btn_blink_start)
                self.generate_report_btn['bg'] = 'SystemButtonFace'
                self.generate_report_btn['fg'] = 'black'
                self.update_report_btn['bg'] = 'SystemButtonFace'
                self.update_report_btn['fg'] = 'black'
                self.__btn_blink_state = True

        elif self.thread_event.isSet():
            if self.btn_blink_handle is not None:
                self.after_cancel(self.btn_blink_handle)
            self.btn_blink_handle = self.after(150, self.btn_blink_start)
            self.generate_report_btn['bg'] = 'SystemButtonFace'
            self.generate_report_btn['fg'] = 'black'
            self.update_report_btn['bg'] = 'SystemButtonFace'
            self.update_report_btn['fg'] = 'black'
            self.__btn_blink_state = True

    def btn_blink_stop(self):
        if self.btn_blink_handle is not None:
            self.after_cancel(self.btn_blink_handle)
            del self.btn_blink_handle
            self.btn_blink_handle = None
            self.__btn_blink_state = False
            self.generate_report_btn['bg'] = 'SystemButtonFace'
            self.generate_report_btn['fg'] = 'black'
            self.update_report_btn['bg'] = 'SystemButtonFace'
            self.update_report_btn['fg'] = 'black'


    def loading_disp_start(self):
        curr_text = self.loading_tk_var.get()
        if curr_text == 'Processing':
            self.loading_tk_var.set('Processing.')
        elif curr_text == 'Processing.':
            self.loading_tk_var.set('Processing..')
        elif curr_text == 'Processing..':
            self.loading_tk_var.set('Processing...')
        else:
            self.loading_tk_var.set('Processing')

        self.loading_display.place(x=0, y = 0, relx = 0, rely = 0, relheight = 1, relwidth = 0.55, width = -80, height = -15, anchor = 'nw')
        self.left_main_gui.place_forget()
        self.__loading_handle = self.after(300, self.loading_disp_start)

    def loading_disp_stop(self):
        self.loading_display.place_forget()
        self.left_main_gui.place(x=0, y = 0, relx = 0, rely = 0, relheight = 1, relwidth = 0.55, width = -80, height = -15, anchor = 'nw')
        self.loading_tk_var.set('Processing')
        if self.__loading_handle is not None:
            self.after_cancel(self.__loading_handle)
